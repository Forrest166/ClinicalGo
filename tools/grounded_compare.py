import json
import threading
import traceback
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font

from common.paths import data_path, interrupt_path, project_root
from common.step2_support.extractor_core import (
    EXPORT_HEADERS,
    CheckpointStore,
    ExtractionError,
    ExtractionRow,
    TokenUsage,
    UserCancelledError,
    _http_post_json,
    estimate_tokens,
    infer_fallback_row,
    is_depression_related_indication,
    iter_pubmed_records_from_file,
    load_checkpoint_progress,
    normalize_row,
    normalize_source_index_value,
    parse_record_structured,
    resolve_gemini_model_id,
)


WINDOW_TITLE = "Clinical Extractor - Gemini Grounded Compare"
DEFAULT_OUTPUT_NAME = "clinical_extraction_output_gemini_grounded_compare.xlsx"
DEFAULT_MODEL_LABEL = "Default (Gemini 2.5 Flash)"
DEFAULT_MODEL_ID = "gemini-2.5-flash"
UNSEARCHABLE_NOTE_PREFIX = "Unable to search grounded sources:"
GROUNDING_FAILURE_NOTE_PREFIX = "Grounded search could not extract article content:"

DEFAULT_GROUNDED_PROMPT = """You are extracting structured clinical-trial evidence by using Gemini grounding with Google Search.

Search workflow:
1. First, find the exact study using the provided PMID when available.
2. If PMID is missing, use the provided journal/title/year/NCT metadata to find the exact article.
3. Prefer PubMed, publisher abstract pages, Europe PMC, or other reliable pages that expose the abstract and article keywords.
4. Base the extraction on the article abstract and article keywords when available.
5. If the exact article or abstract cannot be confirmed, return one placeholder row with Notes explaining the grounded-search limitation.

Output requirements:
- Return only JSON.
- Return a JSON object with one key: "rows".
- "rows" must be an array of objects.
- Create one row per intervention-result pair when possible.
- For multi-arm studies, do not merge all arms into one row.
- If a field is missing, use an empty string.
- Do not invent targets, indications, outcomes, or keywords.

Each row object must use exactly these keys:
- source_index
- pmid
- nct_id
- journal
- year
- indication
- population_characteristics
- target
- intervention
- intervention_type
- comparator
- result
- outcome_direction
- phase
- sample_size
- follow_up_time
- evidence_snippet
- statistical_metrics
- notes

Normalization requirements:
- target must be blank unless the source explicitly states a molecular target, receptor, pathway, biomarker mechanism, or biological system.
- intervention must be a concise canonical intervention name only.
- intervention_type must be exactly one of:
  Drug
  Device
  Biological
  Procedure
  Radiation
  Behavioral
  Genetic
  Dietary Supplement
  Combination Product
  Diagnostic Test
  Other
- outcome_direction must be exactly one of:
  Positive
  Neutral
  Negative
  Mixed or Unknown
- evidence_snippet should summarize the grounded abstract evidence concisely and may mention keywords only when the keyword information is explicitly visible in the grounded sources.
- statistical_metrics should include concise statistics such as p-values, CI, OR, RR, HR, beta, mean difference, response/remission rate.

If the article cannot be confidently identified or the abstract is unavailable, return exactly one row with source_index and known metadata filled, blank extraction fields, and Notes explaining the problem.
"""


MODEL_OPTIONS = {
    DEFAULT_MODEL_LABEL: DEFAULT_MODEL_ID,
    "Gemini 2.5 Flash": "gemini-2.5-flash",
    "Gemini 2.0 Flash": "gemini-2.0-flash",
}


def checkpoint_path_for_output(output_path: str) -> str:
    path = Path(output_path)
    return str(interrupt_path(f"{path.name}.checkpoint.jsonl"))


def _strip_code_fences(text: str) -> str:
    cleaned = (text or "").strip()
    if cleaned.startswith("```"):
        lines = cleaned.splitlines()
        if lines:
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        cleaned = "\n".join(lines).strip()
    return cleaned


def _extract_first_json_object(text: str) -> Dict[str, Any]:
    cleaned = _strip_code_fences(text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ExtractionError("Grounded model response did not contain a valid JSON object.")
    snippet = cleaned[start : end + 1]
    try:
        return json.loads(snippet)
    except json.JSONDecodeError as exc:
        raise ExtractionError(f"Grounded model returned invalid JSON: {exc}") from exc


def _note_row(parsed, note: str) -> ExtractionRow:
    row = infer_fallback_row(parsed, note=note)
    row.source_index = str(parsed.record_id)
    row.pmid = row.pmid or parsed.pmid
    row.nct_id = row.nct_id or parsed.nct_id
    row.journal = row.journal or parsed.journal_line
    row.year = row.year or parsed.year
    if not row.intervention_type:
        row.intervention_type = "Other"
    if not row.outcome_direction:
        row.outcome_direction = "Mixed or Unknown"
    return row


def _build_record_lookup(parsed) -> Tuple[str, str]:
    if parsed.pmid:
        return "PMID", f"PMID {parsed.pmid}"

    parts: List[str] = []
    if parsed.title:
        parts.append(f'Title: "{parsed.title}"')
    if parsed.journal_line:
        parts.append(f'Journal: "{parsed.journal_line}"')
    if parsed.year:
        parts.append(f"Year: {parsed.year}")
    if parsed.nct_id:
        parts.append(f"NCT ID: {parsed.nct_id}")

    if parts:
        return "Journal metadata", " | ".join(parts)
    return "", ""


def _build_grounded_prompt(parsed, prompt_template: str) -> str:
    lookup_mode, lookup_value = _build_record_lookup(parsed)
    metadata_lines = [
        f"source_index: {parsed.record_id}",
        f"lookup_mode: {lookup_mode or 'none'}",
        f"lookup_value: {lookup_value or ''}",
        f"pmid: {parsed.pmid}",
        f"nct_id: {parsed.nct_id}",
        f"journal: {parsed.journal_line}",
        f"year: {parsed.year}",
        f"title: {parsed.title}",
    ]
    return (
        f"{prompt_template.strip()}\n\n"
        "Study metadata:\n"
        + "\n".join(metadata_lines)
    )


def _build_grounded_batch_prompt(parsed_records, prompt_template: str) -> str:
    blocks: List[str] = []
    for parsed in parsed_records:
        lookup_mode, lookup_value = _build_record_lookup(parsed)
        blocks.append(
            "\n".join(
                [
                    f"### RECORD {parsed.record_id}",
                    f"source_index: {parsed.record_id}",
                    f"lookup_mode: {lookup_mode or 'none'}",
                    f"lookup_value: {lookup_value or ''}",
                    f"pmid: {parsed.pmid}",
                    f"nct_id: {parsed.nct_id}",
                    f"journal: {parsed.journal_line}",
                    f"year: {parsed.year}",
                    f"title: {parsed.title}",
                ]
            )
        )
    batch_rules = (
        "\n\nBatch rules:\n"
        "- You may search each record independently within the same request.\n"
        "- Never merge information across different source_index values.\n"
        "- Every returned row must contain the correct source_index for exactly one record.\n"
        "- If one record cannot be confidently identified or its abstract/keywords cannot be found, still return one placeholder row for that source_index with Notes explaining the issue.\n"
    )
    return f"{prompt_template.strip()}{batch_rules}\n\nStudy metadata:\n\n" + "\n\n".join(blocks)


def _grounded_api_call(
    prompt: str,
    model: str,
    api_key: str,
    timeout_seconds: int,
    retries: int,
) -> Tuple[str, TokenUsage, Dict[str, Any]]:
    url = (
        "https://generativelanguage.googleapis.com/v1beta/models/"
        f"{urllib.parse.quote(model, safe='')}:generateContent?key={urllib.parse.quote(api_key, safe='')}"
    )
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
    }
    response = _http_post_json(
        url,
        {"Content-Type": "application/json"},
        payload,
        timeout=timeout_seconds,
        retries=retries,
    )
    usage_meta = response.get("usageMetadata", {})
    candidates = response.get("candidates") or []
    if not candidates:
        raise ExtractionError(f"Grounded Gemini returned no candidates: {response}")
    parts = candidates[0].get("content", {}).get("parts", [])
    text = "".join(part.get("text", "") for part in parts if isinstance(part, dict)).strip()
    usage = TokenUsage(
        provider="Gemini Grounded Compare",
        model=model,
        prompt_tokens=int(usage_meta.get("promptTokenCount", estimate_tokens(prompt))),
        completion_tokens=int(usage_meta.get("candidatesTokenCount", estimate_tokens(text) if text else 0)),
        total_tokens=int(
            usage_meta.get(
                "totalTokenCount",
                int(usage_meta.get("promptTokenCount", estimate_tokens(prompt)))
                + int(usage_meta.get("candidatesTokenCount", estimate_tokens(text) if text else 0)),
            )
        ),
        input_tokens_estimated="promptTokenCount" not in usage_meta,
    )
    return text, usage, response


def grounded_extract_record(
    raw_record: str,
    record_id: int,
    model: str,
    api_key: str,
    timeout_seconds: int,
    retries: int,
    prompt_template: str,
) -> Tuple[List[ExtractionRow], TokenUsage]:
    parsed = parse_record_structured(raw_record, record_id)
    lookup_mode, lookup_value = _build_record_lookup(parsed)
    if not lookup_value:
        note = f"{UNSEARCHABLE_NOTE_PREFIX} missing PMID and usable journal/title/year metadata."
        return [_note_row(parsed, note)], TokenUsage(provider="Gemini Grounded Compare", model=model)

    prompt = _build_grounded_prompt(parsed, prompt_template)
    text, usage, _response = _grounded_api_call(prompt, model, api_key, timeout_seconds, retries)
    payload = _extract_first_json_object(text)
    raw_rows = payload.get("rows", [])
    if not isinstance(raw_rows, list):
        raise ExtractionError("Grounded model JSON must contain a 'rows' array.")

    rows: List[ExtractionRow] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        item.setdefault("source_index", str(record_id))
        item.setdefault("pmid", parsed.pmid)
        item.setdefault("nct_id", parsed.nct_id)
        item.setdefault("journal", parsed.journal_line)
        item.setdefault("year", parsed.year)
        rows.append(normalize_row(item))

    if not rows:
        note = f"{GROUNDING_FAILURE_NOTE_PREFIX} model returned no rows for {lookup_mode.lower()} lookup."
        rows = [_note_row(parsed, note)]

    for row in rows:
        row.source_index = normalize_source_index_value(row.source_index or str(record_id))
        row.pmid = row.pmid or parsed.pmid
        row.nct_id = row.nct_id or parsed.nct_id
        row.journal = row.journal or parsed.journal_line
        row.year = row.year or parsed.year
        if not row.notes and not any(
            [
                row.indication,
                row.population_characteristics,
                row.target,
                row.intervention,
                row.comparator,
                row.result,
                row.evidence_snippet,
            ]
        ):
            row.notes = f"{GROUNDING_FAILURE_NOTE_PREFIX} article content was not confidently extracted."

    return rows, usage


def grounded_extract_batch(
    raw_records: Sequence[str],
    record_ids: Sequence[int],
    model: str,
    api_key: str,
    timeout_seconds: int,
    retries: int,
    prompt_template: str,
) -> Tuple[List[ExtractionRow], TokenUsage]:
    parsed_records = [parse_record_structured(raw_record, record_id) for raw_record, record_id in zip(raw_records, record_ids)]
    searchable: List[Any] = []
    rows: List[ExtractionRow] = []

    for parsed in parsed_records:
        _lookup_mode, lookup_value = _build_record_lookup(parsed)
        if lookup_value:
            searchable.append(parsed)
        else:
            note = f"{UNSEARCHABLE_NOTE_PREFIX} missing PMID and usable journal/title/year metadata."
            rows.append(_note_row(parsed, note))

    if not searchable:
        return rows, TokenUsage(provider="Gemini Grounded Compare", model=model)

    prompt = _build_grounded_batch_prompt(searchable, prompt_template)
    text, usage, _response = _grounded_api_call(prompt, model, api_key, timeout_seconds, retries)
    payload = _extract_first_json_object(text)
    raw_rows = payload.get("rows", [])
    if not isinstance(raw_rows, list):
        raise ExtractionError("Grounded model JSON must contain a 'rows' array.")

    rows_by_source: Dict[int, List[ExtractionRow]] = {}
    searchable_ids = {parsed.record_id for parsed in searchable}
    parsed_by_id = {parsed.record_id: parsed for parsed in searchable}

    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        source_index = normalize_source_index_value(item.get("source_index", ""))
        if not source_index.isdigit():
            continue
        source_id = int(source_index)
        if source_id not in searchable_ids:
            continue
        parsed = parsed_by_id[source_id]
        item.setdefault("pmid", parsed.pmid)
        item.setdefault("nct_id", parsed.nct_id)
        item.setdefault("journal", parsed.journal_line)
        item.setdefault("year", parsed.year)
        row = normalize_row(item)
        row.source_index = str(source_id)
        row.pmid = row.pmid or parsed.pmid
        row.nct_id = row.nct_id or parsed.nct_id
        row.journal = row.journal or parsed.journal_line
        row.year = row.year or parsed.year
        rows_by_source.setdefault(source_id, []).append(row)

    for parsed in searchable:
        current_rows = rows_by_source.get(parsed.record_id, [])
        if not current_rows:
            note = f"{GROUNDING_FAILURE_NOTE_PREFIX} model returned no rows for this source_index in batch lookup."
            rows.append(_note_row(parsed, note))
            continue
        for row in current_rows:
            if not row.notes and not any(
                [
                    row.indication,
                    row.population_characteristics,
                    row.target,
                    row.intervention,
                    row.comparator,
                    row.result,
                    row.evidence_snippet,
                ]
            ):
                row.notes = f"{GROUNDING_FAILURE_NOTE_PREFIX} article content was not confidently extracted."
            rows.append(row)

    rows.sort(key=lambda row: (int(normalize_source_index_value(row.source_index) or "999999999"), row.intervention or ""))
    return rows, usage


def process_grounded_file(
    source_path: str,
    model: str,
    api_key: str,
    timeout_seconds: int,
    retries: int,
    batch_size: int,
    prompt_template: str,
    progress: Optional[Callable[[str], None]] = None,
    on_batch_done: Optional[Callable[[int, int, Sequence[ExtractionRow], TokenUsage, Sequence[int]], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
    skip_record_ids: Optional[Sequence[int]] = None,
) -> Tuple[int, List[TokenUsage], List[ExtractionRow]]:
    resolved_model, hint = resolve_gemini_model_id(model, api_key, timeout=min(timeout_seconds, 60))
    if progress:
        progress(f"Grounded model resolved to: {resolved_model}")
        if hint:
            progress(hint)

    skip_ids = {int(x) for x in (skip_record_ids or [])}
    all_records = list(iter_pubmed_records_from_file(source_path))
    pending_pairs = [(idx, record) for idx, record in enumerate(all_records, start=1) if idx not in skip_ids]
    effective_batch_size = max(1, int(batch_size))
    total_requests = (len(pending_pairs) + effective_batch_size - 1) // effective_batch_size

    total_rows = 0
    usage_logs: List[TokenUsage] = []
    preview_rows: List[ExtractionRow] = []
    completed_requests = 0

    for start in range(0, len(pending_pairs), effective_batch_size):
        if should_stop and should_stop():
            raise UserCancelledError("Stopped by user request.")
        chunk = pending_pairs[start : start + effective_batch_size]
        record_ids = [record_id for record_id, _raw_record in chunk]
        raw_records = [_raw_record for _record_id, _raw_record in chunk]

        try:
            rows, usage = grounded_extract_batch(
                raw_records=raw_records,
                record_ids=record_ids,
                model=resolved_model,
                api_key=api_key,
                timeout_seconds=timeout_seconds,
                retries=retries,
                prompt_template=prompt_template,
            )
        except UserCancelledError:
            raise
        except Exception as exc:
            rows = []
            note = f"{GROUNDING_FAILURE_NOTE_PREFIX} {str(exc).strip()}"
            for record_id, raw_record in chunk:
                parsed = parse_record_structured(raw_record, record_id)
                rows.append(_note_row(parsed, note))
            usage = TokenUsage(provider="Gemini Grounded Compare", model=resolved_model)
            if progress:
                progress(
                    f"Grounded batch failed for records {record_ids[0]}-{record_ids[-1]}, saved as placeholders: {exc}"
                )

        completed_requests += 1
        total_rows += len(rows)
        usage_logs.append(usage)
        if len(preview_rows) < 300:
            preview_rows.extend(rows[: 300 - len(preview_rows)])
        if progress:
            progress(
                f"Processed grounded batch {completed_requests}/{total_requests} covering records {record_ids[0]}-{record_ids[-1]}."
            )
        if on_batch_done:
            on_batch_done(completed_requests, max(1, total_requests), rows, usage, record_ids)

    return total_rows, usage_logs, preview_rows


def rebuild_grounded_excel_from_checkpoint(checkpoint_path: str, output_path: str) -> Tuple[int, int]:
    checkpoint = Path(checkpoint_path)
    if not checkpoint.exists():
        return 0, 0

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Results"
    ws_non_dep = wb.create_sheet("Non-Depression Results")
    ws_placeholder = wb.create_sheet("No-Extraction Results")
    ws_unsearchable = wb.create_sheet("Unsearchable Results")
    ws_stats = wb.create_sheet("Request Stats")
    ws_info = wb.create_sheet("Run Info")

    ws_data.append(EXPORT_HEADERS)
    ws_non_dep.append(EXPORT_HEADERS)
    ws_placeholder.append(EXPORT_HEADERS)
    ws_unsearchable.append(EXPORT_HEADERS)
    ws_stats.append(["Batch", "Provider", "Model", "Input Tokens", "Output Tokens", "Total Tokens", "Input Estimated"])
    ws_info.append(["Field", "Value"])
    for ws in (ws_data, ws_non_dep, ws_placeholder, ws_unsearchable, ws_stats, ws_info):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    row_count = 0
    usage_count = 0
    meta: Dict[str, Any] = {}
    pending_rows: List[Dict[str, Any]] = []
    with checkpoint.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if item.get("type") == "meta":
                meta = item
            elif item.get("type") == "row":
                pending_rows.append(item.get("data", {}))
                row_count += 1
            elif item.get("type") == "usage":
                data = item.get("data", {})
                ws_stats.append(
                    [
                        item.get("batch_index", ""),
                        data.get("provider", ""),
                        data.get("model", ""),
                        data.get("prompt_tokens", 0),
                        data.get("completion_tokens", 0),
                        data.get("total_tokens", 0),
                        "Yes" if data.get("input_tokens_estimated") else "No",
                    ]
                )
                usage_count += 1

    def row_sort_key(data: Dict[str, Any], order_index: int) -> Tuple[int, int]:
        record_index = normalize_source_index_value(data.get("Record Index", ""))
        if record_index.isdigit():
            return int(record_index), order_index
        return 10**12, order_index

    sorted_rows = sorted(enumerate(pending_rows), key=lambda pair: row_sort_key(pair[1], pair[0]))

    source_path = str(meta.get("source_path", "")).strip()
    metadata_by_id: Dict[int, Dict[str, str]] = {}
    if source_path and Path(source_path).exists():
        for rid, raw_record in enumerate(iter_pubmed_records_from_file(source_path), start=1):
            parsed = parse_record_structured(raw_record, rid)
            metadata_by_id[rid] = {
                "PMID": parsed.pmid,
                "NCT ID": parsed.nct_id,
                "Journal": parsed.journal_line,
                "Year": parsed.year,
            }

    dep_count = 0
    non_dep_count = 0
    placeholder_count = 0
    unsearchable_count = 0
    for _pos, data in sorted_rows:
        normalized_index = normalize_source_index_value(data.get("Record Index", ""))
        if normalized_index:
            data["Record Index"] = normalized_index
        if normalized_index.isdigit():
            info = metadata_by_id.get(int(normalized_index), {})
            for key in ("PMID", "NCT ID", "Journal", "Year"):
                if info.get(key) and not str(data.get(key, "")).strip():
                    data[key] = info[key]

        indication_text = str(data.get("Indication", "")).strip()
        notes_text = str(data.get("Notes", "")).strip()
        export_values = [data.get(header, "") for header in EXPORT_HEADERS]

        if notes_text.startswith(UNSEARCHABLE_NOTE_PREFIX):
            ws_unsearchable.append(export_values)
            unsearchable_count += 1
        elif notes_text.startswith(GROUNDING_FAILURE_NOTE_PREFIX):
            ws_placeholder.append(export_values)
            placeholder_count += 1
        elif indication_text and not is_depression_related_indication(indication_text):
            ws_non_dep.append(export_values)
            non_dep_count += 1
        else:
            ws_data.append(export_values)
            dep_count += 1

    ws_info.append(["Checkpoint", str(checkpoint)])
    ws_info.append(["Source File", meta.get("source_path", "")])
    ws_info.append(["Provider", meta.get("provider", "")])
    ws_info.append(["Model", meta.get("model", "")])
    ws_info.append(["Started At", meta.get("started_at", "")])
    ws_info.append(["Rebuilt At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_info.append(["Extracted Rows", row_count])
    ws_info.append(["Depression Sheet Rows", dep_count])
    ws_info.append(["Non-Depression Sheet Rows", non_dep_count])
    ws_info.append(["No-Extraction Sheet Rows", placeholder_count])
    ws_info.append(["Unsearchable Sheet Rows", unsearchable_count])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return row_count, usage_count


class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry("1260x840")

        self.source_path = tk.StringVar()
        self.output_path = tk.StringVar(value=str(data_path(DEFAULT_OUTPUT_NAME)))
        self.model_label = tk.StringVar(value=DEFAULT_MODEL_LABEL)
        self.custom_model = tk.StringVar()
        self.api_key = tk.StringVar()
        self.batch_size = tk.IntVar(value=3)
        self.timeout_seconds = tk.IntVar(value=180)
        self.retries = tk.IntVar(value=2)
        self.status_text = tk.StringVar(value="Ready.")
        self.plan_text = tk.StringVar(value="Records: -, Grounded requests: -, Batch size: -")
        self.rows_text = tk.StringVar(value="Extracted rows: 0")
        self.token_text = tk.StringVar(value="Input: 0 | Output: 0 | Total: 0")
        self.progress_text = tk.StringVar(value="Progress: 0/0")
        self.checkpoint_text = tk.StringVar(value="Checkpoint: -")
        self.model_hint_text = tk.StringVar(
            value="Default maps to Gemini 2.5 Flash. Official Google Search grounding also supports Gemini 2.0 Flash."
        )

        self.preview_rows: List[ExtractionRow] = []
        self.total_rows = 0
        self.total_prompt_tokens = 0
        self.total_completion_tokens = 0
        self.total_tokens = 0
        self.is_running = False
        self.stop_requested = False

        self._build_ui()
        self._load_api_key_from_file()
        self._update_plan_preview()

    def _build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        outer = ttk.Frame(self.root, padding=12)
        outer.grid(sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(3, weight=1)

        files = ttk.LabelFrame(outer, text="Files", padding=10)
        files.grid(row=0, column=0, sticky="ew")
        files.columnconfigure(1, weight=1)
        ttk.Label(files, text="TXT input").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(files, textvariable=self.source_path).grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Button(files, text="Browse", command=self._browse_source).grid(row=0, column=2, padx=(8, 0), pady=4)
        ttk.Label(files, text="Excel output").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(files, textvariable=self.output_path).grid(row=1, column=1, sticky="ew", pady=4)
        ttk.Button(files, text="Save As", command=self._browse_output).grid(row=1, column=2, padx=(8, 0), pady=4)

        config = ttk.LabelFrame(outer, text="Gemini Grounded Settings", padding=10)
        config.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        for index in range(6):
            config.columnconfigure(index, weight=1)

        ttk.Label(config, text="Grounded model").grid(row=0, column=0, sticky="w")
        ttk.Combobox(
            config,
            textvariable=self.model_label,
            values=list(MODEL_OPTIONS.keys()),
            state="readonly",
        ).grid(row=1, column=0, sticky="ew", padx=(0, 8), pady=(4, 8))

        ttk.Label(config, text="Custom model").grid(row=0, column=1, sticky="w")
        ttk.Entry(config, textvariable=self.custom_model).grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=(4, 8))

        ttk.Label(config, text="API key").grid(row=0, column=2, sticky="w")
        ttk.Entry(config, textvariable=self.api_key, show="*").grid(row=1, column=2, sticky="ew", padx=(0, 8), pady=(4, 8))

        ttk.Label(config, text="Records per request").grid(row=0, column=3, sticky="w")
        ttk.Spinbox(config, from_=1, to=10, textvariable=self.batch_size, width=8).grid(row=1, column=3, sticky="w", pady=(4, 8))

        ttk.Label(config, text="Timeout (sec)").grid(row=0, column=4, sticky="w")
        ttk.Spinbox(config, from_=30, to=1800, textvariable=self.timeout_seconds, width=8).grid(row=1, column=4, sticky="w", pady=(4, 8))

        ttk.Label(config, text="Retries").grid(row=0, column=5, sticky="w")
        ttk.Spinbox(config, from_=0, to=8, textvariable=self.retries, width=8).grid(row=1, column=5, sticky="w", pady=(4, 8))

        ttk.Label(config, textvariable=self.model_hint_text, wraplength=980, justify="left").grid(
            row=2, column=0, columnspan=6, sticky="w"
        )

        status = ttk.LabelFrame(outer, text="Run Status", padding=10)
        status.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        status.columnconfigure(0, weight=1)
        ttk.Label(status, textvariable=self.plan_text).grid(row=0, column=0, sticky="w")
        self.progress = ttk.Progressbar(status, orient="horizontal", mode="determinate")
        self.progress.grid(row=1, column=0, sticky="ew", pady=(8, 6))
        ttk.Label(status, textvariable=self.progress_text).grid(row=2, column=0, sticky="w")
        ttk.Label(status, textvariable=self.rows_text).grid(row=3, column=0, sticky="w", pady=(4, 0))
        ttk.Label(status, textvariable=self.token_text).grid(row=4, column=0, sticky="w", pady=(4, 0))
        ttk.Label(status, textvariable=self.checkpoint_text).grid(row=5, column=0, sticky="w", pady=(4, 0))

        notebook = ttk.Notebook(outer)
        notebook.grid(row=3, column=0, sticky="nsew", pady=(10, 0))

        prompt_tab = ttk.Frame(notebook, padding=10)
        prompt_tab.columnconfigure(0, weight=1)
        prompt_tab.rowconfigure(1, weight=1)
        notebook.add(prompt_tab, text="Grounded Prompt")
        ttk.Label(
            prompt_tab,
            text="This mode asks Gemini with Google Search grounding to find the article by PMID first, then by journal metadata.",
        ).grid(row=0, column=0, sticky="w")
        self.prompt_box = tk.Text(prompt_tab, wrap="word", height=18)
        self.prompt_box.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        self.prompt_box.insert("1.0", DEFAULT_GROUNDED_PROMPT)

        preview_tab = ttk.Frame(notebook, padding=10)
        preview_tab.columnconfigure(0, weight=1)
        preview_tab.rowconfigure(0, weight=1)
        notebook.add(preview_tab, text="Preview")
        preview_columns = tuple(EXPORT_HEADERS[:10])
        self.tree = ttk.Treeview(preview_tab, columns=preview_columns, show="headings", height=18)
        for column in preview_columns:
            self.tree.heading(column, text=column)
            self.tree.column(column, width=120, anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")

        log_tab = ttk.Frame(notebook, padding=10)
        log_tab.columnconfigure(0, weight=1)
        log_tab.rowconfigure(0, weight=1)
        notebook.add(log_tab, text="Log")
        self.log_box = tk.Text(log_tab, wrap="word")
        self.log_box.grid(row=0, column=0, sticky="nsew")

        actions = ttk.Frame(outer)
        actions.grid(row=4, column=0, sticky="ew", pady=(10, 0))
        ttk.Button(actions, text="Preview Request Plan", command=self._update_plan_preview).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="Run Grounded Compare", command=self._run).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="Continue", command=self._continue_run).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(actions, text="Stop & Build Excel", command=self._request_stop).grid(row=0, column=3, padx=(0, 8))
        ttk.Label(actions, textvariable=self.status_text).grid(row=0, column=4, sticky="w")

        self.source_path.trace_add("write", lambda *_args: self._update_plan_preview())
        self.batch_size.trace_add("write", lambda *_args: self._update_plan_preview())
        self.output_path.trace_add("write", lambda *_args: self._update_checkpoint_label())
        self._update_checkpoint_label()

    def _selected_model(self) -> str:
        return self.custom_model.get().strip() or MODEL_OPTIONS.get(self.model_label.get(), DEFAULT_MODEL_ID)

    def _browse_source(self) -> None:
        selected = filedialog.askopenfilename(title="Choose input TXT", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if selected:
            self.source_path.set(selected)

    def _browse_output(self) -> None:
        selected = filedialog.asksaveasfilename(
            title="Choose output Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=Path(self.output_path.get()).name,
        )
        if selected:
            self.output_path.set(selected)

    def _load_api_key_from_file(self) -> None:
        candidate_paths = [
            Path.cwd() / "api_keys.local.json",
            Path.cwd() / "api_keys.template.json",
            project_root() / "api_keys.local.json",
            project_root() / "api_keys.template.json",
        ]
        config_path = next((p for p in candidate_paths if p.exists()), None)
        if config_path is None:
            return
        try:
            payload = json.loads(config_path.read_text(encoding="utf-8"))
        except Exception as exc:
            self.status_text.set(f"Failed to read {config_path.name}: {exc}")
            return
        value = str(payload.get("gemini_api_key", "")).strip()
        if value and not value.startswith("PASTE_"):
            self.api_key.set(value)
            self.status_text.set(f"Loaded Gemini API key from {config_path.name}")

    def _append_log(self, message: str) -> None:
        self.log_box.insert("end", message + "\n")
        self.log_box.see("end")
        self.status_text.set(message)
        self.root.update_idletasks()

    def _update_checkpoint_label(self) -> None:
        self.checkpoint_text.set(f"Checkpoint: {checkpoint_path_for_output(self.output_path.get().strip() or DEFAULT_OUTPUT_NAME)}")

    def _update_plan_preview(self) -> None:
        source = self.source_path.get().strip()
        if not source or not Path(source).exists():
            self.plan_text.set("Records: -, Grounded requests: -, Batch size: -")
            return
        try:
            record_count = sum(1 for _ in iter_pubmed_records_from_file(source))
            batch_size = max(1, int(self.batch_size.get()))
            request_count = (record_count + batch_size - 1) // batch_size
            self.plan_text.set(
                f"Records: {record_count}, Grounded requests: {request_count}, Batch size: {batch_size}"
            )
        except Exception as exc:
            self.plan_text.set(f"Plan preview failed: {exc}")

    def _reset_run_state(self) -> None:
        self.preview_rows = []
        self.total_rows = 0
        self.total_prompt_tokens = 0
        self.total_completion_tokens = 0
        self.total_tokens = 0
        self.progress["value"] = 0
        self.progress["maximum"] = 1
        self.progress_text.set("Progress: 0/0")
        self.rows_text.set("Extracted rows: 0")
        self.token_text.set("Input: 0 | Output: 0 | Total: 0")
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.stop_requested = False

    def _request_stop(self) -> None:
        if not self.is_running:
            self._append_log("No active run to stop.")
            return
        self.stop_requested = True
        self._append_log("Stop requested. Finishing current grounded request, then rebuilding Excel from checkpoint...")

    def _run(self) -> None:
        self._start_run(resume_only=False)

    def _continue_run(self) -> None:
        self._start_run(resume_only=True)

    def _start_run(self, resume_only: bool) -> None:
        if self.is_running:
            messagebox.showinfo("Running", "Grounded comparison is already in progress.")
            return
        source = self.source_path.get().strip()
        if not source:
            messagebox.showerror("Missing file", "Please choose a TXT input file.")
            return
        if not Path(source).exists():
            messagebox.showerror("Missing file", "The selected TXT input file does not exist.")
            return
        if not self.api_key.get().strip():
            messagebox.showerror("Missing API key", "Please enter a Gemini API key.")
            return

        self.is_running = True
        self._reset_run_state()
        threading.Thread(target=self._run_worker, args=(resume_only,), daemon=True).start()

    def _batch_done_ui(self, request_index: int, planned_request_count: int, rows, usage) -> None:
        self.total_rows += len(rows)
        self.total_prompt_tokens += usage.prompt_tokens
        self.total_completion_tokens += usage.completion_tokens
        self.total_tokens += usage.total_tokens
        self.progress["maximum"] = max(1, planned_request_count)
        self.progress["value"] = min(request_index, planned_request_count)
        overflow = "+" if request_index > planned_request_count else ""
        self.progress_text.set(f"Progress: {request_index}/{planned_request_count}{overflow}")
        self.rows_text.set(f"Extracted rows: {self.total_rows}")
        self.token_text.set(
            f"Input: {self.total_prompt_tokens} | Output: {self.total_completion_tokens} | Total: {self.total_tokens}"
        )
        remaining = 300 - len(self.preview_rows)
        if remaining > 0:
            self.preview_rows.extend(rows[:remaining])
            self._refresh_preview()

    def _run_worker(self, resume_only: bool = False) -> None:
        checkpoint = None
        checkpoint_path = checkpoint_path_for_output(self.output_path.get().strip())
        output_path = self.output_path.get().strip()
        try:
            source_path = self.source_path.get().strip()
            provider = "Gemini Grounded Compare"
            model = self._selected_model()

            checkpoint = CheckpointStore(checkpoint_path)
            checkpoint_meta = checkpoint.read_meta() if Path(checkpoint_path).exists() else {}
            resume_ids = set()
            if checkpoint.has_progress():
                same_source = checkpoint_meta.get("source_path", "") == source_path
                same_provider = checkpoint_meta.get("provider", "") == provider
                same_model = str(checkpoint_meta.get("model", "")).strip() == model.strip()
                if same_source and same_provider and same_model:
                    resume_ids, existing_rows, existing_usages = load_checkpoint_progress(checkpoint_path)
                    self.root.after(
                        0,
                        self._append_log,
                        f"Resume mode enabled from checkpoint: {existing_rows} rows, {existing_usages} requests already completed.",
                    )
                else:
                    if resume_only:
                        raise ExtractionError(
                            "Continue failed: checkpoint does not match current source/provider/model. "
                            "Keep the same settings or use Run Grounded Compare to start fresh."
                        )
                    checkpoint.reset(source_path, provider, model)
                    self.root.after(
                        0,
                        self._append_log,
                        "Existing checkpoint does not match current grounded-compare settings. Starting fresh checkpoint.",
                    )
            else:
                if resume_only:
                    raise ExtractionError("Continue failed: no checkpoint progress found. Please run grounded comparison first.")
                checkpoint.reset(source_path, provider, model)

            self.root.after(0, self._append_log, f"Checkpoint file: {checkpoint_path}")
            self.root.after(0, self._append_log, "Starting grounded comparison...")

            def on_batch_done(batch_index: int, batch_count: int, rows, usage, completed_record_ids) -> None:
                checkpoint.append_batch(batch_index, rows, usage, completed_record_ids)
                self.root.after(0, self._batch_done_ui, batch_index, batch_count, rows, usage)

            total_rows, _usage_logs, _preview_rows = process_grounded_file(
                source_path=source_path,
                model=model,
                api_key=self.api_key.get().strip(),
                timeout_seconds=max(30, int(self.timeout_seconds.get())),
                retries=max(0, int(self.retries.get())),
                batch_size=max(1, int(self.batch_size.get())),
                prompt_template=self.prompt_box.get("1.0", "end").strip() or DEFAULT_GROUNDED_PROMPT,
                progress=lambda message: self.root.after(0, self._append_log, message),
                on_batch_done=on_batch_done,
                should_stop=lambda: self.stop_requested,
                skip_record_ids=resume_ids,
            )
            rebuilt_rows, _usage_count = rebuild_grounded_excel_from_checkpoint(checkpoint_path, output_path)
            self.root.after(0, self._append_log, f"Grounded comparison finished with {total_rows} rows.")
            self.root.after(0, self._append_log, f"Excel rebuilt from checkpoint with {rebuilt_rows} rows.")
            self.root.after(0, lambda: self.status_text.set(f"Done. Exported to {output_path}"))
        except UserCancelledError:
            recovered_note = ""
            if checkpoint is not None and checkpoint.has_progress():
                rebuilt_rows, _usage_count = rebuild_grounded_excel_from_checkpoint(checkpoint_path, output_path)
                recovered_note = f"Stopped by user. Recovered {rebuilt_rows} rows into {output_path}."
                self.root.after(0, self._append_log, recovered_note)
                self.root.after(0, lambda: self.status_text.set(f"Stopped. Exported partial results to {output_path}"))
            else:
                self.root.after(0, self._append_log, "Stopped by user before any grounded request completed. No rows to export.")
                self.root.after(0, lambda: self.status_text.set("Stopped. No completed rows yet."))
        except ExtractionError as exc:
            recovered_note = ""
            if checkpoint is not None and checkpoint.has_progress():
                rebuilt_rows, _usage_count = rebuild_grounded_excel_from_checkpoint(checkpoint_path, output_path)
                recovered_note = f" Recovered {rebuilt_rows} rows from checkpoint into {output_path}."
                self.root.after(0, self._append_log, recovered_note.strip())
            self.root.after(0, messagebox.showerror, "Grounded comparison error", str(exc) + recovered_note)
            self.root.after(0, self._append_log, f"Grounded comparison failed: {exc}")
        except Exception as exc:
            detail = "".join(traceback.format_exception(exc))
            recovered_note = ""
            if checkpoint is not None and checkpoint.has_progress():
                rebuilt_rows, _usage_count = rebuild_grounded_excel_from_checkpoint(checkpoint_path, output_path)
                recovered_note = f" Recovered {rebuilt_rows} rows from checkpoint into {output_path}."
                self.root.after(0, self._append_log, recovered_note.strip())
            self.root.after(0, messagebox.showerror, "Unexpected error", detail + recovered_note)
            self.root.after(0, self._append_log, f"Unexpected failure: {exc}")
        finally:
            self.is_running = False

    def _refresh_preview(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.preview_rows:
            export_row = row.to_export_dict()
            self.tree.insert("", "end", values=tuple(export_row.get(header, "") for header in EXPORT_HEADERS[:10]))


def main() -> None:
    root = tk.Tk()
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
