import json
import math
import re
import threading
import time
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
import urllib.parse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterator, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from common import provider_catalog as shared_provider_catalog
from common.gemini_models import (
    GeminiModelError,
    check_gemini_model_support as shared_check_gemini_model_support,
    resolve_gemini_model_id as shared_resolve_gemini_model_id,
)
from common.request_stability import apply_runtime_guards
from common.text_utils import estimate_token_count
from common.step2_support.normalization import normalize_row_fields
from common.step2_support.parsing import (
    ParsedRecord,
    format_parsed_record_for_llm,
    iter_pubmed_records_from_file,
    parse_record_metadata,
    parse_record_structured,
)
from common.step2_support.scheduler import (
    covered_source_ids_from_rows as scheduler_covered_source_ids_from_rows,
    is_hard_request_quota_error as scheduler_is_hard_request_quota_error,
    recover_missing_rows as scheduler_recover_missing_rows,
    run_batch_adaptive as scheduler_run_batch_adaptive,
    should_split_batch as scheduler_should_split_batch,
)
from common.step2_support.transport import (
    StopRequested as TransportStopRequested,
    TransportError,
    compact_quota_message as transport_compact_quota_message,
    http_post_json as transport_http_post_json,
    is_daily_request_quota_body as transport_is_daily_request_quota_body,
    parse_retry_delay_seconds as transport_parse_retry_delay_seconds,
)
from common.step2_support.prompt_template import DEFAULT_PROMPT_TEMPLATE

EXPORT_HEADERS = [
    "Record Index",
    "PMID",
    "NCT ID",
    "Journal",
    "Year",
    "Indication",
    "Population Characteristics",
    "Population Raw",
    "Target",
    "Intervention",
    "Intervention Type",
    "Comparator",
    "Result",
    "Outcome Direction",
    "Phase",
    "Sample Size",
    "Follow-up Time",
    "Evidence Snippet",
    "Statistical Metrics",
    "Notes",
    "Population Age Type",
    "Population Age Value",
    "Population Age Descriptor",
    "Population Age Evidence Span",
    "Population Gender",
    "Population Gender Evidence Span",
    "Population Severity",
    "Population Severity Evidence Span",
    "Population Ethnicity",
    "Population Ethnicity Evidence Span",
    "Population Occupation",
    "Population Occupation Evidence Span",
    "Population Social Status",
    "Population Social Status Evidence Span",
    "Population Previous Treatment",
    "Population Previous Treatment Evidence Span",
]

RECORD_ID_HEADER = "Record ID"
RESULT_RECORD_ID_PREFIX = "S2"
FALLBACK_RECORD_ID_PREFIX = "FB"
STAGE2_FALLBACK_RECORD_ID_PREFIX = "SFB"


CTGOV_INTERVENTION_TYPES = [
    "Drug",
    "Device",
    "Biological",
    "Procedure",
    "Radiation",
    "Behavioral",
    "Genetic",
    "Dietary Supplement",
    "Combination Product",
    "Diagnostic Test",
    "Other",
]


NO_EXTRACTION_NOTE = "No extractable intervention-result pair found in this record."
NO_ABSTRACT_NOTE = "No abstract body detected in source record."


def sanitize_record_id_source_label(raw_source_path: str) -> str:
    name = Path(str(raw_source_path or "")).stem.strip() or "input"
    name = re.sub(r"[^A-Za-z0-9]+", "-", name)
    name = name.strip("-")
    if not name:
        name = "input"
    return name[:80]


def build_prefixed_record_id(prefix: str, source_label: str, serial: int) -> str:
    return f"{prefix}-{source_label}-{int(serial)}"


def parse_prefixed_record_id(record_id: Any, prefix: str) -> Optional[Tuple[str, int]]:
    text = str(record_id or "").strip()
    if not text or not text.startswith(f"{prefix}-"):
        return None
    match = re.match(rf"^{re.escape(prefix)}-(.+)-(\d+)$", text)
    if not match:
        return None
    return match.group(1), int(match.group(2))


class ExtractionError(Exception):
    pass


class MalformedModelOutputError(ExtractionError):
    pass


class UserCancelledError(ExtractionError):
    pass


def _parse_retry_delay_seconds(error_body: str) -> Optional[int]:
    return transport_parse_retry_delay_seconds(error_body)


def _compact_quota_message(status: int, body: str) -> str:
    return transport_compact_quota_message(status, body)


def _is_daily_request_quota_body(body: str) -> bool:
    return transport_is_daily_request_quota_body(body)


@dataclass
class TokenUsage:
    provider: str
    model: str
    prompt_tokens: int = 0
    completion_tokens: int = 0
    total_tokens: int = 0
    input_tokens_estimated: bool = False


@dataclass
class ExtractionRow:
    source_index: str = ""
    pmid: str = ""
    nct_id: str = ""
    journal: str = ""
    year: str = ""
    indication: str = ""
    population_characteristics: str = ""
    population_raw: str = ""
    target: str = ""
    intervention: str = ""
    intervention_type: str = ""
    comparator: str = ""
    result: str = ""
    outcome_direction: str = ""
    phase: str = ""
    sample_size: str = ""
    follow_up_time: str = ""
    evidence_snippet: str = ""
    statistical_metrics: str = ""
    notes: str = ""
    population_age_type: str = ""
    population_age_value: str = ""
    population_age_sd: str = ""
    population_age_unit: str = ""
    population_age_descriptor: str = ""
    population_age_evidence_span: str = ""
    population_gender: str = ""
    population_gender_evidence_span: str = ""
    population_severity: str = ""
    population_severity_evidence_span: str = ""
    population_ethnicity: str = ""
    population_ethnicity_evidence_span: str = ""
    population_occupation: str = ""
    population_occupation_evidence_span: str = ""
    population_social_status: str = ""
    population_social_status_evidence_span: str = ""
    population_previous_treatment: str = ""
    population_previous_treatment_evidence_span: str = ""

    def to_export_dict(self) -> Dict[str, str]:
        return {
            "Record Index": self.source_index,
            "PMID": self.pmid,
            "NCT ID": self.nct_id,
            "Journal": self.journal,
            "Year": self.year,
            "Indication": self.indication,
            "Population Characteristics": self.population_characteristics,
            "Population Raw": self.population_raw,
            "Target": self.target,
            "Intervention": self.intervention,
            "Intervention Type": self.intervention_type,
            "Comparator": self.comparator,
            "Result": self.result,
            "Outcome Direction": self.outcome_direction,
            "Phase": self.phase,
            "Sample Size": self.sample_size,
            "Follow-up Time": self.follow_up_time,
            "Evidence Snippet": self.evidence_snippet,
            "Statistical Metrics": self.statistical_metrics,
            "Notes": self.notes,
            "Population Age Type": self.population_age_type,
            "Population Age Value": self.population_age_value,
            "Population Age Descriptor": self.population_age_descriptor,
            "Population Age Evidence Span": self.population_age_evidence_span,
            "Population Gender": self.population_gender,
            "Population Gender Evidence Span": self.population_gender_evidence_span,
            "Population Severity": self.population_severity,
            "Population Severity Evidence Span": self.population_severity_evidence_span,
            "Population Ethnicity": self.population_ethnicity,
            "Population Ethnicity Evidence Span": self.population_ethnicity_evidence_span,
            "Population Occupation": self.population_occupation,
            "Population Occupation Evidence Span": self.population_occupation_evidence_span,
            "Population Social Status": self.population_social_status,
            "Population Social Status Evidence Span": self.population_social_status_evidence_span,
            "Population Previous Treatment": self.population_previous_treatment,
            "Population Previous Treatment Evidence Span": self.population_previous_treatment_evidence_span,
        }


@dataclass
class BatchResult:
    rows: List[ExtractionRow]
    usage: TokenUsage
    raw_json: str


@dataclass
class ProcessingStats:
    record_count: int
    batch_count: int
    estimated_input_tokens: int


@dataclass
class DeferredRecoveryItem:
    record_id: int
    llm_text: str
    parsed: ParsedRecord
    metadata: Dict[str, str]
    attempts: int = 0


MODEL_CATALOG = shared_provider_catalog.MODEL_CATALOG


def recommended_summary(provider: str) -> str:
    return shared_provider_catalog.recommended_summary(provider)


def get_models(provider: str) -> List[str]:
    return shared_provider_catalog.get_models(provider)


MODEL_ALIASES = shared_provider_catalog.MODEL_ALIASES


def resolve_model_name(model: str) -> str:
    return shared_provider_catalog.resolve_model_name(model)


def estimate_tokens(text: str) -> int:
    return estimate_token_count(text)


def normalize_source_index_value(raw_value: Any) -> str:
    text = "" if raw_value is None else str(raw_value).strip()
    if not text:
        return ""
    if text.isdigit():
        return text
    match = re.search(r"\b(\d+)\b", text)
    if match:
        return match.group(1)
    return text


def is_depression_related_indication(text: str) -> bool:
    lowered = str(text or "").strip().lower()
    if not lowered:
        return False
    depression_terms = [
        "depress",
        "mdd",
        "major depressive",
        "treatment-resistant depression",
        "trd",
        "postpartum depression",
        "ppd",
        "dysthymi",
        "depressive episode",
        "bipolar depression",
        "perinatal depression",
    ]
    return any(term in lowered for term in depression_terms)


def summarize_input(source_path: str, batch_size: int, prompt_template: str) -> ProcessingStats:
    record_count = 0
    text_tokens = 0
    for record in iter_pubmed_records_from_file(source_path):
        record_count += 1
        parsed = parse_record_structured(record, record_count)
        text_tokens += estimate_tokens(format_parsed_record_for_llm(parsed))
    batch_count = math.ceil(record_count / max(1, batch_size))
    estimated_input_tokens = text_tokens + batch_count * estimate_tokens(prompt_template)
    return ProcessingStats(record_count=record_count, batch_count=batch_count, estimated_input_tokens=estimated_input_tokens)


def chunk_records_iter(source_path: str, batch_size: int) -> Iterator[Tuple[List[str], List[int]]]:
    batch: List[str] = []
    ids: List[int] = []
    idx = 0
    for record in iter_pubmed_records_from_file(source_path):
        idx += 1
        batch.append(record)
        ids.append(idx)
        if len(batch) >= batch_size:
            yield batch, ids
            batch, ids = [], []
    if batch:
        yield batch, ids


def build_batch_payload(records: Sequence[str], record_ids: Sequence[int]) -> str:
    return "\n\n".join(f"### RECORD {record_id}\n{record}" for record_id, record in zip(record_ids, records))


def extract_follow_up_time_from_text(*texts: str) -> str:
    merged = " ".join(str(t or "") for t in texts)
    if not merged.strip():
        return ""

    cleaned = re.sub(r"\s+", " ", merged)
    duration = r"\d+(?:\.\d+)?(?:\s*(?:to|\-|–)\s*\d+(?:\.\d+)?)?\s*(?:-|–|\s)?(?:day|days|week|weeks|month|months|year|years)"
    patterns = [
        rf"\bfollow[- ]?up(?:\s*(?:period|duration|time))?\s*(?:of|for|at|was|is|:)?\s*({duration})\b",
        rf"\b({duration})\s*(?:of\s*)?follow[- ]?up\b",
        rf"\bfollowed\s+(?:participants|patients|subjects)?\s*(?:for)?\s*({duration})\b",
        rf"\bpost[- ]?treatment\s*(?:follow[- ]?up)?\s*(?:at|of|for)?\s*({duration})\b",
    ]

    hits: List[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, cleaned, flags=re.I):
            value = re.sub(r"\s+", " ", match.group(1)).strip(" ,;.")
            if value and value.lower() not in {h.lower() for h in hits}:
                hits.append(value)
            if len(hits) >= 2:
                break
        if len(hits) >= 2:
            break

    if not hits:
        return ""
    return "; ".join(hits)


def _clean_fallback_text(text: str) -> str:
    value = re.sub(r"[\t\r\n]+", " ", str(text or ""))
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ;|,-")


def _first_sentence(text: str, max_chars: int = 320) -> str:
    cleaned = _clean_fallback_text(text)
    if not cleaned:
        return ""
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", cleaned)
    sentence = parts[0].strip()
    if len(sentence) > max_chars:
        sentence = sentence[:max_chars].rstrip(" ,;:-")
    return sentence


def _trim_candidate_phrase(text: str) -> str:
    cleaned = _clean_fallback_text(text)
    if not cleaned:
        return ""
    cleaned = re.split(
        r"\b(?:who|which|that|where|because|while|during|after|before|when|were|was|are|is|had|received|underwent|enrolled|recruited|randomized|randomised|assigned|allocated|followed)\b",
        cleaned,
        maxsplit=1,
        flags=re.I,
    )[0]
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip(" ,;:-")


def _extract_sample_size_from_text(*texts: str) -> str:
    merged = " ".join(_clean_fallback_text(text) for text in texts if text).strip()
    if not merged:
        return ""
    patterns = [
        r"\bN\s*=\s*(\d{1,5})\b",
        r"\bn\s*=\s*(\d{1,5})\b",
        r"\b(\d{2,5})\s+(?:participants|patients|subjects|adults|women|men|children|adolescents|individuals)\b",
        r"\b(?:included|enrolled|recruited|randomi[sz]ed)\s+(\d{2,5})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, merged, flags=re.I)
        if match:
            return match.group(1)
    return ""


def _extract_phase_from_text(*texts: str) -> str:
    merged = " ".join(_clean_fallback_text(text) for text in texts if text).strip()
    if not merged:
        return ""
    match = re.search(r"\bphase\s+((?:[1-4]|i{1,3}|iv)(?:/(?:[1-4]|i{1,3}|iv))?)\b", merged, flags=re.I)
    if not match:
        return ""
    return f"Phase {match.group(1).upper()}"


def _extract_population_candidate(parsed: ParsedRecord) -> str:
    for key in ("PATIENTS", "PARTICIPANTS", "SETTING", "METHODS", "SUMMARY"):
        text = parsed.sections.get(key, "")
        if not text:
            continue
        cleaned_text = _clean_fallback_text(text)
        patterns = [
            r"\b(?:participants|patients|subjects|adults|women|men|children|adolescents|people|individuals)\s+were\s+([^.;]{5,180})",
            r"(?:^|[.]\s+)([^.;]{5,180}?)\s+(?:were|was)\s+(?:randomized|randomised|assigned|allocated|enrolled|recruited|included)\b",
        ]
        for pattern in patterns:
            match = re.search(pattern, cleaned_text, flags=re.I)
            if match:
                candidate = re.split(r"\b(?:who|and who|and were)\b", match.group(1), maxsplit=1, flags=re.I)[0]
                candidate = _trim_candidate_phrase(candidate)
                if any(token in candidate.lower() for token in ("trial", "study", "institution", "analysis")):
                    continue
                if candidate and len(candidate) >= 8:
                    return candidate
        sentence = _first_sentence(cleaned_text)
        sentence = _trim_candidate_phrase(sentence)
        if sentence and len(sentence) >= 8 and sentence.lower() not in {"participants", "patients", "subjects"}:
            return sentence
    return ""


def _pick_condition_candidate(candidates: Sequence[str]) -> str:
    keywords = [
        "depress",
        "anxiety",
        "ptsd",
        "stress disorder",
        "fear of flying",
        "dementia",
        "alzheimer",
        "psychosis",
        "schizophrenia",
        "bipolar",
        "cancer",
        "disease",
        "disorder",
        "syndrome",
        "myocardial",
        "coronary",
        "diabetes",
        "obesity",
        "pain",
        "insomnia",
        "trauma",
        "phobia",
    ]
    cleaned_candidates = []
    for candidate in candidates:
        trimmed = _trim_candidate_phrase(candidate)
        if trimmed and len(trimmed) >= 3:
            cleaned_candidates.append(trimmed)
    for candidate in cleaned_candidates:
        lowered = candidate.lower()
        if any(keyword in lowered for keyword in keywords):
            return candidate
    return cleaned_candidates[0] if cleaned_candidates else ""


def _extract_indication_candidate(parsed: ParsedRecord) -> str:
    candidates: List[str] = []
    text_sources = [
        parsed.title,
        parsed.sections.get("OBJECTIVE", ""),
        parsed.sections.get("OBJECTIVES", ""),
        parsed.sections.get("BACKGROUND", ""),
        parsed.sections.get("PATIENTS", ""),
        parsed.sections.get("PARTICIPANTS", ""),
        parsed.sections.get("SUMMARY", ""),
    ]
    specific_patterns = [
        r"\b(?:patients|adults|subjects|participants|women|men|children|adolescents|people|individuals|veterans|carers)\s+with\s+([^.;:]{3,120})",
        r"\bdiagnosed with\s+([^.;:]{3,120})",
        r"\bsuffering from\s+([^.;:]{3,120})",
        r"\bfor\s+people\s+with\s+([^.;:]{3,120})",
    ]
    for text in text_sources:
        cleaned = _clean_fallback_text(text)
        if not cleaned:
            continue
        for pattern in specific_patterns:
            for match in re.finditer(pattern, cleaned, flags=re.I):
                candidates.append(match.group(1))
        if text == parsed.title:
            for pattern in [
                r"\bfor\s+([^.;:]{3,120})",
                r"\bin\s+([^.;:]{3,120})",
                r"\bamong\s+([^.;:]{3,120})",
            ]:
                match = re.search(pattern, cleaned, flags=re.I)
                if match:
                    candidates.append(match.group(1))
    return _pick_condition_candidate(candidates)


def _extract_intervention_candidate(parsed: ParsedRecord) -> str:
    intervention_section = parsed.sections.get("INTERVENTIONS", "")
    if intervention_section:
        sentence = _first_sentence(intervention_section)
        if sentence:
            return sentence

    methods_text = " ".join(
        _clean_fallback_text(parsed.sections.get(key, ""))
        for key in ("INTERVENTIONS", "METHODS", "PATIENTS", "PARTICIPANTS", "OBJECTIVE", "BACKGROUND")
    ).strip()
    method_patterns = [
        r"\brandomi[sz]ed(?:\s+\w+){0,4}\s+to\s+([^.;:]{3,100}?)(?:\s+(?:or|versus|vs\.?)\s+([^.;:]{3,100}))",
        r"\breceived\s+([^.;:]{3,100})",
        r"\btreated with\s+([^.;:]{3,100})",
        r"\bacute\s+([^.;:]{3,80})",
    ]
    for pattern in method_patterns:
        match = re.search(pattern, methods_text, flags=re.I)
        if match:
            candidate = _trim_candidate_phrase(match.group(1))
            if candidate:
                return candidate

    title = _clean_fallback_text(parsed.title)
    title_patterns = [
        r"\beffect of\s+(.+?)\s+on\b",
        r"\bimpact of\s+(.+?)\s+on\b",
        r"\bevaluation of\s+(.+?)(?:\s+for|\s+in|\s+among|\s+versus|\s+vs\.?|\s*$)",
        r"\btrial of\s+(.+?)(?:\s+for|\s+in|\s+among|\s+versus|\s+vs\.?|\s*$)",
        r"\bby\s+([^.;:]{3,100})$",
        r"^(.+?)\s+(?:for|in|among)\s+",
    ]
    for pattern in title_patterns:
        match = re.search(pattern, title, flags=re.I)
        if match:
            candidate = _trim_candidate_phrase(match.group(1))
            if candidate:
                return candidate
    return ""


def _extract_comparator_candidate(parsed: ParsedRecord) -> str:
    merged = " ".join(
        _clean_fallback_text(parsed.sections.get(key, ""))
        for key in ("INTERVENTIONS", "METHODS", "PATIENTS", "PARTICIPANTS", "RESULTS", "SUMMARY")
    ).strip()
    if not merged:
        return ""
    structured_match = re.search(
        r"\brandomi[sz]ed(?:\s+\w+){0,4}\s+to\s+([^.;:]{3,120}?)\s+(?:or|versus|vs\.?)\s+([^.;:]{3,120})",
        merged,
        flags=re.I,
    )
    if structured_match:
        return _trim_candidate_phrase(structured_match.group(2))
    for pattern in [
        r"\b(?:compared with|versus|vs\.?)\s+([^.;:]{3,100})",
        r"\bor\s+([^.;:]{3,100})\s+group\b",
    ]:
        match = re.search(pattern, merged, flags=re.I)
        if match:
            return _trim_candidate_phrase(match.group(1))
    for token in ("placebo", "usual care", "control group", "standard care"):
        if token in merged.lower():
            return token.title()
    return ""


def _extract_result_candidate(parsed: ParsedRecord) -> Tuple[str, str]:
    results_text = _clean_fallback_text(parsed.sections.get("RESULTS", ""))
    conclusions_text = _clean_fallback_text(parsed.sections.get("CONCLUSIONS", ""))
    evidence = _first_sentence(results_text or conclusions_text)
    result = _first_sentence(conclusions_text or results_text)
    return result, evidence


def infer_fallback_row(parsed: ParsedRecord, note: str = "") -> ExtractionRow:
    result_text, evidence_text = _extract_result_candidate(parsed)
    indication = _extract_indication_candidate(parsed)
    intervention = _extract_intervention_candidate(parsed)
    phase = _extract_phase_from_text(parsed.title, *parsed.sections.values())
    sample_size = _extract_sample_size_from_text(parsed.title, *parsed.sections.values())
    follow_up_time = extract_follow_up_time_from_text(parsed.title, *parsed.sections.values())
    population = _extract_population_candidate(parsed)
    comparator = _extract_comparator_candidate(parsed)
    return normalize_row(
        {
            "source_index": str(parsed.record_id),
            "pmid": parsed.pmid,
            "nct_id": parsed.nct_id,
            "journal": parsed.journal_line,
            "year": parsed.year,
            "indication": indication,
            "population_characteristics": population,
            "population_raw": population,
            "intervention": intervention,
            "intervention_type": "",
            "comparator": comparator,
            "result": result_text,
            "outcome_direction": "",
            "phase": phase,
            "sample_size": sample_size,
            "follow_up_time": follow_up_time,
            "evidence_snippet": evidence_text,
            "statistical_metrics": "",
            "notes": note,
        }
    )


def enrich_row_with_fallback(row: ExtractionRow, parsed: ParsedRecord) -> ExtractionRow:
    fallback = infer_fallback_row(parsed)

    def fill_if_blank(attr: str, fallback_value: str) -> None:
        current = str(getattr(row, attr, "") or "").strip()
        if not current and str(fallback_value or "").strip():
            setattr(row, attr, fallback_value)

    for attr in (
        "indication",
        "population_characteristics",
        "population_raw",
        "intervention",
        "comparator",
        "result",
        "phase",
        "sample_size",
        "follow_up_time",
        "evidence_snippet",
        "statistical_metrics",
    ):
        fill_if_blank(attr, getattr(fallback, attr))

    # Allow fallback to overwrite placeholder population values emitted by model.
    current_population = str(row.population_characteristics or "").strip().lower()
    fallback_population = str(fallback.population_characteristics or "").strip()
    if current_population in {"not provided", "unknown", "n/a", "na"} and fallback_population and fallback_population.lower() not in {"not provided", "unknown", "n/a", "na"}:
        row.population_characteristics = fallback_population

    if (not str(row.intervention_type or "").strip() or row.intervention_type == "Other") and fallback.intervention_type != "Other":
        row.intervention_type = fallback.intervention_type
    if (not str(row.outcome_direction or "").strip() or row.outcome_direction == "Mixed or Unknown") and fallback.outcome_direction != "Mixed or Unknown":
        row.outcome_direction = fallback.outcome_direction
    return row


def normalize_row(data: Dict[str, Any]) -> ExtractionRow:
    return ExtractionRow(**normalize_row_fields(data))


def extract_json_block(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("{") and stripped.endswith("}"):
        return stripped
    start = stripped.find("{")
    end = stripped.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise MalformedModelOutputError("Model response did not contain a complete JSON object.")
    return stripped[start : end + 1]


def _http_post_json(
    url: str,
    headers: Dict[str, str],
    payload: Dict[str, Any],
    timeout: int = 180,
    retries: int = 2,
    retry_delay: float = 2.0,
    should_stop: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    try:
        return transport_http_post_json(
            url=url,
            headers=headers,
            payload=payload,
            timeout=timeout,
            retries=retries,
            retry_delay=retry_delay,
            should_stop=should_stop,
        )
    except TransportStopRequested as exc:
        raise UserCancelledError(str(exc)) from exc
    except TransportError as exc:
        raise ExtractionError(str(exc)) from exc


def check_gemini_model_support(model: str, api_key: str, timeout: int = 60) -> Tuple[bool, str]:
    try:
        return shared_check_gemini_model_support(model, api_key, timeout=timeout)
    except GeminiModelError as exc:
        return False, str(exc)


def resolve_gemini_model_id(model: str, api_key: str, timeout: int = 60) -> Tuple[str, str]:
    try:
        return shared_resolve_gemini_model_id(model, api_key, timeout=timeout)
    except GeminiModelError as exc:
        raise ExtractionError(str(exc)) from exc


class LLMExtractor:
    def __init__(
        self,
        provider: str,
        model: str,
        api_key: str,
        base_url: str = "",
        prompt_template: str = DEFAULT_PROMPT_TEMPLATE,
        timeout_seconds: int = 180,
        retries: int = 2,
        should_stop: Optional[Callable[[], bool]] = None,
    ) -> None:
        self.provider = provider
        self.model = resolve_model_name(model)
        self.api_key = api_key.strip()
        self.base_url = base_url.strip().rstrip("/")
        self.prompt_template = prompt_template.strip()
        self.timeout_seconds = max(30, int(timeout_seconds))
        self.retries = max(0, int(retries))
        self.should_stop = should_stop
        if not self.model:
            raise ExtractionError("Please choose or enter a model name.")
        if not self.api_key:
            raise ExtractionError("API key is required.")

    def run_batch(self, records: Sequence[str], record_ids: Sequence[int]) -> BatchResult:
        prompt = f"{self.prompt_template}\n\nBatch records:\n{build_batch_payload(records, record_ids)}\n\nReturn JSON now."
        if self.provider == "Gemini":
            return self._run_gemini(prompt)
        if self.provider == "Groq":
            # Keep Groq on its own endpoint to avoid accidental OpenAI base-url bleed-through.
            return self._run_openai_compatible(prompt, "https://api.groq.com/openai/v1", "Groq")
        if self.provider == "NVIDIA NIM":
            return self._run_openai_compatible(prompt, "https://integrate.api.nvidia.com/v1", "NVIDIA NIM")
        if self.provider == "GitHub Models":
            return self._run_openai_compatible(prompt, "https://models.github.ai/inference", "GitHub Models")
        if self.provider == "Mistral":
            return self._run_openai_compatible(prompt, "https://api.mistral.ai/v1", "Mistral")
        if self.provider == "OpenAI-Compatible":
            if not self.base_url:
                raise ExtractionError("Base URL is required for OpenAI-Compatible mode.")
            return self._run_openai_compatible(prompt, self.base_url, "OpenAI-Compatible")
        raise ExtractionError(f"Unsupported provider: {self.provider}")

    def _parse_model_json(self, text: str) -> Dict[str, Any]:
        json_text = extract_json_block(text)
        try:
            parsed = json.loads(json_text)
        except json.JSONDecodeError as exc:
            raise MalformedModelOutputError(
                f"Model returned invalid JSON near line {exc.lineno}, column {exc.colno}. "
                f"This often happens when the batch is too large."
            ) from exc
        if isinstance(parsed, list):
            return {"rows": [item for item in parsed if isinstance(item, dict)]}
        if not isinstance(parsed, dict):
            raise MalformedModelOutputError(
                f"Model JSON top-level type must be object or array, got {type(parsed).__name__}."
            )
        rows = parsed.get("rows", [])
        if isinstance(rows, dict):
            rows = [rows]
        elif not isinstance(rows, list):
            rows = []
        parsed["rows"] = [item for item in rows if isinstance(item, dict)]
        return parsed

    def _run_gemini(self, prompt: str) -> BatchResult:
        endpoint_root = f"https://generativelanguage.googleapis.com/v1beta/models/{urllib.parse.quote(self.model, safe='')}"
        headers = {"Content-Type": "application/json"}
        usage = TokenUsage(provider="Gemini", model=self.model)

        count_url = f"{endpoint_root}:countTokens?key={urllib.parse.quote(self.api_key, safe='')}"
        try:
            count_data = _http_post_json(
                count_url,
                headers,
                {"contents": [{"role": "user", "parts": [{"text": prompt}]}]},
                timeout=min(30, self.timeout_seconds),
                retries=0,
                should_stop=self.should_stop,
            )
            usage.prompt_tokens = int(count_data.get("totalTokens", 0))
        except UserCancelledError:
            raise
        except Exception:
            usage.prompt_tokens = estimate_tokens(prompt)
            usage.input_tokens_estimated = True

        # Gemma models on Gemini API do not support developer/system instructions.
        # For those models we inline the instruction into the user turn instead.
        if self.model.startswith("gemma-"):
            contents = [
                {
                    "role": "user",
                    "parts": [
                        {
                            "text": (
                                "Return only valid JSON matching the requested schema.\n\n"
                                f"{prompt}"
                            )
                        }
                    ],
                }
            ]
            payload = {
                "contents": contents,
                "generationConfig": {"temperature": 0.1},
            }
        else:
            payload = {
                "systemInstruction": {"parts": [{"text": "Return only valid JSON matching the requested schema."}]},
                "contents": [{"role": "user", "parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.1, "responseMimeType": "application/json"},
            }

        response = _http_post_json(
            f"{endpoint_root}:generateContent?key={urllib.parse.quote(self.api_key, safe='')}",
            headers,
            payload,
            timeout=self.timeout_seconds,
            retries=self.retries,
            should_stop=self.should_stop,
        )
        if not isinstance(response, dict):
            raise ExtractionError(
                f"Gemini returned malformed response type: {type(response).__name__}. "
                f"Response preview: {str(response)[:500]}"
            )
        candidates = response.get("candidates") or []
        if not candidates:
            raise ExtractionError(f"Gemini returned no candidates: {response}")
        first = candidates[0] if candidates else {}
        if not isinstance(first, dict):
            raise ExtractionError(
                f"Gemini returned malformed candidate type: {type(first).__name__}. "
                f"Response preview: {str(response)[:500]}"
            )
        content_obj = first.get("content", {})
        if not isinstance(content_obj, dict):
            content_obj = {}
        parts_obj = content_obj.get("parts", [])
        if not isinstance(parts_obj, list):
            parts_obj = []
        content = "".join(part.get("text", "") for part in parts_obj if isinstance(part, dict))
        parsed = self._parse_model_json(content)

        meta = response.get("usageMetadata", {})
        if not isinstance(meta, dict):
            meta = {}
        usage.prompt_tokens = int(meta.get("promptTokenCount", usage.prompt_tokens))
        usage.completion_tokens = int(meta.get("candidatesTokenCount", 0))
        # Some models/endpoints omit completion token usage even when text is returned.
        if usage.completion_tokens == 0 and content.strip():
            usage.completion_tokens = estimate_tokens(content)
        usage.total_tokens = int(meta.get("totalTokenCount", usage.prompt_tokens + usage.completion_tokens))
        rows = [normalize_row(item) for item in parsed.get("rows", []) if isinstance(item, dict)]
        return BatchResult(rows=rows, usage=usage, raw_json=content)

    def _run_openai_compatible(self, prompt: str, base_url: str, provider_name: str) -> BatchResult:
        usage = TokenUsage(provider=provider_name, model=self.model, prompt_tokens=estimate_tokens(prompt), input_tokens_estimated=True)
        parsed: Optional[Dict[str, Any]] = None
        url = f"{base_url}/chat/completions"
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {self.api_key}"}
        if provider_name == "GitHub Models":
            headers["Accept"] = "application/vnd.github+json"
            headers["X-GitHub-Api-Version"] = "2022-11-28"
        payload = {
            "model": self.model,
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": "Return only valid JSON matching the requested schema."},
                {"role": "user", "content": prompt},
            ],
        }
        try:
            response = _http_post_json(
                url,
                headers,
                payload,
                timeout=self.timeout_seconds,
                retries=self.retries,
                should_stop=self.should_stop,
            )
        except ExtractionError as exc:
            # Some compatible endpoints/models fail strict json_object validation.
            # Retry once without response_format and parse JSON from plain text output.
            if "json_validate_failed" not in str(exc).lower() and "failed to validate json" not in str(exc).lower():
                raise
            fallback_payload = {
                "model": self.model,
                "temperature": 0.1,
                "messages": [
                    {"role": "system", "content": "Return only valid JSON matching the requested schema. No markdown."},
                    {"role": "user", "content": prompt},
                ],
            }
            response = _http_post_json(
                url,
                headers,
                fallback_payload,
                timeout=self.timeout_seconds,
                retries=max(0, self.retries - 1),
                should_stop=self.should_stop,
            )
        if isinstance(response, str):
            parsed = self._parse_model_json(response)
            choices_raw = None
        elif isinstance(response, dict):
            choices_raw = response.get("choices")
        else:
            raise ExtractionError(
                f"{provider_name} returned malformed response type: {type(response).__name__}. "
                f"Response preview: {str(response)[:500]}"
            )
        if choices_raw is None:
            # Some providers may return plain text in compatibility mode.
            if isinstance(response, dict):
                direct_text = response.get("output_text") or response.get("text") or ""
                if not str(direct_text).strip():
                    if parsed is None:
                        raise ExtractionError(f"{provider_name} returned no choices: {response}")
                else:
                    parsed = self._parse_model_json(str(direct_text))
            elif parsed is None:
                raise ExtractionError(f"{provider_name} returned no choices.")
        else:
            if isinstance(choices_raw, list):
                choices = choices_raw
            elif isinstance(choices_raw, dict):
                choices = [choices_raw]
            else:
                raise ExtractionError(
                    f"{provider_name} returned malformed choices type: {type(choices_raw).__name__}. "
                    f"Response preview: {str(response)[:500]}"
                )
            if not choices:
                raise ExtractionError(f"{provider_name} returned no choices: {response}")

            first_choice = choices[0]
            if isinstance(first_choice, dict):
                message_obj = first_choice.get("message", first_choice)
                if isinstance(message_obj, dict):
                    content_obj = message_obj.get("content", "")
                else:
                    content_obj = message_obj
            else:
                content_obj = first_choice

            if isinstance(content_obj, str):
                content_text = content_obj
            elif isinstance(content_obj, list):
                parts: List[str] = []
                for item in content_obj:
                    if isinstance(item, dict):
                        parts.append(str(item.get("text", item.get("content", ""))))
                    else:
                        parts.append(str(item))
                content_text = "".join(parts)
            elif isinstance(content_obj, dict):
                content_text = str(content_obj.get("text", content_obj.get("content", "")))
            else:
                content_text = str(content_obj)

            if not content_text.strip():
                raise ExtractionError(
                    f"{provider_name} returned empty message content. "
                    f"Response preview: {str(response)[:500]}"
                )
            parsed = self._parse_model_json(content_text)
        if parsed is None:
            raise ExtractionError(f"{provider_name} returned no parseable JSON content.")
        api_usage = response.get("usage") if isinstance(response, dict) else {}
        if not isinstance(api_usage, dict):
            api_usage = {}
        usage.prompt_tokens = int(api_usage.get("prompt_tokens", usage.prompt_tokens))
        usage.completion_tokens = int(api_usage.get("completion_tokens", 0))
        usage.total_tokens = int(api_usage.get("total_tokens", usage.prompt_tokens + usage.completion_tokens))
        usage.input_tokens_estimated = False if api_usage else True
        rows = [normalize_row(item) for item in parsed.get("rows", []) if isinstance(item, dict)]
        return BatchResult(rows=rows, usage=usage, raw_json=json.dumps(parsed, ensure_ascii=False))


class CheckpointStore:
    def __init__(self, checkpoint_path: str) -> None:
        self.path = Path(checkpoint_path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def reset(self, source_path: str, provider: str, model: str) -> None:
        if self.path.exists():
            self.path.unlink()
        self.append(
            {
                "type": "meta",
                "source_path": source_path,
                "provider": provider,
                "model": model,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    def append(self, item: Dict[str, Any]) -> None:
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(item, ensure_ascii=False) + "\n")
            handle.flush()

    def append_batch(
        self,
        batch_index: int,
        rows: Sequence[ExtractionRow],
        usage: TokenUsage,
        completed_record_ids: Optional[Sequence[int]] = None,
    ) -> None:
        for row in rows:
            self.append({"type": "row", "batch_index": batch_index, "data": row.to_export_dict()})
        self.append(
            {
                "type": "usage",
                "batch_index": batch_index,
                "completed_record_ids": [int(x) for x in (completed_record_ids or [])],
                "data": {
                    "provider": usage.provider,
                    "model": usage.model,
                    "prompt_tokens": usage.prompt_tokens,
                    "completion_tokens": usage.completion_tokens,
                    "total_tokens": usage.total_tokens,
                    "input_tokens_estimated": usage.input_tokens_estimated,
                },
            }
        )

    def has_rows(self) -> bool:
        if not self.path.exists():
            return False
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if payload.get("type") == "row":
                    return True
        return False

    def has_progress(self) -> bool:
        if not self.path.exists():
            return False
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if payload.get("type") in {"row", "usage"}:
                    return True
        return False

    def read_meta(self) -> Dict[str, Any]:
        if not self.path.exists():
            return {}
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if payload.get("type") == "meta":
                    return payload
        return {}


def load_checkpoint_progress(checkpoint_path: str) -> Tuple[set[int], int, int]:
    path = Path(checkpoint_path)
    if not path.exists():
        return set(), 0, 0

    completed_record_ids: set[int] = set()
    row_count = 0
    usage_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            if payload.get("type") == "row":
                data = payload.get("data", {})
                source_index = normalize_source_index_value(data.get("Record Index", ""))
                if source_index.isdigit():
                    completed_record_ids.add(int(source_index))
                row_count += 1
            elif payload.get("type") == "usage":
                usage_count += 1
                for rid in payload.get("completed_record_ids", []) or []:
                    try:
                        completed_record_ids.add(int(rid))
                    except Exception:
                        continue
    return completed_record_ids, row_count, usage_count


def rebuild_excel_from_checkpoint(checkpoint_path: str, output_path: str) -> Tuple[int, int]:
    checkpoint = Path(checkpoint_path)
    if not checkpoint.exists():
        return 0, 0

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Results"
    ws_non_dep = wb.create_sheet("Non-Depression Results")
    ws_placeholder = wb.create_sheet("No-Extraction Results")
    ws_stats = wb.create_sheet("Request Stats")
    ws_info = wb.create_sheet("Run Info")

    export_headers_with_record_id = [RECORD_ID_HEADER, *EXPORT_HEADERS]

    ws_data.append(export_headers_with_record_id)
    ws_non_dep.append(export_headers_with_record_id)
    ws_placeholder.append(export_headers_with_record_id)
    ws_stats.append(["Batch", "Provider", "Model", "Input Tokens", "Output Tokens", "Total Tokens", "Input Estimated"])
    ws_info.append(["Field", "Value"])
    for ws in (ws_data, ws_non_dep, ws_placeholder, ws_stats, ws_info):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    row_count = 0
    usage_count = 0
    meta: Dict[str, Any] = {}
    pending_rows: List[Dict[str, Any]] = []
    with checkpoint.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if item.get("type") == "meta":
                meta = item
            elif item.get("type") == "row":
                data = item.get("data", {})
                pending_rows.append(data)
                row_count += 1
            elif item.get("type") == "usage":
                data = item.get("data", {})
                ws_stats.append(
                    [
                        item.get("batch_index", ""),
                        data.get("provider", ""),
                        data.get("model", ""),
                        data.get("prompt_tokens", 0),
                        data.get("completion_tokens", 0),
                        data.get("total_tokens", 0),
                        "Yes" if data.get("input_tokens_estimated") else "No",
                    ]
                )
                usage_count += 1

    def row_sort_key(data: Dict[str, Any], order_index: int) -> Tuple[int, int]:
        record_index = normalize_source_index_value(data.get("Record Index", ""))
        if record_index.isdigit():
            return int(record_index), order_index
        # Put malformed/empty record indices at the end but preserve original relative order.
        return 10**12, order_index

    sorted_rows = sorted(
        enumerate(pending_rows),
        key=lambda pair: row_sort_key(pair[1], pair[0]),
    )

    source_path = str(meta.get("source_path", "")).strip()
    metadata_by_id: Dict[int, Dict[str, str]] = {}
    ids_needing_backfill: set[int] = set()
    if source_path and Path(source_path).exists():
        for _pos, row_data in sorted_rows:
            record_index = normalize_source_index_value(row_data.get("Record Index", ""))
            if not record_index.isdigit():
                continue
            rid = int(record_index)
            if not row_data.get("PMID") or not row_data.get("Journal") or not row_data.get("Year") or not row_data.get("NCT ID"):
                ids_needing_backfill.add(rid)

        if ids_needing_backfill:
            for rec_text, rec_ids in chunk_records_iter(source_path, 200):
                # Keep this as a fast pass; parse only needed IDs.
                for raw_record, rid in zip(rec_text, rec_ids):
                    if rid not in ids_needing_backfill:
                        continue
                    parsed = parse_record_structured(raw_record, rid)
                    metadata_by_id[rid] = {
                        "PMID": parsed.pmid,
                        "NCT ID": parsed.nct_id,
                        "Journal": parsed.journal_line,
                        "Year": parsed.year,
                    }
                if len(metadata_by_id) >= len(ids_needing_backfill):
                    break

    dep_count = 0
    non_dep_count = 0
    placeholder_count = 0
    promoted_to_results_count = 0
    next_result_record_id = 1
    next_fallback_record_id = 1

    source_label = sanitize_record_id_source_label(source_path)

    def _export_row_values(_sheet_name: str, export_data: Dict[str, Any]) -> List[str]:
        nonlocal next_result_record_id
        nonlocal next_fallback_record_id
        if _sheet_name == "Results":
            record_id = build_prefixed_record_id(RESULT_RECORD_ID_PREFIX, source_label, next_result_record_id)
            next_result_record_id += 1
        else:
            record_id = build_prefixed_record_id(FALLBACK_RECORD_ID_PREFIX, source_label, next_fallback_record_id)
            next_fallback_record_id += 1
        return [record_id, *[export_data.get(header, "") for header in EXPORT_HEADERS]]

    prepared_rows: List[Dict[str, Any]] = []
    result_pmids: set[str] = set()

    for _pos, data in sorted_rows:
        normalized_index = normalize_source_index_value(data.get("Record Index", ""))
        if normalized_index:
            data["Record Index"] = normalized_index
        if normalized_index.isdigit():
            info = metadata_by_id.get(int(normalized_index), {})
            if info:
                for key in ("PMID", "NCT ID", "Journal", "Year"):
                    if not str(data.get(key, "")).strip() and str(info.get(key, "")).strip():
                        data[key] = info[key]

        normalized_row = normalize_row(
            {
                "source_index": data.get("Record Index", ""),
                "pmid": data.get("PMID", ""),
                "nct_id": data.get("NCT ID", ""),
                "journal": data.get("Journal", ""),
                "year": data.get("Year", ""),
                "indication": data.get("Indication", ""),
                "population_characteristics": data.get("Population Characteristics", ""),
                "population_raw": data.get("Population Raw", ""),
                "target": data.get("Target", ""),
                "intervention": data.get("Intervention", ""),
                "intervention_type": data.get("Intervention Type", ""),
                "comparator": data.get("Comparator", ""),
                "result": data.get("Result", ""),
                "outcome_direction": data.get("Outcome Direction", ""),
                "phase": data.get("Phase", ""),
                "sample_size": data.get("Sample Size", ""),
                "follow_up_time": data.get("Follow-up Time", ""),
                "evidence_snippet": data.get("Evidence Snippet", ""),
                "statistical_metrics": data.get("Statistical Metrics", ""),
                "notes": data.get("Notes", ""),
                "population_age_type": data.get("Population Age Type", ""),
                "population_age_value": data.get("Population Age Value", ""),
                "population_age_sd": data.get("Population Age SD", ""),
                "population_age_unit": data.get("Population Age Unit", ""),
                "population_age_descriptor": data.get("Population Age Descriptor", ""),
                "population_age_evidence_span": data.get("Population Age Evidence Span", ""),
                "population_gender": data.get("Population Gender", ""),
                "population_gender_evidence_span": data.get("Population Gender Evidence Span", ""),
                "population_severity": data.get("Population Severity", ""),
                "population_severity_evidence_span": data.get("Population Severity Evidence Span", ""),
                "population_ethnicity": data.get("Population Ethnicity", ""),
                "population_ethnicity_evidence_span": data.get("Population Ethnicity Evidence Span", ""),
                "population_occupation": data.get("Population Occupation", ""),
                "population_occupation_evidence_span": data.get("Population Occupation Evidence Span", ""),
                "population_social_status": data.get("Population Social Status", ""),
                "population_social_status_evidence_span": data.get("Population Social Status Evidence Span", ""),
                "population_previous_treatment": data.get("Population Previous Treatment", ""),
                "population_previous_treatment_evidence_span": data.get("Population Previous Treatment Evidence Span", ""),
            }
        )
        export_data = normalized_row.to_export_dict()
        indication_text = str(export_data.get("Indication", "")).strip()
        notes_text = str(export_data.get("Notes", "")).strip()
        pmid_text = str(export_data.get("PMID", "")).strip()
        if notes_text.startswith(NO_EXTRACTION_NOTE) or notes_text.startswith(NO_ABSTRACT_NOTE):
            sheet_name = "No-Extraction Results"
        elif indication_text and not is_depression_related_indication(indication_text):
            sheet_name = "Non-Depression Results"
        else:
            sheet_name = "Results"
            if pmid_text:
                result_pmids.add(pmid_text)
        prepared_rows.append(
            {
                "sheet_name": sheet_name,
                "pmid": pmid_text,
                "export_data": export_data,
            }
        )

    for item in prepared_rows:
        sheet_name = str(item.get("sheet_name", "")).strip() or "Results"
        pmid_text = str(item.get("pmid", "")).strip()
        export_data = dict(item.get("export_data", {}))
        if sheet_name != "Results" and pmid_text and pmid_text in result_pmids:
            sheet_name = "Results"
            promoted_to_results_count += 1
        if sheet_name == "No-Extraction Results":
            ws_placeholder.append(_export_row_values(sheet_name, export_data))
            placeholder_count += 1
        elif sheet_name == "Non-Depression Results":
            ws_non_dep.append(_export_row_values(sheet_name, export_data))
            non_dep_count += 1
        else:
            ws_data.append(_export_row_values(sheet_name, export_data))
            dep_count += 1

    ws_info.append(["Checkpoint", str(checkpoint)])
    ws_info.append(["Source File", meta.get("source_path", "")])
    ws_info.append(["Provider", meta.get("provider", "")])
    ws_info.append(["Model", meta.get("model", "")])
    ws_info.append(["Started At", meta.get("started_at", "")])
    ws_info.append(["Rebuilt At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_info.append(["Extracted Rows", row_count])
    ws_info.append(["Depression Sheet Rows", dep_count])
    ws_info.append(["Non-Depression Sheet Rows", non_dep_count])
    ws_info.append(["No-Extraction Sheet Rows", placeholder_count])
    ws_info.append(["Promoted To Results By PMID Overlap", promoted_to_results_count])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return row_count, usage_count


def _should_split_batch(exc: Exception, record_count: int) -> bool:
    return scheduler_should_split_batch(exc, record_count)


def _is_hard_request_quota_error(exc: Exception) -> bool:
    return scheduler_is_hard_request_quota_error(exc)


def _run_batch_adaptive(
    extractor: LLMExtractor,
    records: Sequence[str],
    record_ids: Sequence[int],
    progress: Optional[Callable[[str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
) -> List[Tuple[BatchResult, List[int]]]:
    return scheduler_run_batch_adaptive(
        run_batch_fn=lambda recs, ids: extractor.run_batch(recs, ids),
        records=records,
        record_ids=record_ids,
        progress=progress,
        should_stop=should_stop,
        cancel_exception_cls=UserCancelledError,
    )


def _covered_source_ids_from_rows(rows: Sequence[ExtractionRow]) -> List[int]:
    return scheduler_covered_source_ids_from_rows(rows, normalize_source_index_value)


def _recover_missing_rows(
    extractor: LLMExtractor,
    records: Sequence[str],
    record_ids: Sequence[int],
    progress: Optional[Callable[[str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
) -> List[Tuple[BatchResult, List[int]]]:
    return scheduler_recover_missing_rows(
        run_batch_fn=lambda recs, ids: extractor.run_batch(recs, ids),
        records=records,
        record_ids=record_ids,
        covered_source_ids_fn=lambda rows: scheduler_covered_source_ids_from_rows(rows, normalize_source_index_value),
        progress=progress,
        should_stop=should_stop,
        cancel_exception_cls=UserCancelledError,
    )


def process_file(
    source_path: str,
    provider: str,
    model: str,
    api_key: str,
    batch_size: int,
    prompt_template: str,
    base_url: str = "",
    timeout_seconds: int = 180,
    retries: int = 2,
    progress: Optional[Callable[[str], None]] = None,
    on_batch_done: Optional[
        Callable[
            [
                int,  # request_index
                int,  # main_batch_total
                int,  # main_batch_done
                int,  # recovery_batch_done
                int,  # recovery_batch_total
                int,  # extracted_rows_done
                int,  # extracted_rows_target
                List[ExtractionRow],
                TokenUsage,
                List[int],
            ],
            None,
        ]
    ] = None,
    should_stop: Optional[Callable[[], bool]] = None,
    skip_record_ids: Optional[set[int]] = None,
    concurrency: int = 1,
) -> Tuple[int, List[TokenUsage], List[ExtractionRow]]:
    MAX_DEFERRED_RECOVERY_ATTEMPTS = 3
    stats = summarize_input(source_path, batch_size, prompt_template)
    if stats.record_count == 0:
        raise ExtractionError("No records were detected in the input file.")
    if progress:
        progress(
            f"Detected {stats.record_count} records. "
            f"Planned requests: {stats.batch_count}. "
            f"Estimated input tokens: {stats.estimated_input_tokens}."
        )

    extractor = LLMExtractor(
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        prompt_template=prompt_template,
        timeout_seconds=timeout_seconds,
        retries=retries,
        should_stop=should_stop,
    )
    if provider == "Gemini":
        resolved_model, resolution_hint = resolve_gemini_model_id(
            extractor.model,
            api_key=extractor.api_key,
            timeout=min(60, extractor.timeout_seconds),
        )
        extractor.model = resolved_model
        if progress and resolution_hint:
            progress(resolution_hint)
    if progress:
        progress(f"Using resolved model id: {extractor.model}")

    skipped = skip_record_ids or set()
    if skipped and progress:
        progress(f"Resume mode: skipping {len(skipped)} previously completed records from checkpoint.")

    concurrency = max(1, int(concurrency))
    allow_parallel = provider in {"NVIDIA NIM", "GitHub Models", "Mistral"} or (
        provider == "Gemini"
        and extractor.model in {
            "gemini-3.1-flash-lite",
            "gemini-3.1-flash-lite-preview",
            "gemma-3-27b-it",
            "gemma-3-12b-it",
        }
    )
    if not allow_parallel:
        concurrency = 1
    else:
        concurrency = min(100, concurrency)
    guarded = apply_runtime_guards(
        provider=provider,
        model=extractor.model,
        timeout_seconds=extractor.timeout_seconds,
        retries=extractor.retries,
        concurrency=concurrency,
        progress=progress,
        label=f"Model-specific guard for {extractor.model}",
    )
    extractor.timeout_seconds = guarded.timeout_seconds
    extractor.retries = guarded.retries
    concurrency = guarded.concurrency
    if progress and concurrency > 1:
        progress(f"Parallel mode enabled: concurrency={concurrency} for {extractor.model}.")

    total_rows = 0
    usages: List[TokenUsage] = []
    preview_rows: List[ExtractionRow] = []
    logical_batch_index = 0
    deferred_lock = threading.Lock()
    deferred_recoveries: List[DeferredRecoveryItem] = []
    recovery_batches_done = 0
    recovery_batch_inflight = 0

    planned_batches: List[Tuple[int, List[str], List[int]]] = []
    for batch_records, batch_ids in chunk_records_iter(source_path, batch_size):
        if skipped:
            filtered_pairs = [(r, i) for r, i in zip(batch_records, batch_ids) if i not in skipped]
            if not filtered_pairs:
                continue
            batch_records = [x[0] for x in filtered_pairs]
            batch_ids = [x[1] for x in filtered_pairs]
        planned_batches.append((len(planned_batches) + 1, batch_records, batch_ids))
    remaining_record_count = sum(len(batch_ids) for _, _, batch_ids in planned_batches)
    effective_batch_count = len(planned_batches)
    extracted_row_target = remaining_record_count
    extracted_source_ids: set[int] = set()
    if progress and skipped:
        progress(
            f"Resume remaining: {remaining_record_count} records, "
            f"{effective_batch_count} requests (after skipping completed records)."
        )
    if effective_batch_count == 0:
        if progress:
            progress("All records were already completed in checkpoint. Nothing new to process.")
        return 0, [], []

    def build_fallback_rows(items: Sequence[DeferredRecoveryItem], note: str = NO_EXTRACTION_NOTE) -> List[ExtractionRow]:
        rows: List[ExtractionRow] = []
        for item in items:
            fallback_row = infer_fallback_row(item.parsed, note=note)
            if not fallback_row.intervention_type:
                fallback_row.intervention_type = "Other"
            if not fallback_row.outcome_direction:
                fallback_row.outcome_direction = "Mixed or Unknown"
            if not fallback_row.follow_up_time:
                fallback_row.follow_up_time = item.metadata.get("follow_up_time", "")
            rows.append(fallback_row)
        return rows

    def normalize_sub_results(sub_results: List[Tuple[BatchResult, List[int]]]) -> List[Tuple[BatchResult, List[int]]]:
        normalized: List[Tuple[BatchResult, List[int]]] = []
        for result, _ in sub_results:
            normalized.append((result, _covered_source_ids_from_rows(result.rows)))
        return normalized

    completed_planned_batches = 0

    def consume_sub_results(
        sub_results: List[Tuple[BatchResult, List[int]]],
        completed_batch_count: int,
    ) -> None:
        nonlocal total_rows
        nonlocal recovery_batches_done
        nonlocal recovery_batch_inflight
        for result, completed_ids in sub_results:
            if should_stop and should_stop():
                raise UserCancelledError("Stopped by user request.")
            total_rows += len(result.rows)
            usages.append(result.usage)
            for cid in completed_ids:
                extracted_source_ids.add(cid)
            if len(preview_rows) < 300:
                preview_rows.extend(result.rows[: 300 - len(preview_rows)])
            if on_batch_done:
                with deferred_lock:
                    queued_recovery_batches = math.ceil(len(deferred_recoveries) / max(1, batch_size))
                recovery_total = recovery_batches_done + recovery_batch_inflight + queued_recovery_batches
                on_batch_done(
                    len(usages),
                    effective_batch_count,
                    completed_batch_count,
                    recovery_batches_done,
                    max(recovery_batches_done, recovery_total),
                    len(extracted_source_ids),
                    extracted_row_target,
                    result.rows,
                    result.usage,
                    completed_ids,
                )
            if progress:
                progress(
                    f"Completed request {len(usages)}: rows={len(result.rows)}, "
                    f"input={result.usage.prompt_tokens}, output={result.usage.completion_tokens}, total={result.usage.total_tokens}"
                )

    def enqueue_deferred(items: Sequence[DeferredRecoveryItem]) -> None:
        if not items:
            return
        with deferred_lock:
            deferred_recoveries.extend(items)

    def drain_deferred_recoveries(force_fallback: bool = False) -> None:
        nonlocal recovery_batches_done
        nonlocal recovery_batch_inflight
        round_index = 0
        while True:
            with deferred_lock:
                if not deferred_recoveries:
                    return
                current_round_items = list(deferred_recoveries)
                deferred_recoveries.clear()

            round_index += 1
            round_chunk_count = math.ceil(len(current_round_items) / max(1, batch_size))
            if progress:
                progress(
                    f"Starting deferred recovery round {round_index}: "
                    f"{len(current_round_items)} records in {round_chunk_count} repacked batches."
                )

            next_round_items: List[DeferredRecoveryItem] = []
            for chunk_start in range(0, len(current_round_items), batch_size):
                chunk = current_round_items[chunk_start : chunk_start + batch_size]
                recovery_batch_inflight = 1
                try:
                    if force_fallback or (should_stop and should_stop()):
                        fallback_rows = build_fallback_rows(chunk, note=NO_EXTRACTION_NOTE)
                        fallback_result = BatchResult(
                            rows=fallback_rows,
                            usage=TokenUsage(provider=extractor.provider, model=extractor.model, prompt_tokens=0, completion_tokens=0, total_tokens=0),
                            raw_json="{}",
                        )
                        consume_sub_results(normalize_sub_results([(fallback_result, [])]), completed_planned_batches)
                        recovery_batches_done += 1
                        continue

                    if progress:
                        progress(
                            f"Running deferred recovery round {round_index} batch "
                            f"{(chunk_start // batch_size) + 1}/{round_chunk_count} with {len(chunk)} records."
                        )

                    recovered = _run_batch_adaptive(
                        extractor,
                        [item.llm_text for item in chunk],
                        [item.record_id for item in chunk],
                        progress,
                        should_stop,
                    )
                    recovered = normalize_sub_results(recovered)

                    item_by_id = {item.record_id: item for item in chunk}
                    seen_ids: set[int] = set()
                    for result, _ in recovered:
                        for row in result.rows:
                            row.source_index = normalize_source_index_value(row.source_index)
                            try:
                                rid = int(str(row.source_index).strip())
                            except Exception:
                                rid = 0
                            item = item_by_id.get(rid)
                            if not item:
                                continue
                            seen_ids.add(rid)
                            row.pmid = item.metadata.get("pmid", "")
                            parsed_nct = item.metadata.get("nct_id", "")
                            if parsed_nct:
                                row.nct_id = parsed_nct
                            row.journal = item.metadata.get("journal", "")
                            parsed_year = item.metadata.get("year", "")
                            if parsed_year:
                                row.year = parsed_year
                            if not row.follow_up_time:
                                row.follow_up_time = item.metadata.get("follow_up_time", "")
                            enrich_row_with_fallback(row, item.parsed)

                    missing_items = [item for item in chunk if item.record_id not in seen_ids]
                    if missing_items:
                        if len(seen_ids) == 0:
                            # User rule: 0/N means stop repacking this chunk.
                            if progress:
                                progress(
                                    f"Deferred recovery round {round_index} batch "
                                    f"{(chunk_start // batch_size) + 1}/{round_chunk_count} returned 0/{len(chunk)}. "
                                    f"Stopping repack for this batch and using placeholder rows."
                                )
                            fallback_result = BatchResult(
                                rows=build_fallback_rows(missing_items, note=NO_EXTRACTION_NOTE),
                                usage=TokenUsage(provider=extractor.provider, model=extractor.model, prompt_tokens=0, completion_tokens=0, total_tokens=0),
                                raw_json="{}",
                            )
                            recovered.append((fallback_result, []))
                        else:
                            if progress:
                                progress(
                                    f"Model returned rows for {len(seen_ids)}/{len(chunk)} in deferred recovery round {round_index} batch "
                                    f"{(chunk_start // batch_size) + 1}/{round_chunk_count}. "
                                    f"Queueing {len(missing_items)} records for next recovery round."
                                )
                            for item in missing_items:
                                next_attempt = item.attempts + 1
                                if next_attempt >= MAX_DEFERRED_RECOVERY_ATTEMPTS:
                                    fallback_result = BatchResult(
                                        rows=build_fallback_rows([item], note=NO_EXTRACTION_NOTE),
                                        usage=TokenUsage(provider=extractor.provider, model=extractor.model, prompt_tokens=0, completion_tokens=0, total_tokens=0),
                                        raw_json="{}",
                                    )
                                    recovered.append((fallback_result, []))
                                else:
                                    next_round_items.append(
                                        DeferredRecoveryItem(
                                            record_id=item.record_id,
                                            llm_text=item.llm_text,
                                            parsed=item.parsed,
                                            metadata=item.metadata,
                                            attempts=next_attempt,
                                        )
                                    )

                    consume_sub_results(recovered, completed_planned_batches)
                    recovery_batches_done += 1
                finally:
                    recovery_batch_inflight = 0

            if not next_round_items:
                return
            with deferred_lock:
                deferred_recoveries.extend(next_round_items)

    def run_one_batch(planned_index: int, batch_records: List[str], batch_ids: List[int]) -> List[Tuple[BatchResult, List[int]]]:
        parsed_records = [
            parse_record_structured(record_text, record_id)
            for record_text, record_id in zip(batch_records, batch_ids)
        ]
        batch_metadata: Dict[int, Dict[str, str]] = {
            parsed.record_id: {
                "pmid": parsed.pmid,
                "nct_id": parsed.nct_id,
                "journal": parsed.journal_line,
                "year": parsed.year,
                "follow_up_time": extract_follow_up_time_from_text(" ".join(parsed.sections.values())),
            }
            for parsed in parsed_records
        }
        parsed_by_id: Dict[int, ParsedRecord] = {parsed.record_id: parsed for parsed in parsed_records}
        llm_records = [format_parsed_record_for_llm(parsed) for parsed in parsed_records]
        if progress:
            progress(f"Running planned batch {planned_index}/{effective_batch_count} with {len(batch_records)} records...")
        sub_results = _run_batch_adaptive(extractor, llm_records, batch_ids, progress, should_stop)
        initially_seen: set[int] = set()
        for result, _completed_ids in sub_results:
            for rid in _covered_source_ids_from_rows(result.rows):
                initially_seen.add(rid)

        deferred_items = [
            DeferredRecoveryItem(
                record_id=rid,
                llm_text=record_text,
                parsed=parsed_by_id[rid],
                metadata=batch_metadata.get(rid, {}),
                attempts=0,
            )
            for record_text, rid in zip(llm_records, batch_ids)
            if rid not in initially_seen and record_text.strip()
        ]
        if deferred_items:
            if progress:
                progress(
                    f"Model omitted {len(deferred_items)} records from batch {planned_index}. "
                    f"Queued them for deferred recovery so they can be repacked into fuller requests."
                )
            enqueue_deferred(deferred_items)

        seen_source_ids: set[int] = set()
        llm_text_by_id = {rid: record_text for record_text, rid in zip(llm_records, batch_ids)}
        for result, _completed_ids in sub_results:
            for row in result.rows:
                row.source_index = normalize_source_index_value(row.source_index)
                try:
                    source_idx = int(str(row.source_index).strip())
                except Exception:
                    source_idx = 0
                metadata = batch_metadata.get(source_idx)
                if metadata:
                    seen_source_ids.add(source_idx)
                    row.pmid = metadata.get("pmid", "")
                    parsed_nct = metadata.get("nct_id", "")
                    if parsed_nct:
                        row.nct_id = parsed_nct
                    row.journal = metadata.get("journal", "")
                    parsed_year = metadata.get("year", "")
                    if parsed_year:
                        row.year = parsed_year
                    if not row.follow_up_time:
                        row.follow_up_time = metadata.get("follow_up_time", "")
                    parsed = parsed_by_id.get(source_idx)
                    if parsed:
                        enrich_row_with_fallback(row, parsed)

        # Guarantee at least one output row per source record.
        missing_ids = [rid for rid in batch_ids if rid not in seen_source_ids and not llm_text_by_id.get(rid, "").strip()]
        if missing_ids:
            target_result: Optional[BatchResult] = sub_results[0][0] if sub_results else None
            if target_result is None:
                target_result = BatchResult(
                    rows=[],
                    usage=TokenUsage(provider=extractor.provider, model=extractor.model, prompt_tokens=0, completion_tokens=0, total_tokens=0),
                    raw_json="{}",
                )
                sub_results = [(target_result, list(batch_ids))]

            for rid in missing_ids:
                parsed = parsed_by_id.get(rid)
                placeholder_note = NO_ABSTRACT_NOTE if not llm_text_by_id.get(rid, "").strip() else NO_EXTRACTION_NOTE
                if parsed:
                    fallback_row = infer_fallback_row(parsed, note=placeholder_note)
                    if not fallback_row.intervention_type:
                        fallback_row.intervention_type = "Other"
                    if not fallback_row.outcome_direction:
                        fallback_row.outcome_direction = "Mixed or Unknown"
                    if not fallback_row.follow_up_time:
                        fallback_row.follow_up_time = batch_metadata.get(rid, {}).get("follow_up_time", "")
                    target_result.rows.append(fallback_row)
                else:
                    metadata = batch_metadata.get(rid, {})
                    target_result.rows.append(
                        ExtractionRow(
                            source_index=str(rid),
                            pmid=metadata.get("pmid", ""),
                            nct_id=metadata.get("nct_id", ""),
                            journal=metadata.get("journal", ""),
                            year=metadata.get("year", ""),
                            intervention_type="Other",
                            outcome_direction="Mixed or Unknown",
                            follow_up_time=metadata.get("follow_up_time", ""),
                            notes=placeholder_note,
                        )
                    )
        return normalize_sub_results(sub_results)

    try:
        if concurrency == 1:
            for planned_index, batch_records, batch_ids in planned_batches:
                if should_stop and should_stop():
                    raise UserCancelledError("Stopped by user request.")
                logical_batch_index += 1
                sub_results = run_one_batch(planned_index, batch_records, batch_ids)
                completed_planned_batches += 1
                consume_sub_results(sub_results, completed_planned_batches)
            drain_deferred_recoveries(force_fallback=bool(should_stop and should_stop()))
            return total_rows, usages, preview_rows

        # Parallel execution path for provider/model combinations that are allowed above.
        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            remaining = list(planned_batches)
            in_flight = {}
            last_parallel_heartbeat = time.time()

            def submit_more() -> None:
                while remaining and len(in_flight) < concurrency:
                    planned_index, records_part, ids_part = remaining.pop(0)
                    future = pool.submit(run_one_batch, planned_index, records_part, ids_part)
                    in_flight[future] = (planned_index, list(ids_part))

            submit_more()
            while in_flight:
                if should_stop and should_stop():
                    for fut in in_flight:
                        fut.cancel()
                    raise UserCancelledError("Stopped by user request.")
                done, _ = wait(in_flight.keys(), timeout=0.5, return_when=FIRST_COMPLETED)
                if not done:
                    now = time.time()
                    if progress and now - last_parallel_heartbeat >= 8:
                        progress(
                            f"Parallel worker heartbeat: in_flight={len(in_flight)}, "
                            f"queued={len(remaining)}, completed_planned={completed_planned_batches}/{effective_batch_count}, "
                            f"extracted={len(extracted_source_ids)}/{extracted_row_target}."
                        )
                        last_parallel_heartbeat = now
                    continue
                for fut in done:
                    planned_index, _batch_ids = in_flight.pop(fut)
                    logical_batch_index = max(logical_batch_index, planned_index)
                    sub_results = fut.result()
                    completed_planned_batches += 1
                    consume_sub_results(sub_results, completed_planned_batches)
                submit_more()

        drain_deferred_recoveries(force_fallback=bool(should_stop and should_stop()))
        return total_rows, usages, preview_rows
    except UserCancelledError:
        drain_deferred_recoveries(force_fallback=True)
        raise
    except Exception:
        drain_deferred_recoveries(force_fallback=True)
        raise
