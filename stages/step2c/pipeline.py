from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from common.paths import data_path
from stages.step2c.age_normalizer import normalize_age_expression
from .config import RESULT_SHEET_NAME, RULES_DIR


RAW_HEADERS = [
    "Record ID",
    "Record Index",
    "PMID",
    "NCT ID",
    "Journal",
    "Year",
    "Indication",
    "Severity",
    "Intervention",
    "Intervention Type",
    "Comparator",
    "Outcome Direction",
    "Phase",
    "Sample Size",
    "Follow-up Time",
    "Age",
    "Gender",
    "Ethnicity",
    "Occupation",
    "Social Status",
    "Treatment History",
]

BASE_NORMALIZED_HEADERS = [
    "Record ID",
    "Record Index",
    "PMID",
    "NCT ID",
    "Journal",
    "Year",
    "Indication",
    "Intervention",
    "Intervention Type",
    "Comparator",
    "Outcome Direction",
    "Phase",
    "Sample Size",
    "Follow-up Time (Months)",
    "Severity",
    "Age [Mean_Age,SD_Age,Min_Age,Max_Age,N]",
    "Age Error",
    "Gender Male Proportion",
    "{ETHNICITY_VECTOR_HEADER}",
    "Occupation",
    "Social Status",
    "Treatment History",
]

STEP2A_REQUIRED_HEADERS = [
    "Record ID",
    "Record Index",
    "PMID",
    "NCT ID",
    "Journal",
    "Year",
    "Indication",
    "Population Raw",
    "Severity",
    "Population Status",
    "Target",
    "Intervention",
    "Intervention Type",
    "Comparator",
    "Outcome Direction",
    "Phase",
    "Sample Size",
    "Follow-up Time",
    "Evidence Snippet",
]

STEP2B_HEADERS = [
    "Record ID",
    "PMID",
    "Age",
    "Gender",
    "Ethnicity",
    "Occupation",
    "Social Status",
    "Treatment History",
]

TREATMENT_PATTERNS = [
    (r"\bpreviously untreated\b", "Previously untreated"),
    (r"\bcurrently untreated\b|\bnone currently receiving treatment\b|\bnot currently receiving treatment\b", "No current treatment"),
    (r"\buntreated\b", "Untreated"),
    (r"\bunmedicated\b|\bmedication[- ]free\b|\bdrug[- ]free\b|\bpsychotropic medication[- ]free\b", "Medication-free"),
    (r"\bno prior antidepressant treatment\b|\bantidepressant[- ]naive\b|\btreatment[- ]naive\b|\bno prior treatment\b", "Treatment-naive"),
    (r"\bprior antidepressant treatment\b|\bprior treatment\b|\bhistory of treatment\b", "Prior treatment"),
    (r"\btreatment[- ]free\b|\bantidepressant[- ]free\b|\boff medication\b|\bhitherto untreated\b", "Off treatment"),
]

SOCIAL_PATTERNS = [
    (r"\bmothers?\s+of\s+preterms?\b|\bmothers?\s+of\s+preterm\s+infants?\b", "Mothers of preterm infants"),
    (r"\bmothers?\s+of\s+infants?\b", "Mothers of infants"),
    (r"\bmothers?\s+of\s+preschool[- ]aged\s+children\b", "Mothers of preschool-aged children"),
    (r"\bmothers?\s+of\s+young\s+children\b", "Mothers of young children"),
    (r"\bmothers?\s+of\s+children\b", "Mothers of children"),
    (r"\bmothers?\s+of\s+adolescents?\b", "Mothers of adolescents"),
    (r"\bspouses?\s+or\s+parents?\s+of\s+deceased\b|\bspouses?\s+of\s+deceased\b|\bparents?\s+of\s+deceased\b", "Bereaved"),
    (r"\bparents?\s+of\s+preterm\s+infants?\b", "Parents of preterm infants"),
    (r"\bparents?\s+of\s+infants?\b", "Parents of infants"),
    (r"\bparents?\s+of\s+preschool[- ]aged\s+children\b", "Parents of preschool-aged children"),
    (r"\bparents?\s+of\s+young\s+children\b", "Parents of young children"),
    (r"\bparents?\s+of\s+children\b", "Parents of children"),
    (r"\bparents?\s+of\s+adolescents?\b", "Parents of adolescents"),
    (r"\b(?:living with\s+)?hiv(?:/aids)?\b|\bplwha\b|\bhiv[- ]positive\b|\bhiv[- ]infected\b", "HIV-positive"),
    (r"\bpregnant\b|\bpregnancy\b", "Pregnant"),
    (r"\bpostpartum\b|\bpost[- ]birth\b", "Postpartum"),
    (r"\bperinatal\b", "Perinatal"),
    (r"\bperi[- ]menopausal\b|\bpost[- ]menopausal\b|\bmenopausal\b", "Menopausal"),
    (r"\bcaregivers?\b", "Caregivers"),
    (r"\bnew mothers?\b", "New mothers"),
    (r"\bveterans?\b", "Veterans"),
    (r"\bactive duty\b", "Active duty military"),
    (r"\brural\b", "Rural"),
    (r"\burban\b", "Urban"),
    (r"\blow[- ]income\b|\bpoverty\b|\bpoor\b|\beconomically disadvantaged\b", "Low-income"),
    (r"\bhomeless\b", "Homeless"),
    (r"\bincarcerated\b", "Incarcerated"),
    (r"\bimmigrants?\b|\bmigrants?\b|\brefugees?\b", "Immigrant or refugee status"),
    (r"\bmuslim\b", "Muslim"),
    (r"\bbereaved\b", "Bereaved"),
    (r"\bdivorcees?\b", "Divorced"),
    (r"\borphaned\b", "Orphaned"),
    (r"\bhelp[- ]seeking\b", "Help-seeking"),
    (r"\binjection drug use\b", "History of injection drug use"),
    (r"\balcohol(?: use| misuse| abuse| dependence)?\b", "Alcohol use history"),
    (r"\bsubstance use(?: disorder)?\b", "Substance use history"),
    (r"\bsmokers?\b|\bsmoking\b", "Smoking history"),
    (r"\bmethamphetamine(?: use| users?)?\b", "Methamphetamine use history"),
]

ETHNICITY_PATTERNS = [
    (r"\bnon[- ]hispanic\s+white\b", "White"),
    (r"\bnon[- ]hispanic\s+black\b", "Black"),
    (r"\bafrican[- ]american\b", "Black"),
    (r"\bblack\b", "Black"),
    (r"\bwhite\b|\bcaucasian\b", "White"),
    (r"\bhan chinese\b|\bchinese han\b", "Asian"),
    (r"\basian\b", "Asian"),
    (r"(?<!non[- ])\bhispanic\b|(?<!non[- ])\blatino\b|(?<!non[- ])\blatina\b", "Hispanic/Latino"),
    (r"\bpacific islanders?\b", "Pacific Islander"),
    (r"\bindigenous\b|\bnative american\b", "Indigenous"),
    (r"\bmixed\b|\bmultiracial\b", "Mixed"),
    (r"\bother\b(?=\s*(?:race|ethnicity|racial|ethnic|$|[+/,;|]))", "Other"),
]

OCCUPATION_PATTERNS = [
    (r"\b(?:schoolchildren|high school students?|college students?|university students?|undergraduates?|students?)\b", "Student"),
    (r"\bemployees?\b|\bworkers?\b", "Employee"),
    (r"\bnurses?\b", "Nurse"),
    (r"\bteachers?\b", "Teacher"),
    (r"\bphysicians?\b|\bdoctors?\b|\bclinicians?\b", "Clinician"),
    (r"\bcaregivers?\b", "Caregiver"),
]

OTHER_DISEASE_PATTERN = (
    r"\b(?:cancer|diabetes|obesity|fracture|osa|sleep apnea|copd|parkinson|traumatic brain injury|chronic pain|insomnia|"
    r"hiv|aids|alcohol use disorder|substance use disorder|methamphetamine use)\b"
)


def clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_ascii(text: str) -> str:
    return (
        clean_text(text)
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2212", "-")
        .replace("\u00b1", "+/-")
    )


def dedupe_keep_order(items: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    ordered: List[str] = []
    for item in items:
        value = clean_text(item)
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(value)
    return ordered


def split_fragments(*texts: str) -> List[str]:
    fragments: List[str] = []
    for text in texts:
        cleaned = clean_text(text)
        if not cleaned:
            continue
        fragments.extend(part.strip() for part in re.split(r"\s*\|\|\s*|;", cleaned) if part.strip())
    return dedupe_keep_order(fragments)


def merge_labels(labels: Iterable[str], limit: int = 120) -> str:
    merged = " || ".join(dedupe_keep_order(list(labels)))
    if len(merged) <= limit:
        return merged
    return merged[: limit - 3].rstrip(" ,;:-") + "..."


def format_number(value: float | None, precision: int = 2) -> str:
    if value is None:
        return ""
    return f"{float(value):.{precision}f}"


def format_age_number(value: float | None) -> str:
    if value is None:
        return "0"
    number = float(value)
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def format_list_or_scalar(values: Sequence[int]) -> str:
    unique = []
    seen = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        unique.append(value)
    if not unique:
        return ""
    if len(unique) == 1:
        return str(unique[0])
    return "[" + ",".join(str(value) for value in unique) + "]"


def write_workbook(path: str, headers: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RESULT_SHEET_NAME
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_workbook_rows(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook[RESULT_SHEET_NAME] if RESULT_SHEET_NAME in workbook.sheetnames else workbook.active
        headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows: List[Dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in {None, ""} for value in values):
                continue
            payload: Dict[str, str] = {}
            for index, header in enumerate(headers):
                payload[header] = clean_text(values[index] if index < len(values) else "")
            rows.append(payload)
        return headers, rows
    finally:
        workbook.close()


def load_json_rule(name: str) -> Dict[str, Any]:
    return json.loads((RULES_DIR / name).read_text(encoding="utf-8"))


def extract_labels(text: str, patterns: Sequence[Tuple[str, str]]) -> List[str]:
    hits: List[str] = []
    cleaned = normalize_ascii(text).lower()
    if not cleaned:
        return hits
    for pattern, label in patterns:
        if re.search(pattern, cleaned, flags=re.I):
            hits.append(label)
    return dedupe_keep_order(hits)


def normalize_treatment_history_text(*texts: str) -> str:
    labels: List[str] = []
    for fragment in split_fragments(*texts):
        labels.extend(extract_labels(fragment, TREATMENT_PATTERNS))
    return merge_labels(labels)


def normalize_social_status_text(*texts: str) -> str:
    labels: List[str] = []
    for fragment in split_fragments(*texts):
        labels.extend(extract_labels(fragment, SOCIAL_PATTERNS))
    return merge_labels(labels)


def normalize_occupation_text(text: str) -> str:
    labels = extract_labels(text, OCCUPATION_PATTERNS)
    if not labels:
        raw = clean_text(text)
        if raw and len(raw) <= 80 and not re.search(r"\b(pregnan|postpartum|hiv|low-income|untreated|medication-free)\b", raw, flags=re.I):
            labels = [raw]
    return merge_labels(labels, limit=100)


def extract_step2a_treatment_history(row: Dict[str, str]) -> str:
    return normalize_treatment_history_text(row.get("Treatment History", ""), row.get("Population Raw", ""))


def insert_after(headers: Sequence[str], target: str, new_header: str) -> List[str]:
    output = list(headers)
    if new_header in output:
        return output
    if target in output:
        index = output.index(target) + 1
        output.insert(index, new_header)
    else:
        output.append(new_header)
    return output


def upgrade_step2a_rows(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    upgraded: List[Dict[str, str]] = []
    for row in rows:
        payload = {header: clean_text(row.get(header, "")) for header in STEP2A_REQUIRED_HEADERS}
        payload["Treatment History"] = clean_text(row.get("Treatment History", "")) or extract_step2a_treatment_history(row)
        upgraded.append(payload)
    return upgraded


def upgrade_step2b_rows(rows: Sequence[Dict[str, str]], step2a_by_record_id: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    upgraded: List[Dict[str, str]] = []
    for row in rows:
        record_id = clean_text(row.get("Record ID", ""))
        step2a_row = step2a_by_record_id.get(record_id, {})
        population_helper = " || ".join(
            dedupe_keep_order(
                [
                    clean_text(step2a_row.get("Population Raw", "")),
                    clean_text(row.get("Social Status", "")),
                ]
            )
        )
        social_status = normalize_social_status_text(row.get("Social Status", ""), population_helper)
        treatment_history = normalize_treatment_history_text(
            row.get("Treatment History", ""),
            row.get("Social Status", ""),
            step2a_row.get("Treatment History", ""),
            step2a_row.get("Population Raw", ""),
        )
        payload = {
            "Record ID": record_id,
            "PMID": clean_text(row.get("PMID", "")) or clean_text(step2a_row.get("PMID", "")),
            "Age": clean_text(row.get("Age", "")),
            "Gender": clean_text(row.get("Gender", "")),
            "Ethnicity": clean_text(row.get("Ethnicity", "")),
            "Occupation": normalize_occupation_text(row.get("Occupation", "")),
            "Social Status": social_status,
            "Treatment History": treatment_history,
        }
        upgraded.append(payload)
    return upgraded


def copy_population_rescue_manifest_if_present(step2a_input: str, step2a_output: str) -> str:
    input_path = Path(step2a_input)
    output_path = Path(step2a_output)
    candidates = [input_path.with_name(f"{input_path.stem}_population_rescue.json")]
    destination = output_path.with_name(f"{output_path.stem}_population_rescue.json")
    for candidate in candidates:
        if candidate.exists():
            if candidate.resolve() == destination.resolve():
                return str(destination)
            shutil.copyfile(candidate, destination)
            return str(destination)
    return ""


def build_raw_rows(step2a_rows: Sequence[Dict[str, str]], step2b_by_record_id: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    merged: List[Dict[str, str]] = []
    for row in step2a_rows:
        record_id = clean_text(row.get("Record ID", ""))
        step2b_row = step2b_by_record_id.get(record_id, {})
        merged.append(
            {
                "Record ID": record_id,
                "Record Index": clean_text(row.get("Record Index", "")),
                "PMID": clean_text(row.get("PMID", "")),
                "NCT ID": clean_text(row.get("NCT ID", "")),
                "Journal": clean_text(row.get("Journal", "")),
                "Year": clean_text(row.get("Year", "")),
                "Indication": clean_text(row.get("Indication", "")),
                "Severity": clean_text(row.get("Severity", "")),
                "Intervention": clean_text(row.get("Intervention", "")),
                "Intervention Type": clean_text(row.get("Intervention Type", "")),
                "Comparator": clean_text(row.get("Comparator", "")),
                "Outcome Direction": clean_text(row.get("Outcome Direction", "")),
                "Phase": clean_text(row.get("Phase", "")),
                "Sample Size": clean_text(row.get("Sample Size", "")),
                "Follow-up Time": clean_text(row.get("Follow-up Time", "")),
                "Age": clean_text(step2b_row.get("Age", "")),
                "Gender": clean_text(step2b_row.get("Gender", "")),
                "Ethnicity": clean_text(step2b_row.get("Ethnicity", "")),
                "Occupation": clean_text(step2b_row.get("Occupation", "")),
                "Social Status": clean_text(step2b_row.get("Social Status", "")),
                "Treatment History": clean_text(step2b_row.get("Treatment History", "")) or clean_text(row.get("Treatment History", "")),
            }
        )
    return merged


def normalize_sample_size_value(text: str) -> str:
    cleaned = clean_text(text)
    if not cleaned:
        return ""
    direct = re.sub(r"[^\d]", "", cleaned)
    return str(int(direct)) if direct else ""


def normalize_follow_up_months(text: str, follow_rule: Dict[str, Any]) -> str:
    cleaned = normalize_ascii(text).lower()
    if not cleaned:
        return ""
    factors = follow_rule.get("conversion", {})
    days_per_month = float(factors.get("days_per_month", 30.0))
    weeks_per_month = float(factors.get("weeks_per_month", 4.0))
    months_per_year = float(factors.get("months_per_year", 12.0))
    matches = re.findall(r"(\d+(?:\.\d+)?(?:\s*(?:to|-)\s*\d+(?:\.\d+)?)?)\s*(day|days|week|weeks|month|months|year|years)", cleaned, flags=re.I)
    if not matches:
        return ""
    values: List[float] = []
    for amount_text, unit in matches:
        amount_raw = amount_text.split("to")[-1] if "to" in amount_text else amount_text.split("-")[-1]
        amount = float(clean_text(amount_raw))
        unit_lower = unit.lower()
        if unit_lower.startswith("day"):
            values.append(amount / days_per_month)
        elif unit_lower.startswith("week"):
            values.append(amount / weeks_per_month)
        elif unit_lower.startswith("month"):
            values.append(amount)
        elif unit_lower.startswith("year"):
            values.append(amount * months_per_year)
    if not values:
        return ""
    return format_number(max(values))


def normalize_severity_code(text: str, severity_rule: Dict[str, Any]) -> str:
    cleaned = normalize_ascii(text).lower()
    if not cleaned:
        return ""
    code_map = severity_rule.get("codes", {})
    hits: List[int] = []
    phrase_rules = severity_rule.get("phrase_rules", [])
    for rule in phrase_rules:
        pattern = rule.get("pattern", "")
        codes = rule.get("codes", [])
        if pattern and re.search(pattern, cleaned, flags=re.I):
            hits.extend(int(code) for code in codes)
    if not hits:
        scale_rules = severity_rule.get("scale_rules", [])
        for rule in scale_rules:
            pattern = rule.get("pattern", "")
            codes = rule.get("codes", [])
            if pattern and re.search(pattern, cleaned, flags=re.I):
                hits.extend(int(code) for code in codes)
                break
    if not hits:
        for label, code in code_map.items():
            if re.search(rf"\b{re.escape(label)}\b", cleaned, flags=re.I):
                hits.append(int(code))
    hits = sorted(set(hits), reverse=True)
    return format_list_or_scalar(hits) if hits else clean_text(text)


def normalize_gender_male_proportion(text: str) -> str:
    cleaned = normalize_ascii(text).lower()
    if not cleaned:
        return ""
    male_match = re.search(r"(\d+(?:\.\d+)?)\s*%\s*male", cleaned, flags=re.I)
    female_match = re.search(r"(\d+(?:\.\d+)?)\s*%\s*female", cleaned, flags=re.I)
    if male_match:
        return format_number(float(male_match.group(1)) / 100.0)
    if female_match:
        return format_number(1.0 - (float(female_match.group(1)) / 100.0))
    if re.fullmatch(r"male", cleaned, flags=re.I):
        return format_number(1.0)
    if re.fullmatch(r"female", cleaned, flags=re.I):
        return format_number(0.0)
    return ""


def classify_ethnicity_label(text: str) -> str:
    cleaned = normalize_ascii(text).lower()
    for pattern, label in ETHNICITY_PATTERNS:
        if re.search(pattern, cleaned, flags=re.I):
            return label
    return ""


def parse_ethnicity_distribution(text: str) -> Dict[str, float]:
    cleaned = normalize_ascii(text)
    if not cleaned:
        return {}
    segments = re.split(r"\s*\+\s*|\s*\|\|\s*|;", cleaned)
    weighted: Dict[str, float] = {}
    unweighted: List[str] = []
    for segment in segments:
        label = classify_ethnicity_label(segment)
        if not label:
            continue
        pct_match = re.search(r"(\d+(?:\.\d+)?)\s*%", segment)
        if pct_match:
            weighted[label] = weighted.get(label, 0.0) + float(pct_match.group(1)) / 100.0
        else:
            unweighted.append(label)
    if not weighted and not unweighted:
        label = classify_ethnicity_label(cleaned)
        if label:
            return {label: 1.0}
        return {}
    if not weighted and unweighted:
        share = 1.0 / len(unweighted)
        for label in unweighted:
            weighted[label] = weighted.get(label, 0.0) + share
    elif unweighted:
        remaining = max(0.0, 1.0 - sum(weighted.values()))
        share = remaining / len(unweighted) if remaining > 0 else 0.0
        for label in unweighted:
            weighted[label] = weighted.get(label, 0.0) + share
    total = sum(weighted.values())
    if total > 0:
        weighted = {label: value / total for label, value in weighted.items()}
    return weighted


def build_ethnicity_codebook(rows: Sequence[Dict[str, str]]) -> Tuple[Dict[str, str], List[Dict[str, Any]]]:
    counter: Counter[str] = Counter()
    for row in rows:
        for label in parse_ethnicity_distribution(row.get("Ethnicity", "")).keys():
            counter[label] += 1
    labels = [label for label, _count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))]
    top6_codes = {label: code for label, code in zip(labels[:6], list("ABCDEF"))}
    codebook_rows: List[Dict[str, Any]] = []
    for index, label in enumerate(labels, start=1):
        codebook_rows.append(
            {
                "numeric_id": index,
                "normalized_label": label,
                "frequency": counter[label],
                "top6_code": top6_codes.get(label, ""),
            }
        )
    return top6_codes, codebook_rows


def ethnicity_vector_header(top6_codes: Dict[str, str]) -> str:
    label_by_code = {code: label for label, code in top6_codes.items()}
    slots = [f"{code}={label_by_code.get(code, 'Unassigned')}" for code in "ABCDEF"]
    return "Ethnicity [" + ",".join(slots) + "]"


def normalized_headers(top6_codes: Dict[str, str]) -> List[str]:
    ethnicity_header = ethnicity_vector_header(top6_codes)
    return [ethnicity_header if header == "{ETHNICITY_VECTOR_HEADER}" else header for header in BASE_NORMALIZED_HEADERS]


def normalize_occupation_labels(text: str) -> List[str]:
    normalized = normalize_occupation_text(text)
    return dedupe_keep_order(normalized.split("||")) if normalized else []


def build_occupation_codebook(rows: Sequence[Dict[str, str]]) -> Tuple[Dict[str, int], List[Dict[str, Any]]]:
    counter: Counter[str] = Counter()
    for row in rows:
        for label in normalize_occupation_labels(row.get("Occupation", "")):
            counter[label] += 1
    labels = [label for label, _count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))]
    mapping = {label: index for index, label in enumerate(labels, start=1)}
    codebook_rows = [
        {"occupation_id": mapping[label], "normalized_label": label, "frequency": counter[label]}
        for label in labels
    ]
    return mapping, codebook_rows


def social_status_codes(text: str, social_rule: Dict[str, Any]) -> str:
    combined = clean_text(text).lower()
    if not combined:
        return ""
    hits: List[int] = []
    if re.search(r"\bpregnan|postpartum|perinatal|new mothers?\b", combined, flags=re.I):
        hits.append(1)
    if re.search(r"\balcohol|substance use|injection drug use|methamphetamine|smoking|smokers?\b", combined, flags=re.I):
        hits.append(2)
    if re.search(r"\bhiv|aids|plwha\b", combined, flags=re.I):
        hits.append(3)
    if re.search(r"\blow[- ]income|poverty|economically disadvantaged|homeless\b", combined, flags=re.I):
        hits.append(4)
    if 3 not in hits and re.search(OTHER_DISEASE_PATTERN, combined, flags=re.I):
        hits.append(5)
    if not hits:
        hits.append(6)
    return format_list_or_scalar(sorted(hits))


def treatment_history_codes(text: str, rule: Dict[str, Any]) -> Tuple[str, str]:
    labels = extract_labels(text, TREATMENT_PATTERNS)
    normalized = merge_labels(labels)
    code_map = {label: int(code) for code, label in rule.get("codes", {}).items()}
    codes = [code_map[label] for label in labels if label in code_map]
    return normalized, format_list_or_scalar(sorted(codes))


def age_normalization_bundle(text: str) -> Dict[str, Any]:
    result = normalize_age_expression(text)
    row = result.to_row()
    detected_type = clean_text(row.get("age_detected_type", ""))
    source_kind = clean_text(row.get("age_source_kind", ""))
    norm_method = clean_text(row.get("age_norm_method", "")).lower()
    mean_age: float | None = None
    sd_age: float | None = None
    min_age: float | None = None
    max_age: float | None = None
    age_n: float | None = None

    if detected_type == "numeric_range":
        min_age = row.get("age_reported_lower_years") or None
        max_age = row.get("age_reported_upper_years") or None
    elif detected_type == "single_value":
        mean_age = row.get("age_reported_center_years") or None
    elif detected_type == "mean_sd":
        mean_age = row.get("age_reported_center_years") or None
        sd_age = row.get("age_reported_scale_years") or None
    elif detected_type in {"mean", "median"}:
        mean_age = row.get("age_reported_center_years") or None
        sd_age = row.get("age_reported_scale_years") or None
        if "with_range" in norm_method:
            min_age = row.get("age_reported_lower_years") or None
            max_age = row.get("age_reported_upper_years") or None
    elif detected_type == "lower_bound":
        min_age = row.get("age_reported_lower_years") or None
    elif detected_type == "upper_bound":
        max_age = row.get("age_reported_upper_years") or None
    elif source_kind == "mapped_stage":
        min_age = row.get("age_inferred_lower_years") or None
        max_age = row.get("age_inferred_upper_years") or None

    cleaned = normalize_ascii(text).lower()
    max_age_for_error = row.get("age_reported_upper_years") or row.get("age_inferred_upper_years") or ""
    min_age_for_error = row.get("age_reported_lower_years") or row.get("age_inferred_lower_years") or ""
    under_one = any(
        [
            re.search(r"\binfants?\b|\bnewborns?\b|\bneonates?\b", cleaned, flags=re.I),
            max_age_for_error != "" and float(max_age_for_error) < 1.0,
            min_age_for_error != "" and float(min_age_for_error) < 1.0 and max_age_for_error != "" and float(max_age_for_error) <= 1.0,
        ]
    )
    vector = "[" + ",".join(
        [
            format_age_number(mean_age),
            format_age_number(sd_age),
            format_age_number(min_age),
            format_age_number(max_age),
            format_age_number(age_n),
        ]
    ) + "]"
    return {
        "Age [Mean_Age,SD_Age,Min_Age,Max_Age,N]": vector,
        "Age Error": "error" if under_one else "",
    }


def ethnicity_vector(text: str, top6_codes: Dict[str, str]) -> str:
    distribution = parse_ethnicity_distribution(text)
    slots = {code: 0.0 for code in "ABCDEF"}
    for label, proportion in distribution.items():
        code = top6_codes.get(label)
        if code:
            slots[code] += proportion
    ordered_values = [format_number(slots[code]) for code in "ABCDEF"]
    return "[" + ",".join(ordered_values) + "]"


def occupation_codes(text: str, mapping: Dict[str, int]) -> Tuple[str, str]:
    labels = normalize_occupation_labels(text)
    normalized = merge_labels(labels, limit=100)
    codes = [mapping[label] for label in labels if label in mapping]
    return normalized, format_list_or_scalar(codes)


def build_normalized_rows(raw_rows: Sequence[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    follow_rule = load_json_rule("follow_up_time_rule.json")
    severity_rule = load_json_rule("severity_codebook.json")
    social_rule = load_json_rule("social_status_codebook.json")
    top6_codes, ethnicity_codebook = build_ethnicity_codebook(raw_rows)
    occupation_mapping, occupation_codebook = build_occupation_codebook(raw_rows)
    headers = normalized_headers(top6_codes)
    ethnicity_header = ethnicity_vector_header(top6_codes)

    normalized_rows: List[Dict[str, Any]] = []
    age_error_count = 0
    for row in raw_rows:
        age_bundle = age_normalization_bundle(row.get("Age", ""))
        if age_bundle["Age Error"]:
            age_error_count += 1
        ethnicity_vector_value = ethnicity_vector(row.get("Ethnicity", ""), top6_codes)
        occupation_normalized, occupation_code = occupation_codes(row.get("Occupation", ""), occupation_mapping)
        normalized_rows.append(
            {
                "Record ID": row.get("Record ID", ""),
                "Record Index": row.get("Record Index", ""),
                "PMID": row.get("PMID", ""),
                "NCT ID": row.get("NCT ID", ""),
                "Journal": row.get("Journal", ""),
                "Year": row.get("Year", ""),
                "Indication": row.get("Indication", ""),
                "Intervention": row.get("Intervention", ""),
                "Intervention Type": row.get("Intervention Type", ""),
                "Comparator": row.get("Comparator", ""),
                "Outcome Direction": row.get("Outcome Direction", ""),
                "Phase": row.get("Phase", ""),
                "Sample Size": normalize_sample_size_value(row.get("Sample Size", "")),
                "Follow-up Time (Months)": normalize_follow_up_months(row.get("Follow-up Time", ""), follow_rule),
                "Severity": normalize_severity_code(row.get("Severity", ""), severity_rule),
                **age_bundle,
                "Gender Male Proportion": normalize_gender_male_proportion(row.get("Gender", "")),
                ethnicity_header: ethnicity_vector_value,
                "Occupation": occupation_normalized or occupation_code,
                "Social Status": social_status_codes(row.get("Social Status", ""), social_rule),
                "Treatment History": row.get("Treatment History", ""),
            }
        )

    summary = {
        "raw_row_count": len(raw_rows),
        "normalized_row_count": len(normalized_rows),
        "age_error_count": age_error_count,
        "ethnicity_codebook_size": len(ethnicity_codebook),
        "occupation_codebook_size": len(occupation_codebook),
        "ethnicity_top6_codes": top6_codes,
        "rules_dir": str(RULES_DIR),
    }
    return normalized_rows, {
        "summary": summary,
        "normalized_headers": headers,
        "ethnicity_codebook": ethnicity_codebook,
        "occupation_codebook": occupation_codebook,
    }


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def run_step2c(
    *,
    step2a_input: str,
    step2b_input: str,
    step2a_output: str,
    step2b_output: str,
    raw_output: str,
    normalized_output: str,
    summary_output: str,
) -> Dict[str, Any]:
    RULES_DIR.mkdir(parents=True, exist_ok=True)

    step2a_headers, step2a_rows = read_workbook_rows(step2a_input)
    upgraded_step2a_rows = upgrade_step2a_rows(step2a_rows)
    step2a_output_headers = insert_after(step2a_headers or STEP2A_REQUIRED_HEADERS, "Severity", "Treatment History")
    write_workbook(step2a_output, step2a_output_headers, upgraded_step2a_rows)
    rescue_copy = copy_population_rescue_manifest_if_present(step2a_input, step2a_output)

    step2a_by_record_id = {clean_text(row.get("Record ID", "")): row for row in upgraded_step2a_rows if clean_text(row.get("Record ID", ""))}

    _step2b_headers, step2b_rows = read_workbook_rows(step2b_input)
    upgraded_step2b_rows = upgrade_step2b_rows(step2b_rows, step2a_by_record_id)
    write_workbook(step2b_output, STEP2B_HEADERS, upgraded_step2b_rows)
    step2b_by_record_id = {clean_text(row.get("Record ID", "")): row for row in upgraded_step2b_rows if clean_text(row.get("Record ID", ""))}

    raw_rows = build_raw_rows(upgraded_step2a_rows, step2b_by_record_id)
    write_workbook(raw_output, RAW_HEADERS, raw_rows)

    normalized_rows, artifacts = build_normalized_rows(raw_rows)
    write_workbook(normalized_output, artifacts["normalized_headers"], normalized_rows)

    write_csv(
        RULES_DIR / "ethnicity_codebook.csv",
        ["numeric_id", "normalized_label", "frequency", "top6_code"],
        artifacts["ethnicity_codebook"],
    )
    write_csv(
        RULES_DIR / "occupation_codebook.csv",
        ["occupation_id", "normalized_label", "frequency"],
        artifacts["occupation_codebook"],
    )

    summary = {
        "step2a_input": str(Path(step2a_input).resolve()),
        "step2b_input": str(Path(step2b_input).resolve()),
        "step2a_output": str(Path(step2a_output).resolve()),
        "step2b_output": str(Path(step2b_output).resolve()),
        "raw_output": str(Path(raw_output).resolve()),
        "normalized_output": str(Path(normalized_output).resolve()),
        "population_rescue_copy": rescue_copy,
        **artifacts["summary"],
    }
    Path(summary_output).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary
