from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
import sys
from typing import Any, Callable, Dict, List, Tuple

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config import default_output_path, latest_step2a_output
from pipeline import (
    AGE_SIGNAL_PATTERN,
    ETHNICITY_PATTERNS,
    GENDER_SIGNAL_PATTERN,
    OCCUPATION_SIGNAL_PATTERN,
    SOCIAL_STATUS_SIGNAL_PATTERN,
    _build_rule_based_output,
    _population_source_text,
    load_input_rows,
)


FieldDetector = Callable[[str], bool]


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _clip_text(value: Any, limit: int = 180) -> str:
    text = _clean_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _load_step2b_rows(output_path: str) -> Dict[str, Dict[str, str]]:
    workbook = load_workbook(output_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: Dict[str, Dict[str, str]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        payload = dict(zip(headers, ["" if value is None else str(value).strip() for value in values]))
        record_id = _clean_text(payload.get("Record ID", ""))
        if record_id:
            rows[record_id] = payload
    return rows


def _has_ethnicity_signal(text: str) -> bool:
    low = _clean_text(text).lower()
    if not low:
        return False
    if any(re.search(pattern, low, flags=re.I) for pattern, _ in ETHNICITY_PATTERNS):
        return True
    return bool(
        re.search(
            r"\b\d+(?:\.\d+)?%\s*(?:african[- ]american|black|white|caucasian|asian|"
            r"hispanic|latino|latina|indigenous|native american|mixed|multiracial)\b",
            low,
            flags=re.I,
        )
    )


def _field_detectors() -> List[Tuple[str, str, FieldDetector]]:
    return [
        ("Age", "age", lambda text: bool(re.search(AGE_SIGNAL_PATTERN, _clean_text(text).lower(), flags=re.I))),
        (
            "Gender",
            "gender",
            lambda text: bool(re.search(GENDER_SIGNAL_PATTERN, _clean_text(text).lower(), flags=re.I)),
        ),
        ("Ethnicity", "ethnicity", _has_ethnicity_signal),
        (
            "Occupation",
            "occupation",
            lambda text: bool(re.search(OCCUPATION_SIGNAL_PATTERN, _clean_text(text).lower(), flags=re.I)),
        ),
        (
            "Social Status",
            "social_status",
            lambda text: bool(re.search(SOCIAL_STATUS_SIGNAL_PATTERN, _clean_text(text).lower(), flags=re.I)),
        ),
    ]


def _build_summary(step2a_path: str, step2b_path: str, *, example_limit: int) -> Dict[str, Any]:
    input_rows, sheet_name = load_input_rows(step2a_path)
    output_rows = _load_step2b_rows(step2b_path)

    provider_context = re.compile(
        r"\b(?:provided|delivered|administered|led|facilitated|conducted|trained)\b.{0,120}\bby\b",
        flags=re.I,
    )
    services_context = re.compile(r"\b(?:services?|advisor|providers?)\b", flags=re.I)
    age_numeric_signal = re.compile(
        r"\b(?:aged?\s+(?:between\s+)?\d|mean age|median age|"
        r"\d+(?:\.\d+)?\s*(?:-|to)\s*\d+(?:\.\d+)?\s*years?|"
        r"\d+(?:\.\d+)?\s*years?\s+or older)\b",
        flags=re.I,
    )

    field_metrics: Dict[str, Dict[str, Any]] = {}
    for column_name, attr_name, detector in _field_detectors():
        source_signal_rows = 0
        saved_non_empty = 0
        saved_signal_hits = 0
        current_rule_non_empty = 0
        current_rule_signal_hits = 0
        saved_vs_current_rule_diff = 0

        for input_row in input_rows:
            source_text = _population_source_text(input_row)
            has_signal = detector(source_text)
            if has_signal:
                source_signal_rows += 1

            saved_value = _clean_text(output_rows.get(input_row.record_id, {}).get(column_name, ""))
            current_rule_value = _clean_text(getattr(_build_rule_based_output(input_row), attr_name))

            if saved_value:
                saved_non_empty += 1
                if has_signal:
                    saved_signal_hits += 1

            if current_rule_value:
                current_rule_non_empty += 1
                if has_signal:
                    current_rule_signal_hits += 1

            if saved_value != current_rule_value:
                saved_vs_current_rule_diff += 1

        field_metrics[column_name] = {
            "source_signal_rows": source_signal_rows,
            "saved_step2b_non_empty": saved_non_empty,
            "saved_step2b_signal_recall_proxy": round(saved_signal_hits / source_signal_rows, 4)
            if source_signal_rows
            else None,
            "current_rule_non_empty": current_rule_non_empty,
            "current_rule_signal_recall_proxy": round(current_rule_signal_hits / source_signal_rows, 4)
            if source_signal_rows
            else None,
            "saved_vs_current_rule_diff_rows": saved_vs_current_rule_diff,
        }

    example_buckets: Dict[str, List[Dict[str, str]]] = {
        "age_student_identity_in_age": [],
        "age_numeric_signal_but_saved_empty": [],
        "gender_two_sex_signal_but_saved_output_partial": [],
        "ethnicity_other_output_suspicious": [],
        "occupation_missing_in_saved_step2b": [],
        "occupation_false_positive_risk": [],
        "social_status_missing_in_saved_step2b": [],
        "social_status_false_positive_risk": [],
    }
    example_counts = {key: 0 for key in example_buckets}

    for input_row in input_rows:
        source_text = _population_source_text(input_row)
        low_source = source_text.lower()
        saved_row = output_rows.get(input_row.record_id, {})
        saved_age = _clean_text(saved_row.get("Age", ""))
        saved_gender = _clean_text(saved_row.get("Gender", ""))
        saved_ethnicity = _clean_text(saved_row.get("Ethnicity", ""))
        saved_occupation = _clean_text(saved_row.get("Occupation", ""))
        saved_social_status = _clean_text(saved_row.get("Social Status", ""))
        current_rule = _build_rule_based_output(input_row)

        if current_rule.age in {"College students", "High school students", "Students", "Schoolchildren"} or re.fullmatch(
            r"\d{1,2}(?:st|nd|rd|th) grade students",
            current_rule.age or "",
            flags=re.I,
        ):
            if not current_rule.occupation:
                example_counts["age_student_identity_in_age"] += 1
                if len(example_buckets["age_student_identity_in_age"]) < example_limit:
                    example_buckets["age_student_identity_in_age"].append(
                        {
                            "record_id": input_row.record_id,
                            "current_rule_age": current_rule.age,
                            "source_text": _clip_text(source_text),
                        }
                    )

        if age_numeric_signal.search(low_source) and not saved_age:
            example_counts["age_numeric_signal_but_saved_empty"] += 1
            if len(example_buckets["age_numeric_signal_but_saved_empty"]) < example_limit:
                example_buckets["age_numeric_signal_but_saved_empty"].append(
                    {
                        "record_id": input_row.record_id,
                        "saved_age": saved_age,
                        "current_rule_age": current_rule.age,
                        "source_text": _clip_text(source_text),
                    }
                )

        male_signal = bool(re.search(r"\b(?:male|males|men|boys)\b", low_source))
        female_signal = bool(re.search(r"\b(?:female|females|women|girls)\b", low_source))
        if male_signal and female_signal and saved_gender:
            if "Male + Female" not in saved_gender and saved_gender.count("%") < 2:
                example_counts["gender_two_sex_signal_but_saved_output_partial"] += 1
                if len(example_buckets["gender_two_sex_signal_but_saved_output_partial"]) < example_limit:
                    example_buckets["gender_two_sex_signal_but_saved_output_partial"].append(
                        {
                            "record_id": input_row.record_id,
                            "saved_gender": saved_gender,
                            "source_text": _clip_text(source_text),
                        }
                    )

        current_ethnicity = _clean_text(current_rule.ethnicity)
        if saved_ethnicity.lower() == "other" or current_ethnicity.lower() == "other":
            if re.search(r"\bother\b", low_source) and not re.search(
                r"\b(?:african[- ]american|black|white|caucasian|asian|hispanic|latino|latina|"
                r"indigenous|native american|mixed|multiracial)\b",
                low_source,
                flags=re.I,
            ):
                example_counts["ethnicity_other_output_suspicious"] += 1
                if len(example_buckets["ethnicity_other_output_suspicious"]) < example_limit:
                    example_buckets["ethnicity_other_output_suspicious"].append(
                        {
                            "record_id": input_row.record_id,
                            "saved_ethnicity": saved_ethnicity,
                            "current_rule_ethnicity": current_ethnicity,
                            "source_text": _clip_text(source_text),
                        }
                    )

        if current_rule.occupation and not saved_occupation:
            example_counts["occupation_missing_in_saved_step2b"] += 1
            if len(example_buckets["occupation_missing_in_saved_step2b"]) < example_limit:
                example_buckets["occupation_missing_in_saved_step2b"].append(
                    {
                        "record_id": input_row.record_id,
                        "current_rule_occupation": current_rule.occupation,
                        "source_text": _clip_text(source_text),
                    }
                )

        if current_rule.occupation and (
            provider_context.search(source_text) or services_context.search(source_text) or "without a doctor" in low_source
        ):
            example_counts["occupation_false_positive_risk"] += 1
            if len(example_buckets["occupation_false_positive_risk"]) < example_limit:
                example_buckets["occupation_false_positive_risk"].append(
                    {
                        "record_id": input_row.record_id,
                        "current_rule_occupation": current_rule.occupation,
                        "source_text": _clip_text(source_text),
                    }
                )

        if current_rule.social_status and not saved_social_status:
            example_counts["social_status_missing_in_saved_step2b"] += 1
            if len(example_buckets["social_status_missing_in_saved_step2b"]) < example_limit:
                example_buckets["social_status_missing_in_saved_step2b"].append(
                    {
                        "record_id": input_row.record_id,
                        "current_rule_social_status": current_rule.social_status,
                        "source_text": _clip_text(source_text),
                    }
                )

        if current_rule.social_status and re.search(
            r"\b(?:Alcohol use|Substance use|Smoking|Methamphetamine|Rural|Urban)\b",
            current_rule.social_status,
            flags=re.I,
        ):
            example_counts["social_status_false_positive_risk"] += 1
            if len(example_buckets["social_status_false_positive_risk"]) < example_limit:
                example_buckets["social_status_false_positive_risk"].append(
                    {
                        "record_id": input_row.record_id,
                        "current_rule_social_status": current_rule.social_status,
                        "source_text": _clip_text(source_text),
                    }
                )

    return {
        "step2a_input_path": str(Path(step2a_path).resolve()),
        "step2a_sheet_name": sheet_name,
        "step2b_output_path": str(Path(step2b_path).resolve()),
        "row_counts": {
            "step2a_rows": len(input_rows),
            "step2b_rows": len(output_rows),
        },
        "field_metrics": field_metrics,
        "example_bucket_counts": example_counts,
        "likely_bad_examples_to_check_first": example_buckets,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit Step2B outputs against real Step2A inputs.")
    parser.add_argument("--step2a", default=latest_step2a_output(), help="Path to the Step2A workbook.")
    parser.add_argument("--step2b", default=default_output_path(), help="Path to the Step2B workbook.")
    parser.add_argument("--example-limit", type=int, default=6, help="Examples to keep per bucket.")
    parser.add_argument(
        "--json-out",
        default="",
        help="Optional path to write the audit JSON. Prints to stdout when omitted.",
    )
    args = parser.parse_args()

    summary = _build_summary(args.step2a, args.step2b, example_limit=max(1, args.example_limit))
    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    if args.json_out:
        output_path = Path(args.json_out)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(payload, encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
