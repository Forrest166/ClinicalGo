from __future__ import annotations

import json
import re
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from common.api_resilience import is_rate_limit_error, should_split_batch
from common.request_stability import AdaptiveRateLimitController
from common.structured_llm import UserCancelledError, build_runtime_settings
from common.text_utils import normalize_ascii_key
from config import DEFAULT_SHEET_NAME, OUTPUT_HEADERS
from llm_client import Step2BExtractionClient
from models import Step2BInputRow, Step2BOutputRow


class Step2BError(Exception):
    pass


HEADER_ALIASES = {
    "record_id": {"recordid", "record id", "record_id"},
    "pmid": {"pmid"},
    "population_raw": {"populationraw", "population raw", "population_raw"},
    "treatment_history": {"treatmenthistory", "treatment history", "treatment_history"},
}

ETHNICITY_PATTERNS = [
    (r"\bnon[- ]hispanic\s+white\b", "White"),
    (r"\bnon[- ]hispanic\s+black\b", "Black"),
    (r"\bafrican[- ]american\b", "Black"),
    (r"\bblack\b", "Black"),
    (r"\bwhite\b|\bcaucasian\b", "White"),
    (r"\bhan chinese\b|\bchinese han\b", "Asian"),
    (r"\basian\b", "Asian"),
    (r"(?<!non[- ])\bhispanic\b|(?<!non[- ])\blatino\b|(?<!non[- ])\blatina\b", "Hispanic/Latino"),
    (r"\bpacific islanders?\b", "Pacific Islander"),
    (r"\bindigenous\b|\bnative american\b", "Indigenous"),
    (r"\bmixed\b|\bmultiracial\b", "Mixed"),
    (r"\bother\b(?=\s*(?:race|ethnicity|racial|ethnic|$|[+/,;|]))", "Other"),
]

SCALE_NAME_PATTERN = r"(?:GRID-)?(?:HAM-?D|HDRS|HRSD|MADRS|PHQ-?9|PHQ-?8|BDI(?:-?II)?|CGI(?:-?[SI])?|CSDD|QIDS(?:-SR)?|IDS(?:-SR)?)"
AGE_SIGNAL_PATTERN = (
    r"\b(?:age(?:d|s)?\s+(?:between\s+)?\d+(?:\.\d+)?|aged?\s+\d+(?:\.\d+)?\+|aged?\s*(?:>=|=>|<=|=<|>|<)\s*\d+(?:\.\d+)?(?:\s*years?)?|"
    r"age\s*(?:>=|=>|<=|=<|>|<)\s*\d+(?:\.\d+)?(?:\s*years?)?|mean age|median age|"
    r"at least\s+\d+(?:\.\d+)?\s*years?\s+old|"
    r"\d+(?:\.\d+)?\+|"
    r"aged?\s+\d+(?:\.\d+)?\s+(?:or|and)\s+(?:older|over|above|more)|"
    r"\d+(?:\.\d+)?\s*years?\s+(?:or|and)\s+(?:older|over|above|more)|"
    r"\d+(?:\.\d+)?\s*(?:-|to)\s*\d+(?:\.\d+)?\s*years?|"
    r"\d+(?:\.\d+)?\s*-\s*to\s*\d+(?:\.\d+)?-year-olds?|"
    r"\d+(?:\.\d+)?\s*years?\s+or older|"
    r"older adults?|young adults?|adolescents?|children|schoolchildren|elderly|geriatric|late-life|"
    r"college students|university students|high school students|undergraduates?)\b"
)
GENDER_SIGNAL_PATTERN = r"\b(?:male|males|female|females|men|women|boys|girls)\b|\b\d+(?:\.\d+)?%\s*(?:male|female|men|women)\b"
OCCUPATION_SIGNAL_PATTERN = (
    r"\b(?:students?|schoolchildren|workers?|employees?|nurses?|teachers?|physicians?|doctors?|clinicians?)\b"
)
SOCIAL_STATUS_SIGNAL_PATTERN = (
    r"\b(?:pregnant|pregnancy|postpartum|perinatal|peri[- ]menopausal|post[- ]menopausal|"
    r"peri[- ]?and\s+post[- ]?menopausal|menopausal|hiv(?:/aids)?|hiv[- ]positive|hiv[- ]infected|"
    r"caregivers?|mothers?\s+of|parents?\s+of|new mothers?|refugees?|immigrants?|migrants?|incarcerated|homeless|"
    r"low[- ]income|poverty|economically disadvantaged|rural|urban|substance use|alcohol use|smokers?|"
    r"veterans?|muslim|bereaved|divorcees?|post-birth|orphaned|help[- ]seeking|active duty)\b"
)
TREATMENT_HISTORY_SIGNAL_PATTERN = (
    r"\b(?:previously untreated|currently untreated|untreated|unmedicated|medication[- ]free|drug[- ]free|"
    r"treatment[- ]naive|antidepressant[- ]naive|no prior antidepressant treatment|"
    r"none currently receiving treatment|not currently receiving treatment|"
    r"treatment[- ]free|antidepressant[- ]free|off medication|prior antidepressant treatment|"
    r"prior treatment|history of treatment|no prior treatment|no history of treatment|"
    r"psychotropic medication[- ]free|hitherto untreated)\b"
)
RISK_SIGNAL_PATTERN = (
    r"\b(?:at risk(?:\s+for)?|high risk(?:\s+for)?|low risk(?:\s+for)?|moderate risk(?:\s+for)?|"
    r"familial risk(?:\s+for)?|individual risk(?:\s+for)?|parental depression|"
    r"subsyndromal symptoms?|clinically relevant depressive symptoms?|elevated depressive symptoms?)\b"
)
STUDY_NOISE_PATTERN = (
    r"\b(?:randomized|randomised|trial|study|intervention|program|session|sessions|group|groups|"
    r"provided by|delivered by|compared with|control group|outcome|results?|treatment)\b"
)
SAMPLE_COUNT_PATTERN = (
    r"\b(?:N\s*=\s*\d{1,5}(?:,\d{3})*|"
    r"\d{1,3}(?:,\d{3})*\s+(?:participants?|patients?|subjects?|individuals?|adults?|women|men|"
    r"children|adolescents|outpatients?|inpatients?|completers?))\b"
)


class Step2BCheckpointStore:
    def __init__(self, checkpoint_path: str) -> None:
        self.path = Path(checkpoint_path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def append(self, item: Dict[str, Any]) -> None:
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(item, ensure_ascii=False) + "\n")
            handle.flush()

    def reset(self, *, input_path: str, sheet_name: str, provider: str, model: str) -> None:
        if self.path.exists():
            self.path.unlink()
        self.append(
            {
                "type": "meta",
                "input_path": str(Path(input_path).resolve()),
                "sheet_name": sheet_name,
                "provider": provider,
                "model": model,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    def read_meta(self) -> Dict[str, Any]:
        if not self.path.exists():
            return {}
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                try:
                    payload = json.loads(line)
                except Exception:
                    continue
                if payload.get("type") == "meta":
                    return payload
        return {}

    def append_row(self, row: Step2BOutputRow) -> None:
        self.append({"type": "row", "record_id": row.record_id, "data": row.to_export_dict()})

    def append_usage(self, usage: Dict[str, Any], completed_record_ids: Sequence[str]) -> None:
        self.append({"type": "usage", "completed_record_ids": list(completed_record_ids), "data": dict(usage)})


def load_checkpoint_progress(checkpoint_path: str) -> Tuple[set[str], int, int]:
    path = Path(checkpoint_path)
    if not path.exists():
        return set(), 0, 0
    completed: set[str] = set()
    row_count = 0
    usage_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                payload = json.loads(line)
            except Exception:
                continue
            if payload.get("type") == "row":
                record_id = str(payload.get("record_id", "")).strip()
                if record_id:
                    completed.add(record_id)
                row_count += 1
            elif payload.get("type") == "usage":
                usage_count += 1
                for record_id in payload.get("completed_record_ids", []) or []:
                    if str(record_id).strip():
                        completed.add(str(record_id).strip())
    return completed, row_count, usage_count


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _dedupe_keep_order(items: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    ordered: List[str] = []
    for item in items:
        cleaned = _clean_text(item)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(cleaned)
    return ordered


def _clip_text(value: Any, limit: int = 120) -> str:
    text = _clean_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _format_percent(value: float) -> str:
    rounded = round(float(value), 1)
    if abs(rounded - round(rounded)) < 1e-9:
        return f"{int(round(rounded))}%"
    return f"{rounded:.1f}%"


def _normalize_percent(value: str) -> str:
    text = _clean_text(value).replace("%", "")
    if not text:
        return ""
    try:
        number = float(text)
    except Exception:
        return ""
    if 0 <= number <= 1:
        number *= 100
    if number < 0 or number > 100:
        return ""
    return _format_percent(number)


def _infer_gender_percentages(source_text: str) -> Tuple[str, str]:
    text = str(source_text or "")
    for pattern in [
        r"(?P<male>\d{1,5})\s*(?:male|males|men|boys)\b[^\d]{0,40}(?P<female>\d{1,5})\s*(?:female|females|women|girls)\b",
        r"(?P<female>\d{1,5})\s*(?:female|females|women|girls)\b[^\d]{0,40}(?P<male>\d{1,5})\s*(?:male|males|men|boys)\b",
    ]:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue
        male = float(match.group("male"))
        female = float(match.group("female"))
        total = male + female
        if total > 0:
            return _format_percent((male / total) * 100.0), _format_percent((female / total) * 100.0)
    male_pct = re.search(r"(\d{1,3}(?:\.\d+)?)\s*%\s*(?:male|males|men|boys)\b", text, flags=re.I)
    female_pct = re.search(r"(\d{1,3}(?:\.\d+)?)\s*%\s*(?:female|females|women|girls)\b", text, flags=re.I)
    return (male_pct.group(1) + "%" if male_pct else "", female_pct.group(1) + "%" if female_pct else "")


def _format_gender_value(male_present: bool, female_present: bool, male_pct: str = "", female_pct: str = "") -> str:
    if male_pct and female_pct:
        return f"{male_pct} Male + {female_pct} Female"
    if male_pct:
        return f"{male_pct} Male"
    if female_pct:
        return f"{female_pct} Female"
    if male_present and female_present:
        return "Male + Female"
    if male_present:
        return "Male"
    if female_present:
        return "Female"
    return ""


def _normalize_gender_value(raw_row: Dict[str, Any], source_text: str) -> str:
    raw_gender = _clean_text(raw_row.get("gender", ""))
    if raw_gender:
        low_gender = raw_gender.lower()
        if low_gender not in {"unknown", "n/a", "na", "not available", "not provided"}:
            male_pct, female_pct = _infer_gender_percentages(raw_gender)
            male_present = bool(re.search(r"\b(male|males|men|boys)\b", low_gender))
            female_present = bool(re.search(r"\b(female|females|women|girls)\b", low_gender))
            normalized = _format_gender_value(male_present, female_present, male_pct, female_pct)
            if normalized:
                return normalized
            return _clip_text(raw_gender, 80)

    raw_male = _clean_text(raw_row.get("gender_male", "")).lower()
    raw_female = _clean_text(raw_row.get("gender_female", "")).lower()
    male_pct = _normalize_percent(str(raw_row.get("gender_male_pct", "")))
    female_pct = _normalize_percent(str(raw_row.get("gender_female_pct", "")))
    inferred_male_pct, inferred_female_pct = _infer_gender_percentages(source_text)
    if not male_pct:
        male_pct = inferred_male_pct
    if not female_pct:
        female_pct = inferred_female_pct
    low_source = source_text.lower()
    male_present = raw_male in {"male", "males"} or bool(re.search(r"\b(male|males|men|boys)\b", low_source))
    female_present = raw_female in {"female", "females"} or bool(re.search(r"\b(female|females|women|girls)\b", low_source))
    if not female_present and re.search(r"\b(mothers?|pregnant|pregnancy|postpartum|postnatal|perinatal)\b", low_source):
        female_present = True
    return _format_gender_value(male_present, female_present, male_pct, female_pct)


def _split_candidate_fragments(*texts: str) -> List[str]:
    fragments: List[str] = []
    for text in texts:
        cleaned = _clean_text(text)
        if not cleaned:
            continue
        for fragment in re.split(r"\s*\|\|\s*|(?<=[.;])\s+(?=[A-Z0-9])", cleaned):
            value = _clean_text(fragment)
            if value:
                fragments.append(value)
    return _dedupe_keep_order(fragments)


def _has_population_signal(text: str) -> bool:
    low = _clean_text(text).lower()
    if not low:
        return False
    return bool(
        re.search(AGE_SIGNAL_PATTERN, low, flags=re.I)
        or re.search(GENDER_SIGNAL_PATTERN, low, flags=re.I)
        or re.search(OCCUPATION_SIGNAL_PATTERN, low, flags=re.I)
        or re.search(SOCIAL_STATUS_SIGNAL_PATTERN, low, flags=re.I)
        or re.search(TREATMENT_HISTORY_SIGNAL_PATTERN, low, flags=re.I)
        or any(re.search(pattern, low, flags=re.I) for pattern, _ in ETHNICITY_PATTERNS)
    )


def _strip_leading_sample_count(text: str) -> str:
    cleaned = _clean_text(text)
    if not cleaned:
        return ""
    cleaned = re.sub(r"^\(?N\s*=\s*\d{1,5}(?:,\d{3})*\)?\s*[:,-]?\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(
        r"^\d{1,3}(?:,\d{3})*\s+(?:participants?|patients?|subjects?|individuals?|adults?|women|men|children|adolescents|outpatients?|inpatients?|completers?)\b\s*[:,-]?\s*",
        "",
        cleaned,
        flags=re.I,
    )
    return _clean_text(cleaned)


def _clean_population_fragment(fragment: str) -> str:
    text = _strip_leading_sample_count(fragment)
    if not text:
        return ""
    text = re.sub(rf"\b{SCALE_NAME_PATTERN}\b[^|;,.]{{0,80}}", " ", text, flags=re.I)
    text = re.sub(r"\b\d{1,3}(?:,\d{3})*\s+at risk(?:\s+for)?\b[^|;,.]{0,90}", " ", text, flags=re.I)
    text = re.sub(rf"{RISK_SIGNAL_PATTERN}[^|;,.]{{0,90}}", " ", text, flags=re.I)
    text = re.sub(r"\b\d{1,3}(?:,\d{3})*\s+completers?\b", " ", text, flags=re.I)
    text = re.sub(r"\bcompleters?\b", " ", text, flags=re.I)
    text = re.sub(
        r"\b(?:diagnosed with|meeting criteria for|with)\s+(?P<prefix>(?:postpartum|perinatal|pregnant)\s+)?(?:major depressive disorder|mdd|depression|depressive symptoms?|unipolar depressive disorder)\b",
        lambda match: match.group("prefix") or " ",
        text,
        flags=re.I,
    )
    text = re.sub(r"\b(?:major depressive disorder|mdd|depression|depressive symptoms?|unipolar depressive disorder)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(?:patients?|participants?|subjects?|individuals?|outpatients?|inpatients?)\b", " ", text, flags=re.I)
    text = re.sub(r"\(\s*\)", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"^(?:with|and|or|among|for)\b\s*", "", text, flags=re.I)
    text = re.sub(r"\b(?:with|and|or|who|having)\b\s*$", "", text, flags=re.I)
    return _clean_text(text)


def _looks_useful_population_fragment(text: str) -> bool:
    cleaned = _clean_text(text)
    if not cleaned:
        return False
    low = cleaned.lower()
    if re.fullmatch(r"(?:adults?|women|men|children|adolescents?|older adults?|young adults?)", low):
        return True
    if not _has_population_signal(cleaned):
        return False
    if re.search(SAMPLE_COUNT_PATTERN, cleaned, flags=re.I):
        return False
    if re.search(rf"\b{SCALE_NAME_PATTERN}\b", cleaned, flags=re.I):
        return False
    if re.search(RISK_SIGNAL_PATTERN, cleaned, flags=re.I):
        return False
    if len(cleaned) > 120 and re.search(STUDY_NOISE_PATTERN, cleaned, flags=re.I):
        return False
    if low in {"depression", "major depressive disorder", "mdd"}:
        return False
    return True


def _extract_population_signal_snippets(text: str) -> List[str]:
    cleaned = _clean_text(text)
    if not cleaned:
        return []
    patterns = [
        r"\bage(?:d|s)?\s+(?:between\s+)?\d+(?:\.\d+)?(?:\s*(?:-|to|and)\s*\d+(?:\.\d+)?)?\s*years?\b(?:\s+or older)?",
        r"\baged?\s+\d+(?:\.\d+)?\+",
        r"\baged?\s*(?:>=|=>|<=|=<|>|<)\s*\d+(?:\.\d+)?(?:\s*years?)?\b",
        r"\bage\s*(?:>=|=>|<=|=<|>|<)\s*\d+(?:\.\d+)?(?:\s*years?)?\b",
        r"\bat least\s+\d+(?:\.\d+)?\s*years?\s+old\b",
        r"\baged?\s+\d+(?:\.\d+)?\s+(?:or|and)\s+(?:older|over|above|more)\b",
        r"\b\d+(?:\.\d+)?\s*years?\s+(?:or|and)\s+(?:older|over|above|more)\b",
        r"\b\d+(?:\.\d+)?\+",
        r"\b\d+(?:\.\d+)?\s*years?\s*(?:s\.?d\.?|sd)\s*[=:]?\s*\d+(?:\.\d+)?\b",
        r"\bmean age[^\d]{0,12}\d+(?:\.\d+)?(?:\s*(?:\+/-|±)\s*\d+(?:\.\d+)?)?\b",
        r"\bmedian age[^\d]{0,12}\d+(?:\.\d+)?\b",
        r"\b\d+(?:\.\d+)?\s*(?:-|to)\s*\d+(?:\.\d+)?\s*years?\b",
        r"\b\d+(?:\.\d+)?\s*-\s*to\s*\d+(?:\.\d+)?-year-olds?\b",
        r"\b\d+(?:\.\d+)?\s*years?\s+or older\b",
        r"\b\d{1,2}(?:st|nd|rd|th)?\s+grade students?\b",
        r"\bcollege students\b|\buniversity students\b|\bundergraduates?\b|\bhigh school students\b",
        r"\bmothers?\s+of\s+(?:young\s+|preschool[- ]aged\s+)?children\b|\bparents?\s+of\s+(?:young\s+|preschool[- ]aged\s+)?children\b|\bmothers?\s+of\s+adolescents?\b|\bparents?\s+of\s+adolescents?\b",
        r"\b\d+(?:\.\d+)?%\s*(?:female|male|women|men)\b",
        GENDER_SIGNAL_PATTERN,
        r"\b\d+(?:\.\d+)?%\s*(?:african[- ]american|black|white|caucasian|asian|hispanic|latino|latina|indigenous|native american|mixed|multiracial)\b",
        *[pattern for pattern, _ in ETHNICITY_PATTERNS],
        OCCUPATION_SIGNAL_PATTERN,
        SOCIAL_STATUS_SIGNAL_PATTERN,
        TREATMENT_HISTORY_SIGNAL_PATTERN,
        r"\bolder adults?\b|\byoung adults?\b|\badolescents?\b|\bchildren\b|\bschoolchildren\b|\belderly\b|\bgeriatric\b",
    ]
    hits: List[str] = []
    provider_context = re.compile(
        r"\b(?:provided|delivered|administered|led|facilitated|conducted|trained)\b.{0,120}\bby\b",
        flags=re.I,
    )
    for pattern in patterns:
        for match in re.finditer(pattern, cleaned, flags=re.I):
            if pattern == OCCUPATION_SIGNAL_PATTERN:
                window = cleaned[max(0, match.start() - 100) : min(len(cleaned), match.end() + 30)]
                if provider_context.search(window):
                    continue
            hit = _clean_text(match.group(0))
            if hit:
                hits.append(hit)
    return _dedupe_keep_order(hits)


def _build_population_focus(raw_text: str) -> str:
    fragments: List[str] = []
    for fragment in _split_candidate_fragments(raw_text):
        cleaned = _clean_population_fragment(fragment)
        fragment_has_noise = bool(
            re.search(SAMPLE_COUNT_PATTERN, fragment, flags=re.I)
            or re.search(rf"\b{SCALE_NAME_PATTERN}\b", fragment, flags=re.I)
            or re.search(RISK_SIGNAL_PATTERN, fragment, flags=re.I)
            or re.search(STUDY_NOISE_PATTERN, fragment, flags=re.I)
        )
        if _looks_useful_population_fragment(cleaned) and not fragment_has_noise:
            fragments.append(_clip_text(cleaned, 180))
            continue
        for hit in _extract_population_signal_snippets(fragment):
            cleaned_hit = _clean_population_fragment(hit)
            if _looks_useful_population_fragment(cleaned_hit):
                fragments.append(_clip_text(cleaned_hit, 180))
    return " || ".join(_dedupe_keep_order(fragments[:4]))


def _normalize_age(raw_value: str, source_text: str) -> str:
    text = _clean_text(raw_value or source_text)
    if not text:
        return ""
    low = text.lower().replace("±", "+/-").replace("–", "-").replace("—", "-")
    range_match = re.search(r"\bfrom\s+(\d+(?:\.\d+)?)\s+to\s+(\d+(?:\.\d+)?)\b", low)
    if not range_match:
        range_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)", low)
    if range_match:
        return f"{range_match.group(1)}-{range_match.group(2)}"
    pm_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:\+/-|±)\s*(\d+(?:\.\d+)?)", low)
    if not pm_match:
        pm_match = re.search(r"\bmean\b[^\d]{0,12}(\d+(?:\.\d+)?).*?\bsd\b[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if pm_match:
        return f"{pm_match.group(1)}+/-{pm_match.group(2)}"
    median_match = re.search(r"\bmedian(?: age)?[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if median_match:
        return f"Median {median_match.group(1)}"
    number_match = re.fullmatch(r"\d+(?:\.\d+)?", low)
    if number_match:
        return number_match.group(0)
    return _clip_text(text, 80)


def _normalize_ethnicity(raw_value: str, source_text: str) -> str:
    text = _clean_text(raw_value or source_text)
    if not text:
        return ""
    low = text.lower()
    hits: List[str] = []
    for pattern, label in ETHNICITY_PATTERNS:
        for percent_pattern in [
            rf"(\d{{1,3}}(?:\.\d+)?)\s*%\s*(?:{pattern})",
            rf"(?:{pattern})\s*[:=]?\s*(\d{{1,3}}(?:\.\d+)?)\s*%",
        ]:
            for match in re.finditer(percent_pattern, low, flags=re.I):
                hits.append(f"{match.group(1)}% {label}")
    if hits:
        ordered: List[str] = []
        seen: set[str] = set()
        for hit in hits:
            if hit.lower() in seen:
                continue
            seen.add(hit.lower())
            ordered.append(hit)
        return " + ".join(ordered)
    labels: List[str] = []
    for pattern, label in ETHNICITY_PATTERNS:
        if re.search(pattern, low, flags=re.I) and label.lower() not in {item.lower() for item in labels}:
            labels.append(label)
    return " + ".join(labels)


def _normalize_simple_field(raw_value: str) -> str:
    text = _clip_text(raw_value, 100)
    if text.lower() in {"", "not provided", "unknown", "n/a", "na"}:
        return ""
    return text


def _split_field_labels(value: str) -> List[str]:
    text = _clean_text(value)
    if not text:
        return []
    return [part for part in (_clean_text(piece) for piece in re.split(r"\s*(?:\|\||;|,|\+)\s*", text)) if part]


def _merge_labels(*values: str, limit: int = 100) -> str:
    merged: List[str] = []
    for value in values:
        for label in _split_field_labels(value):
            low = label.lower()
            skip = False
            for idx, existing in enumerate(list(merged)):
                existing_low = existing.lower()
                if low == existing_low or low in existing_low:
                    skip = True
                    break
                if existing_low in low:
                    merged[idx] = label
                    skip = True
                    break
            if not skip:
                merged.append(label)
    if not merged:
        return ""
    return _clip_text(" + ".join(_dedupe_keep_order(merged)), limit)


def _extract_pattern_labels(text: str, patterns: Sequence[Tuple[str, Any]], *, first_only: bool = False) -> List[str]:
    cleaned = _clean_text(text)
    if not cleaned:
        return []
    hits: List[str] = []
    for pattern, label in patterns:
        for match in re.finditer(pattern, cleaned, flags=re.I):
            resolved = _clean_text(label(match) if callable(label) else label)
            if not resolved:
                continue
            hits.append(resolved)
            if first_only:
                return _dedupe_keep_order(hits)
    return _dedupe_keep_order(hits)


def _extract_occupation_fallback(text: str) -> str:
    cleaned = _clean_text(text)
    if not cleaned:
        return ""
    patterns: List[Tuple[str, Any]] = [
        (r"\b\d{1,2}(?:st|nd|rd|th)\s+grade students?\b", lambda match: match.group(0).title()),
        (r"\bcollege students\b|\buniversity students\b|\bundergraduates?\b", "College students"),
        (r"\bhigh school students\b", "High school students"),
        (r"\bschoolchildren\b", "Schoolchildren"),
        (r"\bstudents?\b", "Students"),
        (r"\bhealth\s*care workers?\b|\bhealthcare workers?\b", "Healthcare workers"),
        (r"\bphysicians?\b|\bdoctors?\b", "Physicians"),
        (r"\bnurses?\b", "Nurses"),
        (r"\bteachers?\b", "Teachers"),
        (r"\bemployees?\b", "Employees"),
        (r"\bworkers?\b", "Workers"),
        (r"\bclinicians?\b", "Clinicians"),
    ]
    provider_context = re.compile(
        r"\b(?:provided|delivered|administered|led|facilitated|conducted|trained)\b.{0,120}\bby\b",
        flags=re.I,
    )
    for pattern, label in patterns:
        for match in re.finditer(pattern, cleaned, flags=re.I):
            window = cleaned[max(0, match.start() - 100) : min(len(cleaned), match.end() + 30)]
            if provider_context.search(window):
                continue
            resolved = _clean_text(label(match) if callable(label) else label)
            if resolved:
                return resolved
    return ""


def _extract_social_status_fallback(text: str) -> str:
    patterns: List[Tuple[str, Any]] = [
        (r"\bmothers?\s+of\s+preterms?\b", "Mothers of preterm infants"),
        (r"\bmothers?\s+of\s+preterm\s+infants?\b", "Mothers of preterm infants"),
        (r"\bmothers?\s+of\s+infants?\b", "Mothers of infants"),
        (r"\bmothers?\s+of\s+preschool[- ]aged\s+children\b", "Mothers of preschool-aged children"),
        (r"\bmothers?\s+of\s+young\s+children\b", "Mothers of young children"),
        (r"\bmothers?\s+of\s+children\b", "Mothers of children"),
        (r"\bmothers?\s+of\s+adolescents?\b", "Mothers of adolescents"),
        (r"\bspouses?\s+or\s+parents?\s+of\s+deceased\b", "Bereaved"),
        (r"\bspouses?\s+of\s+deceased\b|\bparents?\s+of\s+deceased\b", "Bereaved"),
        (r"\bparents?\s+of\s+preterm\s+infants?\b", "Parents of preterm infants"),
        (r"\bparents?\s+of\s+infants?\b", "Parents of infants"),
        (r"\bparents?\s+of\s+preschool[- ]aged\s+children\b", "Parents of preschool-aged children"),
        (r"\bparents?\s+of\s+young\s+children\b", "Parents of young children"),
        (r"\bparents?\s+of\s+children\b", "Parents of children"),
        (r"\bparents?\s+of\s+adolescents?\b", "Parents of adolescents"),
        (r"\b(?:living with\s+)?hiv(?:/aids)?\b|\bplwha\b|\bhiv[- ]positive\b|\bhiv[- ]infected\b", "HIV-positive"),
        (r"\bpregnant\b|\bpregnancy\b", "Pregnant"),
        (r"\bpostpartum\b", "Postpartum"),
        (r"\bpost[- ]birth\b", "Postpartum"),
        (r"\bperinatal\b", "Perinatal"),
        (r"\bperi[- ]menopausal\b|\bpost[- ]menopausal\b|\bmenopausal\b", "Menopausal"),
        (r"\bcaregivers?\b", "Caregivers"),
        (r"\bnew mothers?\b", "New mothers"),
        (r"\bveterans?\b", "Veterans"),
        (r"\bactive duty\b", "Active duty military"),
        (r"\brural\b", "Rural"),
        (r"\burban\b", "Urban"),
        (r"\blow[- ]income\b|\bpoverty\b|\bpoor\b|\beconomically disadvantaged\b", "Low-income"),
        (r"\bhomeless\b", "Homeless"),
        (r"\bincarcerated\b", "Incarcerated"),
        (r"\bimmigrants?\b|\bmigrants?\b|\brefugees?\b", "Immigrant or refugee status"),
        (r"\bmuslim\b", "Muslim"),
        (r"\bbereaved\b", "Bereaved"),
        (r"\bdivorcees?\b", "Divorced"),
        (r"\borphaned\b", "Orphaned"),
        (r"\bhelp[- ]seeking\b", "Help-seeking"),
        (r"\binjection drug use\b", "History of injection drug use"),
        (r"\balcohol(?: use| misuse| abuse| dependence)?\b", "Alcohol use history"),
        (r"\bsubstance use(?: disorder)?\b", "Substance use history"),
        (r"\bsmokers?\b|\bsmoking\b", "Smoking history"),
        (r"\bmethamphetamine(?: use| users?)?\b", "Methamphetamine use history"),
    ]
    return _merge_labels(*_extract_pattern_labels(text, patterns), limit=120)


def _extract_treatment_history_fallback(text: str) -> str:
    patterns: List[Tuple[str, Any]] = [
        (r"\bpreviously untreated\b", "Previously untreated"),
        (r"\bcurrently untreated\b|\bnone currently receiving treatment\b|\bnot currently receiving treatment\b", "No current treatment"),
        (r"\buntreated\b", "Untreated"),
        (r"\bunmedicated\b|\bmedication[- ]free\b|\bdrug[- ]free\b|\bpsychotropic medication[- ]free\b", "Medication-free"),
        (r"\bno prior antidepressant treatment\b|\bantidepressant[- ]naive\b|\btreatment[- ]naive\b|\bno prior treatment\b", "Treatment-naive"),
        (r"\bprior antidepressant treatment\b|\bprior treatment\b|\bhistory of treatment\b", "Prior treatment"),
        (r"\btreatment[- ]free\b|\bantidepressant[- ]free\b|\boff medication\b|\bhitherto untreated\b", "Off treatment"),
    ]
    return _merge_labels(*_extract_pattern_labels(text, patterns), limit=120)


def _legacy_normalize_age_value(raw_value: str, source_text: str) -> str:
    text = _clean_text(raw_value or source_text)
    if not text:
        return ""
    low = (
        text.lower()
        .replace("±", "+/-")
        .replace("卤", "+/-")
        .replace("–", "-")
        .replace("—", "-")
    )
    if re.search(r"\bn\s*=\s*\d+\b", low):
        return ""
    if re.fullmatch(r"\d{2,4}", low):
        try:
            numeric = float(low)
        except Exception:
            numeric = -1
        if numeric > 120:
            return ""
    aged_range_match = re.search(
        r"\baged?\s+(?:between\s+)?(\d+(?:\.\d+)?)\s*(?:-|to|and)\s*(\d+(?:\.\d+)?)\s*years?\b",
        low,
    )
    if aged_range_match:
        return f"{aged_range_match.group(1)}-{aged_range_match.group(2)}"
    range_match = re.search(r"\bfrom\s+(\d+(?:\.\d+)?)\s+to\s+(\d+(?:\.\d+)?)\b", low)
    if not range_match:
        range_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)", low)
    if range_match:
        return f"{range_match.group(1)}-{range_match.group(2)}"
    pm_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:\+/-|卤)\s*(\d+(?:\.\d+)?)", low)
    if not pm_match:
        pm_match = re.search(r"\bmean\b[^\d]{0,12}(\d+(?:\.\d+)?).*?\bsd\b[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if pm_match:
        return f"{pm_match.group(1)}+/-{pm_match.group(2)}"
    median_match = re.search(r"\bmedian(?: age)?[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if median_match:
        return f"Median {median_match.group(1)}"
    if re.search(r"\bolder adults?\b|\belderly\b|\bseniors?\b", low):
        return "Older adults"
    if re.search(r"\byoung adults?\b", low):
        return "Young adults"
    if re.search(r"\badolescents?\b|\bteen(?:ager)?s?\b|\byouth\b", low):
        return "Adolescents"
    if re.search(r"\bchildren\b|\bchild\b|\bpediatric\b", low):
        return "Children"
    if re.search(r"\binfants?\b|\bnewborns?\b|\bneonates?\b", low):
        return "Infants"
    grade_match = re.search(r"\b(\d{1,2}(?:st|nd|rd|th)?\s+grade students?)\b", low)
    if grade_match:
        return _clip_text(grade_match.group(1).title(), 40)
    if re.search(r"\bcollege students\b|\buniversity students\b|\bundergraduates?\b", low):
        return "College students"
    if re.search(r"\bhigh school students\b", low):
        return "High school students"
    if re.search(r"\bstudents?\b", low):
        return "Students"
    if re.search(r"\badults?\b", low):
        return "Adults"
    number_match = re.fullmatch(r"\d+(?:\.\d+)?", low)
    if number_match:
        try:
            numeric = float(number_match.group(0))
        except Exception:
            numeric = -1
        if 0 < numeric <= 120:
            return number_match.group(0)
    return ""


def _looks_like_sample_size_only(text: str) -> bool:
    low = _clean_text(text).lower()
    if not low:
        return False
    return bool(
        re.fullmatch(r"(?:n\s*=\s*)?\d{1,5}", low)
        or re.fullmatch(r"\d{1,5}\s+(?:participants?|patients?|subjects?|individuals?)", low)
    )


def _normalize_age_text(text: str) -> str:
    return (
        _clean_text(text)
        .lower()
        .replace("卤", "+/-")
        .replace("±", "+/-")
        .replace("+-", "+/-")
        .replace("–", "-")
        .replace("—", "-")
    )


def _is_relational_numeric_age_match(low: str, match: Any) -> bool:
    window = low[max(0, match.start() - 90) : min(len(low), match.end() + 20)]
    owner_pattern = (
        r"(?:mothers?|fathers?|parents?|caregivers?|guardians?|famil(?:y|ies) members?|siblings?|"
        r"relatives?|spouses?|partners?|teachers?|providers?|staff|workers?|employees?|persons?|people|"
        r"women|men|adults?)"
    )
    dependent_pattern = (
        r"(?:young\s+|preterm\s+|preschool[- ]aged\s+|school[- ]aged\s+|college\s+|university\s+|high school\s+)?"
        r"(?:children|child|adolescents?|teen(?:ager)?s?|youth|infants?|newborns?|neonates?|schoolchildren|students?)"
    )
    blocked_patterns = [
        rf"\b{owner_pattern}\s+of\s+{dependent_pattern}\s+aged?\b",
        rf"\b{owner_pattern}\s+with\s+{dependent_pattern}\s+aged?\b",
        rf"\b{owner_pattern}\s+for\s+{dependent_pattern}\s+aged?\b",
    ]
    return any(re.search(pattern, window, flags=re.I) for pattern in blocked_patterns)


def _range_has_age_context(low: str, match: Any) -> bool:
    window = low[max(0, match.start() - 24) : min(len(low), match.end() + 24)]
    if re.search(
        r"\b(?:weeks?|months?|days?|hours?|minutes?|sessions?|visits?|follow[- ]up|"
        r"postpartum|postnatal|gestation(?:al)?|postmenstrual|prenatal|trimester)\b",
        window,
        flags=re.I,
    ):
        return False
    if _is_relational_numeric_age_match(low, match):
        return False
    if re.search(
        r"\b(?:age|aged?|years?\s+old|year[- ]olds?|adults?|children|adolescents?|teen(?:ager)?s?|youth|"
        r"elderly|geriatric|older adults?|young adults?|schoolchildren|high school students|grade students|years?)\b",
        window,
        flags=re.I,
    ):
        return True
    return _clean_text(low) == _clean_text(match.group(0))


def _is_relational_age_match(low: str, match: Any) -> bool:
    cue = re.escape(match.group(0).strip().lower())
    window = low[max(0, match.start() - 80) : match.end()]
    owner_pattern = (
        r"(?:mothers?|fathers?|parents?|caregivers?|guardians?|famil(?:y|ies) members?|siblings?|"
        r"relatives?|spouses?|partners?|teachers?|providers?|staff|workers?|employees?|persons?|people|"
        r"women|men|adults?)"
    )
    blocked_patterns = [
        rf"\b{owner_pattern}\s+of\s+(?:young\s+|preterm\s+|preschool[- ]aged\s+|school[- ]aged\s+|college\s+|university\s+|high school\s+)?{cue}\b",
        rf"\b{owner_pattern}\s+with\s+(?:young\s+|preterm\s+|preschool[- ]aged\s+|school[- ]aged\s+)?{cue}\b",
        rf"\b{owner_pattern}\s+for\s+(?:young\s+|preterm\s+|preschool[- ]aged\s+|school[- ]aged\s+)?{cue}\b",
    ]
    return any(re.search(pattern, window, flags=re.I) for pattern in blocked_patterns)


def _first_unblocked_age_label(low: str) -> str:
    age_patterns: List[Tuple[str, Optional[str]]] = [
        (r"\bolder adults?\b|\belderly\b|\bseniors?\b|\bgeriatric\b|\blate-life\b", "Older adults"),
        (r"\byoung adults?\b", "Young adults"),
        (r"\badolescents?\b|\bteen(?:ager)?s?\b|\byouth\b", "Adolescents"),
        (r"\bschoolchildren\b", "Children"),
        (r"\bchildren\b|\bchild\b|\bpediatric\b", "Children"),
        (r"\binfants?\b|\bnewborns?\b|\bneonates?\b", "Infants"),
        (r"\b(\d{1,2}(?:st|nd|rd|th)?\s+grade students?)\b", None),
        (r"\bhigh school students\b", "High school students"),
        (r"\badults?\b", "Adults"),
    ]
    for pattern, label in age_patterns:
        for match in re.finditer(pattern, low, flags=re.I):
            if _is_relational_age_match(low, match):
                continue
            if label is None:
                return _clip_text(match.group(1).title(), 40)
            return label
    return ""


def _normalize_age_value(raw_value: str, source_text: str) -> str:
    text = _clean_text(raw_value or source_text)
    if not text:
        return ""
    low = _normalize_age_text(text)
    if _looks_like_sample_size_only(low):
        return ""
    if re.fullmatch(r"\d{2,4}", low):
        try:
            numeric = float(low)
        except Exception:
            numeric = -1
        if numeric > 120:
            return ""
    aged_range_match = re.search(
        r"\baged?\s+(?:between\s+)?(\d+(?:\.\d+)?)\s*(?:-|to|and)\s*(\d+(?:\.\d+)?)\s*years?\b",
        low,
    )
    if aged_range_match and not _is_relational_numeric_age_match(low, aged_range_match):
        return f"{aged_range_match.group(1)}-{aged_range_match.group(2)}"
    for plus_pattern in [
        r"\b(?:aged?|age)\s*(?:>=|=>|>|at\s+least)\s*(\d+(?:\.\d+)?)\s*(?:years?)?\b",
        r"\bat least\s+(\d+(?:\.\d+)?)\s*years?\s+old\b",
        r"\b(?:aged?\s+)?(\d+(?:\.\d+)?)\+",
        r"\baged?\s+(\d+(?:\.\d+)?)\s+(?:or|and)\s+(?:older|over|above|more)\b",
        r"\b(\d+(?:\.\d+)?)\s*years?\s+(?:or|and)\s+(?:older|over|above|more)\b",
    ]:
        plus_match = re.search(plus_pattern, low, flags=re.I)
        if plus_match and not _is_relational_numeric_age_match(low, plus_match):
            return f"{plus_match.group(1)}+"
    year_old_range_match = re.search(
        r"\b(\d+(?:\.\d+)?)\s*-\s*to\s*(\d+(?:\.\d+)?)-year-olds?\b",
        low,
        flags=re.I,
    )
    if year_old_range_match:
        return f"{year_old_range_match.group(1)}-{year_old_range_match.group(2)}"
    pm_match = re.search(r"\b(\d+(?:\.\d+)?)\s*(?:\+/-)\s*(\d+(?:\.\d+)?)", low)
    if not pm_match:
        pm_match = re.search(r"\bmean\b[^\d]{0,12}(\d+(?:\.\d+)?).*?\bsd\b[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if not pm_match:
        pm_match = re.search(r"\b(\d+(?:\.\d+)?)\s*years?\s*(?:,?\s*)?(?:s\.?d\.?|sd)\s*[=:]?\s*(\d+(?:\.\d+)?)\b", low)
    if pm_match:
        return f"{pm_match.group(1)}+/-{pm_match.group(2)}"
    mean_age_match = re.search(r"\bmean age(?: of)?[^\d]{0,12}(\d+(?:\.\d+)?)\b", low)
    if mean_age_match:
        return mean_age_match.group(1)
    median_match = re.search(r"\bmedian(?: age)?[^\d]{0,12}(\d+(?:\.\d+)?)", low)
    if median_match:
        return f"Median {median_match.group(1)}"
    range_match = re.search(r"\bfrom\s+(\d+(?:\.\d+)?)\s+to\s+(\d+(?:\.\d+)?)\b", low)
    if range_match and _range_has_age_context(low, range_match):
        return f"{range_match.group(1)}-{range_match.group(2)}"
    for range_match in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:-|to)\s*(\d+(?:\.\d+)?)", low):
        if _range_has_age_context(low, range_match):
            return f"{range_match.group(1)}-{range_match.group(2)}"
    exact_age_match = re.search(r"\bage(?:d|s)?\s+(\d+(?:\.\d+)?)\b", low)
    if exact_age_match and not _is_relational_numeric_age_match(low, exact_age_match):
        return exact_age_match.group(1)
    descriptor = _first_unblocked_age_label(low)
    if descriptor:
        return descriptor
    number_match = re.fullmatch(r"\d+(?:\.\d+)?", low)
    if number_match:
        try:
            numeric = float(number_match.group(0))
        except Exception:
            numeric = -1
        if 0 < numeric <= 120:
            return number_match.group(0)
    return ""


def _population_source_text(source_row: Step2BInputRow) -> str:
    return " || ".join(
        _dedupe_keep_order([_clean_text(source_row.population_focus), _clean_text(source_row.population_raw)])
    )


def _treatment_source_text(source_row: Step2BInputRow) -> str:
    return " || ".join(
        _dedupe_keep_order(
            [
                _clean_text(source_row.treatment_history),
                _clean_text(source_row.population_focus),
                _clean_text(source_row.population_raw),
            ]
        )
    )


def _normalize_occupation_value(raw_value: str, source_text: str) -> str:
    raw_text = _clean_text(raw_value)
    model_value = _extract_occupation_fallback(raw_text)
    if (
        not model_value
        and raw_text
        and not re.search(SOCIAL_STATUS_SIGNAL_PATTERN, raw_text, flags=re.I)
        and not re.search(TREATMENT_HISTORY_SIGNAL_PATTERN, raw_text, flags=re.I)
    ):
        model_value = _normalize_simple_field(raw_text)
    fallback = _extract_occupation_fallback(source_text)
    return _merge_labels(model_value, fallback, limit=100)


def _normalize_social_status_value(raw_value: str, source_text: str) -> str:
    raw_text = _clean_text(raw_value)
    model_value = _extract_social_status_fallback(raw_text)
    if (
        not model_value
        and raw_text
        and not re.search(OCCUPATION_SIGNAL_PATTERN, raw_text, flags=re.I)
        and not re.search(TREATMENT_HISTORY_SIGNAL_PATTERN, raw_text, flags=re.I)
    ):
        model_value = _normalize_simple_field(raw_text)
    fallback = _extract_social_status_fallback(source_text)
    return _merge_labels(model_value, fallback, limit=120)


def _normalize_treatment_history_value(raw_value: str, source_text: str) -> str:
    raw_text = _clean_text(raw_value)
    model_value = _extract_treatment_history_fallback(raw_text)
    if (
        not model_value
        and raw_text
        and not re.search(OCCUPATION_SIGNAL_PATTERN, raw_text, flags=re.I)
        and not re.search(SOCIAL_STATUS_SIGNAL_PATTERN, raw_text, flags=re.I)
    ):
        model_value = _normalize_simple_field(raw_text)
    fallback = _extract_treatment_history_fallback(source_text)
    return _merge_labels(model_value, fallback, limit=120)


def _header_index_map(sheet) -> Dict[str, int]:
    header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized = {normalize_ascii_key(value): index for index, value in enumerate(header_values, start=1)}
    resolved: Dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        found = next((normalized.get(alias) for alias in aliases if alias in normalized), None)
        if found is not None:
            resolved[key] = found
    return resolved


def load_input_rows(input_path: str) -> Tuple[List[Step2BInputRow], str]:
    workbook = load_workbook(input_path, data_only=True)
    candidates: List[Tuple[str, Dict[str, int]]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        resolved = _header_index_map(sheet)
        if all(key in resolved for key in HEADER_ALIASES.keys()):
            candidates.append((sheet_name, resolved))
    if not candidates:
        raise Step2BError("Could not find required columns: Record ID, PMID, Population Raw.")
    sheet_name, index_map = next(
        ((name, resolved) for name, resolved in candidates if name.strip().lower() == DEFAULT_SHEET_NAME.lower()),
        candidates[0],
    )
    rows: List[Step2BInputRow] = []
    sheet = workbook[sheet_name]
    for row_idx in range(2, sheet.max_row + 1):
        record_id = _clean_text(sheet.cell(row=row_idx, column=index_map["record_id"]).value)
        if not record_id:
            continue
        rows.append(
            Step2BInputRow(
                record_id=record_id,
                pmid=_clean_text(sheet.cell(row=row_idx, column=index_map["pmid"]).value),
                population_raw=_clean_text(sheet.cell(row=row_idx, column=index_map["population_raw"]).value),
                treatment_history=_clean_text(
                    sheet.cell(row=row_idx, column=index_map["treatment_history"]).value
                ) if "treatment_history" in index_map else "",
                population_focus=_build_population_focus(sheet.cell(row=row_idx, column=index_map["population_raw"]).value),
            )
        )
    return rows, sheet_name


def _rows_from_checkpoint(checkpoint_path: str) -> Tuple[Dict[str, Step2BOutputRow], Dict[str, int]]:
    rows: Dict[str, Step2BOutputRow] = {}
    usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    path = Path(checkpoint_path)
    if not path.exists():
        return rows, usage
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                payload = json.loads(line)
            except Exception:
                continue
            if payload.get("type") == "row" and isinstance(payload.get("data"), dict):
                data = payload["data"]
                record_id = _clean_text(data.get("Record ID", ""))
                if not record_id:
                    continue
                restored_gender = _clean_text(data.get("Gender", ""))
                if not restored_gender:
                    restored_gender = _format_gender_value(
                        bool(_clean_text(data.get("Gender Male", ""))),
                        bool(_clean_text(data.get("Gender Female", ""))),
                        _normalize_percent(_clean_text(data.get("Gender Male %", ""))),
                        _normalize_percent(_clean_text(data.get("Gender Female %", ""))),
                    )
                rows[record_id] = Step2BOutputRow(
                    record_id=record_id,
                    pmid=_clean_text(data.get("PMID", "")),
                    age=_clean_text(data.get("Age", "")),
                    gender=restored_gender,
                    ethnicity=_clean_text(data.get("Ethnicity", "")),
                    occupation=_clean_text(data.get("Occupation", "")),
                    social_status=_clean_text(data.get("Social Status", "")),
                    treatment_history=_clean_text(data.get("Treatment History", "")),
                )
            elif payload.get("type") == "usage" and isinstance(payload.get("data"), dict):
                usage["prompt_tokens"] += int(payload["data"].get("prompt_tokens", 0) or 0)
                usage["completion_tokens"] += int(payload["data"].get("completion_tokens", 0) or 0)
                usage["total_tokens"] += int(payload["data"].get("total_tokens", 0) or 0)
    return rows, usage


def write_output_workbook(output_path: str, inputs: Sequence[Step2BInputRow], outputs: Dict[str, Step2BOutputRow]) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = DEFAULT_SHEET_NAME
    ws.append(OUTPUT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    count = 0
    for input_row in inputs:
        output_row = outputs.get(input_row.record_id)
        if not output_row:
            continue
        payload = output_row.to_export_dict()
        ws.append([payload.get(header, "") for header in OUTPUT_HEADERS])
        count += 1
    wb.save(output_path)
    return count


def rebuild_excel_from_checkpoint(checkpoint_path: str, output_path: str, input_rows: Sequence[Step2BInputRow]) -> Tuple[int, int]:
    rows, usage = _rows_from_checkpoint(checkpoint_path)
    count = write_output_workbook(output_path, input_rows, rows) if rows else 0
    _, _, usage_entries = load_checkpoint_progress(checkpoint_path)
    return count, usage_entries


def _validate_resume_meta(checkpoint_store: Step2BCheckpointStore, *, input_path: str, sheet_name: str, provider: str, model: str) -> None:
    meta = checkpoint_store.read_meta()
    if not meta:
        raise Step2BError("No compatible checkpoint metadata was found for continue mode.")
    if str(meta.get("input_path", "")).strip() != str(Path(input_path).resolve()):
        raise Step2BError("Continue failed: input workbook does not match the checkpoint input.")
    if str(meta.get("sheet_name", "")).strip() != sheet_name:
        raise Step2BError("Continue failed: input sheet does not match the checkpoint sheet.")
    if str(meta.get("provider", "")).strip() != provider:
        raise Step2BError("Continue failed: provider does not match the checkpoint provider.")
    if str(meta.get("model", "")).strip() != model:
        raise Step2BError("Continue failed: resolved model does not match the checkpoint model.")


def _normalize_output(raw_row: Dict[str, Any], source_row: Step2BInputRow) -> Step2BOutputRow:
    source_text = _population_source_text(source_row)
    treatment_source = _treatment_source_text(source_row)
    return Step2BOutputRow(
        record_id=source_row.record_id,
        pmid=source_row.pmid,
        age=_normalize_age_value(_clean_text(raw_row.get("age", "")), source_text),
        gender=_normalize_gender_value(raw_row, source_text),
        ethnicity=_normalize_ethnicity(_clean_text(raw_row.get("ethnicity", "")), source_text),
        occupation=_normalize_occupation_value(str(raw_row.get("occupation", "")), source_text),
        social_status=_normalize_social_status_value(str(raw_row.get("social_status", "")), source_text),
        treatment_history=_normalize_treatment_history_value(
            str(raw_row.get("treatment_history", "")),
            treatment_source,
        ),
    )


def _build_rule_based_output(source_row: Step2BInputRow) -> Step2BOutputRow:
    return _normalize_output({}, source_row)


def _count_populated_fields(row: Step2BOutputRow) -> int:
    return sum(
        1
        for value in [row.age, row.gender, row.ethnicity, row.occupation, row.social_status, row.treatment_history]
        if _clean_text(value)
    )


def _looks_complex_population_text(text: str) -> bool:
    cleaned = _clean_text(text)
    if not cleaned:
        return False
    fragments = _split_candidate_fragments(cleaned)
    has_percent = bool(re.search(r"\b\d+(?:\.\d+)?\s*%", cleaned))
    has_ratio = bool(re.search(r"\b\d+\s*[:/]\s*\d+\b", cleaned))
    has_coordination = bool(re.search(r"\b(?:and|or)\b", cleaned, flags=re.I))
    return bool(
        has_percent
        or has_ratio
        or len(cleaned) > 160
        or (len(fragments) >= 2 and (len(cleaned) > 50 or has_coordination))
    )


def _needs_targeted_llm_override(raw_text: str, rule_output: Step2BOutputRow) -> bool:
    if not raw_text:
        return False
    low = _clean_text(raw_text).lower()
    if re.search(SOCIAL_STATUS_SIGNAL_PATTERN, low, flags=re.I) and not rule_output.social_status:
        return True
    if re.search(TREATMENT_HISTORY_SIGNAL_PATTERN, low, flags=re.I) and not rule_output.treatment_history:
        return True
    if re.search(
        r"\b(?:age\s*(?:>=|=>|<=|=<|>|<)\s*\d+|aged?\s*(?:>=|=>|<=|=<|>|<)\s*\d+|"
        r"at least\s+\d+\s*years?\s+old|\d+\s*-\s*to\s*\d+-year-olds?|\d+\+\b)\b",
        low,
        flags=re.I,
    ) and not rule_output.age:
        return True
    if re.search(r"\b(?:non[- ]hispanic|han chinese|chinese han|pacific islanders?)\b", low, flags=re.I):
        if not rule_output.ethnicity:
            return True
        if "hispanic/latino" in rule_output.ethnicity.lower() and re.search(r"\bnon[- ]hispanic\b", low, flags=re.I):
            return True
    return False


def _needs_llm_refinement(source_row: Step2BInputRow, rule_output: Step2BOutputRow) -> bool:
    raw_text = _clean_text(source_row.population_raw)
    source_text = _population_source_text(source_row)
    if not raw_text:
        return False
    if _needs_targeted_llm_override(raw_text, rule_output):
        return True
    populated = _count_populated_fields(rule_output)
    if populated == 0:
        return bool(source_row.population_focus) or _has_population_signal(raw_text)
    if populated >= 2:
        return False
    if not _looks_complex_population_text(source_text):
        return False
    if re.search(r"\b\d+(?:\.\d+)?\s*%", source_text) and not (rule_output.gender or rule_output.ethnicity):
        return True
    return len(_split_candidate_fragments(source_text)) >= 2


def _run_batch_adaptive(
    run_fn: Callable[[Sequence[Dict[str, Any]]], Any],
    items: Sequence[Dict[str, Any]],
    progress: Optional[Callable[[str], None]] = None,
    should_stop_fn: Optional[Callable[[], bool]] = None,
) -> List[Tuple[Any, List[Dict[str, Any]]]]:
    if should_stop_fn and should_stop_fn():
        raise UserCancelledError("Stopped by user request.")
    try:
        return [(run_fn(items), list(items))]
    except Exception as exc:
        if len(items) <= 1 or not should_split_batch(exc, len(items)):
            raise
        midpoint = max(1, len(items) // 2)
        if progress:
            progress(
                f"Batch with {len(items)} rows failed ({exc}). "
                f"Retrying as {midpoint} + {len(items) - midpoint}."
            )
        return _run_batch_adaptive(run_fn, items[:midpoint], progress, should_stop_fn) + _run_batch_adaptive(
            run_fn, items[midpoint:], progress, should_stop_fn
        )


def process_file(
    *,
    input_path: str,
    output_path: str,
    checkpoint_path: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    batch_size: int,
    prompt_template: str,
    timeout_seconds: int,
    retries: int,
    concurrency: int,
    progress: Optional[Callable[[str], None]] = None,
    on_state: Optional[Callable[[Dict[str, Any]], None]] = None,
    should_stop_fn: Optional[Callable[[], bool]] = None,
    resume_only: bool = False,
) -> Dict[str, Any]:
    input_rows, sheet_name = load_input_rows(input_path)
    if not input_rows:
        raise Step2BError("No valid input rows were found in the selected workbook.")
    client = Step2BExtractionClient(
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        prompt_template=prompt_template,
        timeout_seconds=timeout_seconds,
        retries=retries,
        should_stop=should_stop_fn,
        user_agent="Step2B/1.0 (+desktop)",
    )
    resolved_model = client.resolve_runtime_model()
    runtime = build_runtime_settings(
        provider=provider,
        model=resolved_model,
        timeout_seconds=timeout_seconds,
        retries=retries,
        concurrency=concurrency,
        progress=progress,
    )
    client.timeout_seconds = runtime.timeout_seconds
    client.retries = runtime.retries
    checkpoint_store = Step2BCheckpointStore(checkpoint_path)
    completed_ids: set[str] = set()
    outputs: Dict[str, Step2BOutputRow] = {}
    usage_totals = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    if resume_only:
        _validate_resume_meta(
            checkpoint_store,
            input_path=input_path,
            sheet_name=sheet_name,
            provider=provider,
            model=resolved_model,
        )
        completed_ids, _, _ = load_checkpoint_progress(checkpoint_path)
        outputs, usage_totals = _rows_from_checkpoint(checkpoint_path)
    else:
        checkpoint_store.reset(input_path=input_path, sheet_name=sheet_name, provider=provider, model=resolved_model)
    if progress:
        progress(f"Loaded {len(input_rows)} rows from sheet `{sheet_name}`.")
        progress(f"Using resolved model id: {resolved_model}")
        if completed_ids:
            progress(f"Resume mode: skipping {len(completed_ids)} previously completed rows.")

    pending_rows = [row for row in input_rows if row.record_id not in completed_ids]
    state = {
        "processed_rows": len(completed_ids),
        "total_rows": len(input_rows),
        "output_rows": len(outputs),
        "prompt_tokens": usage_totals["prompt_tokens"],
        "completion_tokens": usage_totals["completion_tokens"],
        "total_tokens": usage_totals["total_tokens"],
        "checkpoint_path": checkpoint_path,
        "sheet_name": sheet_name,
    }
    if on_state:
        on_state(dict(state))
    llm_rows: List[Step2BInputRow] = []
    local_rows = 0
    for index, row in enumerate(pending_rows, start=1):
        if should_stop_fn and should_stop_fn():
            raise UserCancelledError("Stopped by user request.")
        rule_output = _build_rule_based_output(row)
        if _needs_llm_refinement(row, rule_output):
            llm_rows.append(row)
            continue
        outputs[row.record_id] = rule_output
        checkpoint_store.append_row(rule_output)
        completed_ids.add(row.record_id)
        local_rows += 1
        state["processed_rows"] += 1
        state["output_rows"] = len(outputs)
        if on_state and (local_rows <= 5 or local_rows % 200 == 0 or index == len(pending_rows)):
            on_state(dict(state))
    planned_batches = [
        llm_rows[index : index + max(1, int(batch_size))]
        for index in range(0, len(llm_rows), max(1, int(batch_size)))
    ]
    if progress:
        progress(
            f"Rule-first pass completed {local_rows} rows locally; "
            f"{len(llm_rows)} rows remain for LLM refinement."
        )
    if not planned_batches:
        write_output_workbook(output_path, input_rows, outputs)
        return dict(state)

    controller = AdaptiveRateLimitController(
        max_workers=max(1, runtime.concurrency),
        label=f"Step2B `{resolved_model}`",
        progress=progress,
    )

    def run_batch(batch: Sequence[Step2BInputRow]) -> List[Any]:
        items = [
            {
                "row_id": idx + 1,
                "record_id": row.record_id,
                "pmid": row.pmid,
                "population_raw": row.population_raw,
                "population_focus": row.population_focus,
                "treatment_history": row.treatment_history,
            }
            for idx, row in enumerate(batch)
        ]
        return _run_batch_adaptive(client.run_batch, items, progress, should_stop_fn)

    pending: Dict[Any, int] = {}

    def submit(executor: ThreadPoolExecutor, batch_index: int) -> None:
        pending[executor.submit(run_batch, planned_batches[batch_index])] = batch_index
        if progress:
            progress(
                f"Submitted batch {batch_index + 1}/{len(planned_batches)} "
                f"(rows={len(planned_batches[batch_index])})."
            )

    with ThreadPoolExecutor(max_workers=max(1, runtime.concurrency)) as executor:
        next_index = 0
        while next_index < len(planned_batches) and len(pending) < controller.worker_cap():
            submit(executor, next_index)
            next_index += 1
        while pending or next_index < len(planned_batches):
            if should_stop_fn and should_stop_fn():
                for future in list(pending.keys()):
                    future.cancel()
                raise UserCancelledError("Stopped by user request.")
            if not pending:
                controller.wait_if_needed(
                    should_stop=should_stop_fn,
                    cancel_exception_cls=UserCancelledError,
                    cancel_message="Stopped by user request.",
                )
                while next_index < len(planned_batches) and len(pending) < controller.worker_cap() and controller.pause_remaining() <= 0:
                    submit(executor, next_index)
                    next_index += 1
                continue
            done, _ = wait(list(pending.keys()), timeout=0.2, return_when=FIRST_COMPLETED)
            if not done:
                controller.wait_if_needed(
                    should_stop=should_stop_fn,
                    cancel_exception_cls=UserCancelledError,
                    cancel_message="Stopped by user request.",
                )
                continue
            for future in done:
                batch_index = pending.pop(future)
                batch_rows = planned_batches[batch_index]
                try:
                    responses = future.result()
                except UserCancelledError:
                    raise
                except Exception as exc:
                    if is_rate_limit_error(exc):
                        controller.on_rate_limit_event(exc, 1, 1)
                    raise Step2BError(f"Batch {batch_index + 1} failed: {exc}") from exc
                controller.maybe_relax()
                source_by_id = {str(idx + 1): row for idx, row in enumerate(batch_rows)}
                for response, response_items in responses:
                    response_source_map = {
                        str(item["row_id"]): source_by_id[str(item["row_id"])]
                        for item in response_items
                        if str(item.get("row_id", "")) in source_by_id
                    }
                    seen_record_ids: set[str] = set()
                    completed_record_ids: List[str] = []
                    for raw_row in response.rows:
                        if not isinstance(raw_row, dict):
                            continue
                        source_row = response_source_map.get(str(raw_row.get("row_id", "")))
                        if not source_row:
                            continue
                        output_row = _normalize_output(raw_row, source_row)
                        outputs[source_row.record_id] = output_row
                        checkpoint_store.append_row(output_row)
                        seen_record_ids.add(source_row.record_id)
                        completed_record_ids.append(source_row.record_id)
                    for source_row in response_source_map.values():
                        if source_row.record_id in seen_record_ids:
                            continue
                        fallback_row = _build_rule_based_output(source_row)
                        outputs[source_row.record_id] = fallback_row
                        checkpoint_store.append_row(fallback_row)
                        completed_record_ids.append(source_row.record_id)
                    checkpoint_store.append_usage(response.usage.to_dict(), completed_record_ids)
                    state["processed_rows"] += len(set(completed_record_ids))
                    state["output_rows"] = len(outputs)
                    state["prompt_tokens"] += int(response.usage.prompt_tokens or 0)
                    state["completion_tokens"] += int(response.usage.completion_tokens or 0)
                    state["total_tokens"] += int(response.usage.total_tokens or 0)
                    if on_state:
                        on_state(dict(state))
                while next_index < len(planned_batches) and len(pending) < controller.worker_cap() and controller.pause_remaining() <= 0:
                    submit(executor, next_index)
                    next_index += 1

    write_output_workbook(output_path, input_rows, outputs)
    return dict(state)
