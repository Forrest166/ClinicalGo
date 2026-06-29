from __future__ import annotations

import json
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from batch_recovery import FailedBatchQueue, replay_failed_batches, write_failure_manifest
from common.api_resilience import is_rate_limit_error
from common.request_stability import AdaptiveRateLimitController
from common.structured_llm import UserCancelledError, build_runtime_settings
from common.text_utils import estimate_token_count
from config import OUTPUT_HEADERS, RESULT_SHEET_NAME
from llm_client import Step2AExtractionClient
from local_rules import (
    extract_follow_up_time_from_text,
    normalize_source_index_value,
    normalize_step2a_fields,
)
from models import Step2ARow
from population_rescue import PopulationRescueService
from population_rescue_client import Step2APopulationRescueClient
from population_rescue_manifest import (
    read_population_rescue_manifest,
    update_population_rescue_manifest,
    write_population_rescue_manifest,
)
from common.step2_support.extractor_core import (
    _extract_phase_from_text,
    build_prefixed_record_id,
    sanitize_record_id_source_label,
)
from common.step2_support.parsing import format_parsed_record_for_llm, iter_pubmed_records_from_file, parse_record_structured
from common.step2_support.scheduler import run_batch_adaptive


@dataclass
class SourceRecord:
    record_id: int
    raw_text: str
    llm_text: str
    parsed: Any


class Step2AError(Exception):
    pass


class Step2ACheckpointStore:
    def __init__(self, checkpoint_path: str) -> None:
        self.path = Path(checkpoint_path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def append(self, item: Dict[str, Any]) -> None:
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(item, ensure_ascii=False) + "\n")
            handle.flush()

    def reset(self, *, source_path: str, provider: str, model: str) -> None:
        if self.path.exists():
            self.path.unlink()
        self.append(
            {
                "type": "meta",
                "source_path": str(Path(source_path).resolve()),
                "provider": provider,
                "model": model,
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    def read_meta(self) -> Dict[str, Any]:
        if not self.path.exists():
            return {}
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                try:
                    payload = json.loads(line)
                except Exception:
                    continue
                if payload.get("type") == "meta":
                    return payload
        return {}

    def has_progress(self) -> bool:
        if not self.path.exists():
            return False
        with self.path.open("r", encoding="utf-8") as handle:
            for line in handle:
                try:
                    payload = json.loads(line)
                except Exception:
                    continue
                if payload.get("type") in {"row", "usage"}:
                    return True
        return False

    def append_rows(self, batch_index: int, rows: Sequence[Step2ARow]) -> None:
        for row in rows:
            self.append({"type": "row", "batch_index": batch_index, "data": row.__dict__})

    def append_usage(self, batch_index: int, usage: Dict[str, Any], completed_record_ids: Sequence[int]) -> None:
        self.append(
            {
                "type": "usage",
                "batch_index": batch_index,
                "completed_record_ids": [int(value) for value in completed_record_ids],
                "data": dict(usage),
            }
        )

    def append_no_row_resolution(self, batch_index: int, record_id: int) -> None:
        self.append(
            {
                "type": "no_row_resolution",
                "batch_index": int(batch_index),
                "record_id": int(record_id),
            }
        )


def load_checkpoint_progress(checkpoint_path: str) -> Tuple[set[int], int, int]:
    path = Path(checkpoint_path)
    if not path.exists():
        return set(), 0, 0
    completed_record_ids: set[int] = set()
    row_count = 0
    usage_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                payload = json.loads(line)
            except Exception:
                continue
            if payload.get("type") == "row":
                row_count += 1
                data = payload.get("data")
                if isinstance(data, dict):
                    source_index = normalize_source_index_value(data.get("source_index", ""))
                    if source_index.isdigit():
                        completed_record_ids.add(int(source_index))
            elif payload.get("type") == "usage":
                usage_count += 1
            elif payload.get("type") == "no_row_resolution":
                try:
                    completed_record_ids.add(int(payload.get("record_id", 0) or 0))
                except Exception:
                    continue
    return completed_record_ids, row_count, usage_count


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _clip_text(value: str, limit: int = 320) -> str:
    text = _clean_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _normalize_population_raw(value: str) -> str:
    raw = _clean_text(value).replace("\n", " || ").replace(";", " || ")
    if not raw:
        return ""
    parts: List[str] = []
    seen: set[str] = set()
    for piece in [segment.strip() for segment in raw.split("||") if segment.strip()]:
        clipped = _clip_text(piece, 220)
        key = clipped.lower()
        if key in seen:
            continue
        seen.add(key)
        parts.append(clipped)
        if len(parts) >= 3:
            break
    return " || ".join(parts)


def _record_ids_from_rows(rows: Sequence[Step2ARow]) -> List[int]:
    covered: List[int] = []
    seen: set[int] = set()
    for row in rows:
        source_index = normalize_source_index_value(getattr(row, "source_index", ""))
        if not source_index.isdigit():
            continue
        record_id = int(source_index)
        if record_id in seen:
            continue
        seen.add(record_id)
        covered.append(record_id)
    return covered


def _record_context_text(source: SourceRecord) -> str:
    parts = [_clean_text(source.parsed.title)]
    for section in (
        "OBJECTIVE",
        "OBJECTIVES",
        "BACKGROUND",
        "PATIENTS",
        "PARTICIPANTS",
        "METHODS",
        "INTERVENTIONS",
        "MAIN OUTCOMES AND MEASURES",
        "MAIN OUTCOME MEASURES",
        "RESULTS",
        "CONCLUSIONS",
        "SUMMARY",
    ):
        section_text = _clean_text(source.parsed.sections.get(section, ""))
        if section_text:
            parts.append(section_text)
    return " ".join(parts)


def _record_population_context_text(source: SourceRecord) -> str:
    parts = [_clean_text(source.parsed.title)]
    for section in ("PATIENTS", "PARTICIPANTS", "METHODS", "SUMMARY", "BACKGROUND"):
        section_text = _clean_text(source.parsed.sections.get(section, ""))
        if section_text:
            parts.append(section_text)
    return " ".join(parts)


def _record_evidence_context_text(source: SourceRecord) -> str:
    parts: List[str] = []
    for section in ("RESULTS", "CONCLUSIONS", "MAIN OUTCOMES AND MEASURES", "MAIN OUTCOME MEASURES", "SUMMARY"):
        section_text = _clean_text(source.parsed.sections.get(section, ""))
        if section_text:
            parts.append(section_text)
    if not parts:
        parts.append(_record_context_text(source))
    return " ".join(parts)


def _normalize_output_row(raw_row: Dict[str, Any], source: SourceRecord) -> Optional[Step2ARow]:
    candidate = dict(raw_row or {})
    candidate["source_index"] = normalize_source_index_value(candidate.get("source_index", ""))
    if not candidate["source_index"].isdigit():
        return None
    context = _record_context_text(source)
    normalized = normalize_step2a_fields(
        candidate,
        context_text=context,
        population_context_text=_record_population_context_text(source),
        evidence_context_text=_record_evidence_context_text(source),
    )
    intervention = _clip_text(normalized.get("intervention", "") or candidate.get("intervention", ""), 240)
    comparator = _clip_text(normalized.get("comparator", "") or candidate.get("comparator", ""), 240)
    evidence = _clip_text(normalized.get("evidence_snippet", "") or candidate.get("evidence_snippet", ""), 420)
    if not intervention and not comparator and not evidence:
        return None
    raw_population_present = bool(_clean_text(candidate.get("population_raw", "")))
    raw_severity_present = bool(_clean_text(candidate.get("severity", "")))
    normalized_population = _normalize_population_raw(normalized.get("population_raw", "") or candidate.get("population_raw", ""))
    normalized_severity = _clip_text(normalized.get("severity", ""), 180)
    normalized_treatment_history = _clip_text(normalized.get("treatment_history", ""), 180)
    population_status = ""
    if raw_population_present:
        if raw_severity_present:
            population_status = "main_pass"
        elif normalized_severity:
            population_status = "main_pass_plus_row_local_severity"
        else:
            population_status = "main_pass"
    elif normalized_population:
        if normalized_severity:
            population_status = "row_local_population_and_severity"
        else:
            population_status = "row_local_population"
    elif normalized_severity:
        population_status = "row_local_severity_only"
    return Step2ARow(
        source_index=str(source.record_id),
        pmid=_clean_text(source.parsed.pmid),
        nct_id=_clean_text(source.parsed.nct_id),
        journal=_clean_text(source.parsed.journal_line),
        year=_clean_text(source.parsed.year),
        indication=_clean_text(normalized.get("indication", "")),
        population_raw=normalized_population,
        severity=normalized_severity,
        treatment_history=normalized_treatment_history,
        population_status=population_status,
        target=_clip_text(normalized.get("target", "") or candidate.get("target", ""), 200),
        intervention=intervention,
        intervention_type=_clean_text(normalized.get("intervention_type", "") or candidate.get("intervention_type", "")),
        comparator=comparator,
        outcome_direction=_clean_text(normalized.get("outcome_direction", "")) or "Mixed or Unknown",
        phase=_clip_text(normalized.get("phase", "") or _extract_phase_from_text(context), 120),
        sample_size=_clip_text(normalized.get("sample_size", ""), 80),
        follow_up_time=_clip_text(normalized.get("follow_up_time", "") or extract_follow_up_time_from_text(context), 120),
        evidence_snippet=evidence,
    )


def _normalize_result_rows(raw_rows: Sequence[Any], source_by_id: Dict[int, SourceRecord]) -> List[Step2ARow]:
    normalized_rows: List[Step2ARow] = []
    for raw_row in raw_rows:
        if not isinstance(raw_row, dict):
            continue
        source_index = normalize_source_index_value(raw_row.get("source_index", ""))
        if not source_index.isdigit() or int(source_index) not in source_by_id:
            continue
        normalized_row = _normalize_output_row(raw_row, source_by_id[int(source_index)])
        if normalized_row is not None:
            normalized_rows.append(normalized_row)
    return normalized_rows


def _apply_batch_segments(
    *,
    batch_number: int,
    batch_records: Sequence[SourceRecord],
    segments: Sequence[Tuple[Any, List[int]]],
    checkpoint_store: Step2ACheckpointStore,
    completed_ids: set[int],
    row_covered_ids: set[int],
    all_rows: List[Step2ARow],
    state: Dict[str, Any],
    on_state: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> set[int]:
    source_by_id = {record.record_id: record for record in batch_records}
    missing_record_ids: set[int] = set()
    for raw_result, completed_record_ids in segments:
        normalized_rows = _normalize_result_rows(raw_result.rows, source_by_id)
        covered_record_ids = _record_ids_from_rows(normalized_rows)
        covered_record_id_set = set(covered_record_ids)
        all_rows.extend(normalized_rows)
        checkpoint_store.append_rows(batch_number, normalized_rows)
        checkpoint_store.append_usage(batch_number, raw_result.usage.to_dict(), covered_record_ids)
        for value in covered_record_ids:
            try:
                completed_ids.add(int(value))
                row_covered_ids.add(int(value))
            except Exception:
                continue
        for value in completed_record_ids:
            try:
                record_id = int(value)
            except Exception:
                continue
            if record_id in source_by_id and record_id not in covered_record_id_set:
                missing_record_ids.add(record_id)
        state["processed_records"] = len(completed_ids)
        state["records_with_rows"] = len(row_covered_ids)
        state["output_rows"] = len(all_rows)
        state["prompt_tokens"] += int(raw_result.usage.prompt_tokens or 0)
        state["completion_tokens"] += int(raw_result.usage.completion_tokens or 0)
        state["total_tokens"] += int(raw_result.usage.total_tokens or 0)
        if on_state:
            on_state(dict(state))
    return missing_record_ids


def _append_no_row_resolution(
    *,
    batch_number: int,
    record_id: int,
    checkpoint_store: Step2ACheckpointStore,
    completed_ids: set[int],
    confirmed_no_row_ids: set[int],
    row_covered_ids: set[int],
    state: Dict[str, Any],
    on_state: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> None:
    checkpoint_store.append_no_row_resolution(batch_number, record_id)
    completed_ids.add(int(record_id))
    confirmed_no_row_ids.add(int(record_id))
    state["processed_records"] = len(completed_ids)
    state["records_with_rows"] = len(row_covered_ids)
    state["confirmed_no_row_records"] = len(confirmed_no_row_ids)
    if on_state:
        on_state(dict(state))


def _rows_from_checkpoint(checkpoint_path: str) -> Tuple[List[Step2ARow], Dict[str, int], str]:
    rows: List[Step2ARow] = []
    usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    source_path = ""
    path = Path(checkpoint_path)
    if not path.exists():
        return rows, usage, source_path
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            try:
                payload = json.loads(line)
            except Exception:
                continue
            if payload.get("type") == "meta":
                source_path = str(payload.get("source_path", "") or "")
            elif payload.get("type") == "row" and isinstance(payload.get("data"), dict):
                data = payload["data"]
                rows.append(Step2ARow(**{key: data.get(key, "") for key in Step2ARow.__annotations__.keys()}))
            elif payload.get("type") == "usage" and isinstance(payload.get("data"), dict):
                usage["prompt_tokens"] += int(payload["data"].get("prompt_tokens", 0) or 0)
                usage["completion_tokens"] += int(payload["data"].get("completion_tokens", 0) or 0)
                usage["total_tokens"] += int(payload["data"].get("total_tokens", 0) or 0)
    return rows, usage, source_path


def _sort_key(row: Step2ARow, order: int) -> Tuple[int, int]:
    source_index = normalize_source_index_value(row.source_index)
    if source_index.isdigit():
        return int(source_index), order
    return 10**12, order


_EXPORT_TO_FIELD = {
    "Record Index": "source_index",
    "PMID": "pmid",
    "NCT ID": "nct_id",
    "Journal": "journal",
    "Year": "year",
    "Indication": "indication",
    "Population Raw": "population_raw",
    "Severity": "severity",
    "Treatment History": "treatment_history",
    "Population Status": "population_status",
    "Target": "target",
    "Intervention": "intervention",
    "Intervention Type": "intervention_type",
    "Comparator": "comparator",
    "Outcome Direction": "outcome_direction",
    "Phase": "phase",
    "Sample Size": "sample_size",
    "Follow-up Time": "follow_up_time",
    "Evidence Snippet": "evidence_snippet",
}


def write_output_workbook(output_path: str, rows: Sequence[Step2ARow], source_path: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = RESULT_SHEET_NAME
    ws.append(OUTPUT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    source_label = sanitize_record_id_source_label(source_path)
    for serial, (_pos, row) in enumerate(
        sorted(list(enumerate(rows)), key=lambda item: _sort_key(item[1], item[0])),
        start=1,
    ):
        record_id = build_prefixed_record_id("S2", source_label, serial)
        payload = row.to_export_dict(record_id)
        ws.append([payload.get(header, "") for header in OUTPUT_HEADERS])
    wb.save(output_path)
    return len(rows)


def load_output_workbook_rows(output_path: str) -> List[Step2ARow]:
    workbook_path = Path(output_path)
    if not workbook_path.exists():
        raise Step2AError(f"Output workbook was not found: {output_path}")
    wb = load_workbook(workbook_path)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows: List[Step2ARow] = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in {None, ""} for value in row_values):
            continue
        data = {field_name: "" for field_name in Step2ARow.__annotations__.keys()}
        for index, header in enumerate(headers):
            field_name = _EXPORT_TO_FIELD.get(header)
            if not field_name:
                continue
            data[field_name] = _clean_text(row_values[index])
        rows.append(Step2ARow(**data))
    return rows


def rebuild_excel_from_checkpoint(checkpoint_path: str, output_path: str) -> Tuple[int, int]:
    rows, usage, source_path = _rows_from_checkpoint(checkpoint_path)
    count = write_output_workbook(output_path, rows, source_path) if rows else 0
    _, _, usage_entries = load_checkpoint_progress(checkpoint_path)
    return count, usage_entries


def _load_source_records(source_path: str) -> List[SourceRecord]:
    records: List[SourceRecord] = []
    for record_id, raw_record in enumerate(iter_pubmed_records_from_file(source_path), start=1):
        parsed = parse_record_structured(raw_record, record_id=record_id)
        llm_text = format_parsed_record_for_llm(parsed) or _clean_text(raw_record)
        records.append(SourceRecord(record_id=record_id, raw_text=raw_record, llm_text=llm_text, parsed=parsed))
    return records


def _load_selected_source_records(source_path: str, record_ids: Sequence[int]) -> Dict[int, SourceRecord]:
    wanted = {int(value) for value in record_ids}
    if not wanted:
        return {}
    selected: Dict[int, SourceRecord] = {}
    for record in _load_source_records(source_path):
        if record.record_id in wanted:
            selected[record.record_id] = record
    return selected


def _validate_resume_meta(checkpoint_store: Step2ACheckpointStore, *, source_path: str, provider: str, model: str) -> None:
    meta = checkpoint_store.read_meta()
    if not meta:
        raise Step2AError("No compatible checkpoint metadata was found for continue mode.")
    if str(meta.get("source_path", "")).strip() != str(Path(source_path).resolve()):
        raise Step2AError("Continue failed: source TXT does not match the checkpoint source.")
    if str(meta.get("provider", "")).strip() != provider:
        raise Step2AError("Continue failed: provider does not match the checkpoint provider.")
    if str(meta.get("model", "")).strip() != model:
        raise Step2AError("Continue failed: resolved model does not match the checkpoint model.")


def process_file(
    *,
    source_path: str,
    output_path: str,
    checkpoint_path: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    batch_size: int,
    prompt_template: str,
    timeout_seconds: int,
    retries: int,
    concurrency: int,
    progress: Optional[Callable[[str], None]] = None,
    on_state: Optional[Callable[[Dict[str, Any]], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
    resume_only: bool = False,
) -> Dict[str, Any]:
    records = _load_source_records(source_path)
    if not records:
        raise Step2AError("No records were detected in the input TXT.")
    client = Step2AExtractionClient(
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        prompt_template=prompt_template,
        timeout_seconds=timeout_seconds,
        retries=retries,
        should_stop=should_stop,
        user_agent="Step2A/1.0 (+desktop)",
    )
    resolved_model = client.resolve_runtime_model()
    runtime = build_runtime_settings(
        provider=provider,
        model=resolved_model,
        timeout_seconds=timeout_seconds,
        retries=retries,
        concurrency=concurrency,
        progress=progress,
    )
    client.timeout_seconds = runtime.timeout_seconds
    client.retries = runtime.retries
    population_rescue_planner = PopulationRescueService(
        progress=progress,
        should_stop=should_stop,
        cancel_exception_cls=UserCancelledError,
    )
    checkpoint_store = Step2ACheckpointStore(checkpoint_path)
    completed_ids: set[int] = set()
    all_rows: List[Step2ARow] = []
    usage_totals = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    if resume_only:
        _validate_resume_meta(checkpoint_store, source_path=source_path, provider=provider, model=resolved_model)
        completed_ids, _, _ = load_checkpoint_progress(checkpoint_path)
        all_rows, usage_totals, _ = _rows_from_checkpoint(checkpoint_path)
    else:
        checkpoint_store.reset(source_path=source_path, provider=provider, model=resolved_model)
    row_covered_ids: set[int] = set(_record_ids_from_rows(all_rows))
    confirmed_no_row_ids: set[int] = {record_id for record_id in completed_ids if record_id not in row_covered_ids}
    record_by_id = {record.record_id: record for record in records}
    if progress:
        progress(
            f"Detected {len(records)} records. Estimated input tokens: "
            f"{sum(estimate_token_count(record.llm_text) for record in records)}."
        )
        progress(f"Using resolved model id: {resolved_model}")
        if completed_ids:
            progress(f"Resume mode: skipping {len(completed_ids)} previously completed records.")

    pending_records = [record for record in records if record.record_id not in completed_ids]
    planned_batches = [
        pending_records[index : index + max(1, int(batch_size))]
        for index in range(0, len(pending_records), max(1, int(batch_size)))
    ]
    state: Dict[str, Any] = {
        "processed_records": len(completed_ids),
        "total_records": len(records),
        "records_with_rows": len(row_covered_ids),
        "output_rows": len(all_rows),
        "prompt_tokens": usage_totals["prompt_tokens"],
        "completion_tokens": usage_totals["completion_tokens"],
        "total_tokens": usage_totals["total_tokens"],
        "checkpoint_path": checkpoint_path,
        "failed_batches": 0,
        "failed_records": 0,
        "recovered_failed_batches": 0,
        "recovered_failed_records": 0,
        "failure_manifest_path": "",
        "second_pass_candidate_records": 0,
        "second_pass_recovered_records": 0,
        "second_pass_confirmed_no_row_records": 0,
        "confirmed_no_row_records": len(confirmed_no_row_ids),
        "remaining_unresolved_records": 0,
        "population_rescue_manifest_path": "",
        "population_rescue_candidate_records": 0,
        "population_rescue_population_candidates": 0,
        "population_rescue_severity_candidates": 0,
        "population_rescue_no_cue_records": 0,
    }
    if on_state:
        on_state(dict(state))

    controller = AdaptiveRateLimitController(
        max_workers=max(1, runtime.concurrency),
        label=f"Step2A `{resolved_model}`",
        progress=progress,
    )

    def run_batch(batch: Sequence[SourceRecord]) -> List[Tuple[Any, List[int]]]:
        return run_batch_adaptive(
            client.run_batch,
            [record.llm_text for record in batch],
            [record.record_id for record in batch],
            progress=progress,
            should_stop=should_stop,
            cancel_exception_cls=UserCancelledError,
        )

    def run_recovery_batch(batch: Sequence[SourceRecord]) -> List[Tuple[Any, List[int]]]:
        controller.wait_if_needed(
            should_stop=should_stop,
            cancel_exception_cls=UserCancelledError,
            cancel_message="Stopped by user request.",
        )
        return run_batch(batch)

    synthetic_batch_number = max(0, len(planned_batches))

    def next_synthetic_batch_number() -> int:
        nonlocal synthetic_batch_number
        synthetic_batch_number += 1
        return synthetic_batch_number

    pending: Dict[Any, int] = {}
    failed_batches = FailedBatchQueue()
    coverage_gap_ids: set[int] = set()

    def submit(executor: ThreadPoolExecutor, batch_index: int) -> None:
        pending[executor.submit(run_batch, planned_batches[batch_index])] = batch_index
        if progress:
            progress(
                f"Submitted batch {batch_index + 1}/{len(planned_batches)} "
                f"(records={len(planned_batches[batch_index])})."
            )

    with ThreadPoolExecutor(max_workers=max(1, runtime.concurrency)) as executor:
        next_index = 0
        while next_index < len(planned_batches) and len(pending) < controller.worker_cap():
            submit(executor, next_index)
            next_index += 1
        while pending or next_index < len(planned_batches):
            if should_stop and should_stop():
                for future in list(pending.keys()):
                    future.cancel()
                raise UserCancelledError("Stopped by user request.")
            if not pending:
                controller.wait_if_needed(
                    should_stop=should_stop,
                    cancel_exception_cls=UserCancelledError,
                    cancel_message="Stopped by user request.",
                )
                while next_index < len(planned_batches) and len(pending) < controller.worker_cap() and controller.pause_remaining() <= 0:
                    submit(executor, next_index)
                    next_index += 1
                continue
            done, _ = wait(list(pending.keys()), timeout=0.2, return_when=FIRST_COMPLETED)
            if not done:
                controller.wait_if_needed(
                    should_stop=should_stop,
                    cancel_exception_cls=UserCancelledError,
                    cancel_message="Stopped by user request.",
                )
                continue
            for future in done:
                batch_index = pending.pop(future)
                batch_records = planned_batches[batch_index]
                try:
                    segments = future.result()
                except UserCancelledError:
                    raise
                except Exception as exc:
                    if is_rate_limit_error(exc):
                        controller.on_rate_limit_event(exc, 1, 1)
                    failed_batch = failed_batches.add(
                        batch_number=batch_index + 1,
                        batch_records=batch_records,
                        error=exc,
                    )
                    if progress:
                        progress(
                            f"Batch {failed_batch.batch_number}/{len(planned_batches)} failed during main pass "
                            f"(records={failed_batch.record_count}). Queued for replay after the main pass: {exc}"
                        )
                    state["failed_batches"] = len(failed_batches)
                    state["failed_records"] = sum(item.record_count for item in failed_batches.items())
                    if on_state:
                        on_state(dict(state))
                    while next_index < len(planned_batches) and len(pending) < controller.worker_cap() and controller.pause_remaining() <= 0:
                        submit(executor, next_index)
                        next_index += 1
                    continue
                controller.maybe_relax()
                missing_ids = _apply_batch_segments(
                    batch_number=batch_index + 1,
                    batch_records=batch_records,
                    segments=segments,
                    checkpoint_store=checkpoint_store,
                    completed_ids=completed_ids,
                    row_covered_ids=row_covered_ids,
                    all_rows=all_rows,
                    state=state,
                    on_state=on_state,
                )
                coverage_gap_ids.update(record_id for record_id in missing_ids if record_id not in completed_ids)
                while next_index < len(planned_batches) and len(pending) < controller.worker_cap() and controller.pause_remaining() <= 0:
                    submit(executor, next_index)
                    next_index += 1

    def on_replay_success(failed_batch, segments: List[Tuple[Any, List[int]]]) -> None:
        missing_ids = _apply_batch_segments(
            batch_number=failed_batch.batch_number,
            batch_records=failed_batch.batch_records,
            segments=segments,
            checkpoint_store=checkpoint_store,
            completed_ids=completed_ids,
            row_covered_ids=row_covered_ids,
            all_rows=all_rows,
            state=state,
            on_state=on_state,
        )
        coverage_gap_ids.update(record_id for record_id in missing_ids if record_id not in completed_ids)

    recovery_summary = replay_failed_batches(
        failed_batches.items(),
        run_batch_fn=run_recovery_batch,
        on_success=on_replay_success,
        on_failure=lambda _failed_batch, exc: controller.on_rate_limit_event(exc, 1, 1) if is_rate_limit_error(exc) else None,
        progress=progress,
        should_stop=should_stop,
        cancel_exception_cls=UserCancelledError,
    )
    controller.maybe_relax()
    recovery_summary.manifest_path = write_failure_manifest(output_path, recovery_summary.permanent_failures)
    state["failed_batches"] = recovery_summary.permanent_batch_count
    state["failed_records"] = recovery_summary.permanent_record_count
    state["recovered_failed_batches"] = recovery_summary.recovered_batches
    state["recovered_failed_records"] = recovery_summary.recovered_records
    state["failure_manifest_path"] = recovery_summary.manifest_path if recovery_summary.permanent_failures else ""
    if progress and recovery_summary.collected_batches:
        progress(
            f"Replay summary: recovered {recovery_summary.recovered_batches}/{recovery_summary.collected_batches} "
            f"failed batches. Remaining permanent failures: {recovery_summary.permanent_batch_count} "
            f"batches / {recovery_summary.permanent_record_count} records."
        )
        if recovery_summary.permanent_failures and state["failure_manifest_path"]:
            progress(f"Failed-batch manifest written to: {state['failure_manifest_path']}")

    second_pass_candidates = sorted(record_id for record_id in coverage_gap_ids if record_id not in completed_ids and record_id in record_by_id)
    state["second_pass_candidate_records"] = len(second_pass_candidates)
    if second_pass_candidates and progress:
        progress(
            f"Main pass finished with {len(second_pass_candidates)} records that did not land in workbook rows. "
            f"Starting second-pass record replay."
        )
    second_pass_recovered_before = len(row_covered_ids)
    confirmed_no_row_before = len(confirmed_no_row_ids)
    unresolved_second_pass_ids: set[int] = set()
    if second_pass_candidates:
        second_pass_batches = [
            second_pass_candidates[index : index + max(1, int(batch_size))]
            for index in range(0, len(second_pass_candidates), max(1, int(batch_size)))
        ]
        solo_replay_ids: set[int] = set()
        for replay_index, batch_record_ids in enumerate(second_pass_batches, start=1):
            batch_records = [record_by_id[record_id] for record_id in batch_record_ids if record_id in record_by_id]
            if not batch_records:
                continue
            if progress:
                progress(
                    f"Second-pass replay {replay_index}/{len(second_pass_batches)} "
                    f"(records={len(batch_records)})."
                )
            try:
                segments = run_recovery_batch(batch_records)
            except UserCancelledError:
                raise
            except Exception as exc:
                solo_replay_ids.update(record.record_id for record in batch_records)
                if progress:
                    progress(
                        f"Second-pass replay {replay_index}/{len(second_pass_batches)} failed "
                        f"(records={len(batch_records)}). Deferring to single-record verification: {exc}"
                    )
                continue
            missing_ids = _apply_batch_segments(
                batch_number=next_synthetic_batch_number(),
                batch_records=batch_records,
                segments=segments,
                checkpoint_store=checkpoint_store,
                completed_ids=completed_ids,
                row_covered_ids=row_covered_ids,
                all_rows=all_rows,
                state=state,
                on_state=on_state,
            )
            solo_replay_ids.update(record_id for record_id in missing_ids if record_id not in completed_ids)
        if solo_replay_ids and progress:
            progress(
                f"Second-pass replay still left {len(solo_replay_ids)} records without workbook rows. "
                f"Starting single-record verification."
            )
        for verify_index, record_id in enumerate(sorted(solo_replay_ids), start=1):
            if record_id in completed_ids:
                continue
            record = record_by_id.get(record_id)
            if record is None:
                continue
            if progress:
                progress(
                    f"Single-record verification {verify_index}/{len(solo_replay_ids)} for record {record_id}."
                )
            try:
                segments = run_recovery_batch([record])
            except UserCancelledError:
                raise
            except Exception as exc:
                unresolved_second_pass_ids.add(record_id)
                if progress:
                    progress(f"Single-record verification failed for record {record_id}: {exc}")
                continue
            _apply_batch_segments(
                batch_number=next_synthetic_batch_number(),
                batch_records=[record],
                segments=segments,
                checkpoint_store=checkpoint_store,
                completed_ids=completed_ids,
                row_covered_ids=row_covered_ids,
                all_rows=all_rows,
                state=state,
                on_state=on_state,
            )
            if record_id not in completed_ids:
                _append_no_row_resolution(
                    batch_number=next_synthetic_batch_number(),
                    record_id=record_id,
                    checkpoint_store=checkpoint_store,
                    completed_ids=completed_ids,
                    confirmed_no_row_ids=confirmed_no_row_ids,
                    row_covered_ids=row_covered_ids,
                    state=state,
                    on_state=on_state,
                )
                if progress:
                    progress(
                        f"Record {record_id} produced no extractable Step2A row even after single-record verification. "
                        f"Marked as confirmed no-row."
                    )
    state["second_pass_recovered_records"] = max(0, len(row_covered_ids) - second_pass_recovered_before)
    state["second_pass_confirmed_no_row_records"] = max(0, len(confirmed_no_row_ids) - confirmed_no_row_before)
    state["confirmed_no_row_records"] = len(confirmed_no_row_ids)
    state["remaining_unresolved_records"] = len(
        {record_id for record_id in unresolved_second_pass_ids if record_id not in completed_ids}
    )
    if second_pass_candidates and progress:
        progress(
            f"Second-pass summary: checked {len(second_pass_candidates)} records, "
            f"recovered {state['second_pass_recovered_records']} into workbook rows, "
            f"confirmed {state['second_pass_confirmed_no_row_records']} as no-row, "
            f"unresolved={state['remaining_unresolved_records']}."
        )
    if on_state:
        on_state(dict(state))

    write_output_workbook(output_path, all_rows, source_path)
    completed_source_by_id = {record.record_id: record for record in records if record.record_id in row_covered_ids}
    rescue_plan = population_rescue_planner.plan_rows(all_rows, completed_source_by_id)
    rescue_manifest_path = write_population_rescue_manifest(
        output_path=output_path,
        source_path=source_path,
        record_ids=rescue_plan.candidate_record_ids,
        metadata={
            "candidate_record_count": len(rescue_plan.candidate_record_ids),
            "population_candidate_count": rescue_plan.population_candidate_count,
            "severity_candidate_count": rescue_plan.severity_candidate_count,
            "no_cue_record_count": rescue_plan.no_cue_record_count,
            "failed_record_count": state["failed_records"],
            "generated_from_resume": bool(resume_only),
        },
    )
    state["population_rescue_manifest_path"] = rescue_manifest_path if rescue_plan.candidate_record_ids else ""
    state["population_rescue_candidate_records"] = len(rescue_plan.candidate_record_ids)
    state["population_rescue_population_candidates"] = rescue_plan.population_candidate_count
    state["population_rescue_severity_candidates"] = rescue_plan.severity_candidate_count
    state["population_rescue_no_cue_records"] = rescue_plan.no_cue_record_count
    if progress:
        if rescue_plan.candidate_record_ids:
            progress(
                f"Population rescue manifest written to: {rescue_manifest_path} "
                f"(records={len(rescue_plan.candidate_record_ids)}, "
                f"population={rescue_plan.population_candidate_count}, "
                f"severity={rescue_plan.severity_candidate_count})."
            )
        else:
            progress("Population rescue manifest not needed; no rescue candidates were detected.")
    if on_state:
        on_state(dict(state))
    return dict(state)


def run_population_rescue_stage(
    *,
    manifest_path: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    batch_size: int,
    timeout_seconds: int,
    retries: int,
    concurrency: int,
    progress: Optional[Callable[[str], None]] = None,
    on_state: Optional[Callable[[Dict[str, Any]], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    manifest = read_population_rescue_manifest(manifest_path)
    if manifest.status.lower() == "completed":
        raise Step2AError("This population rescue manifest has already been completed.")
    record_ids = [int(value) for value in manifest.record_ids]
    if not record_ids:
        raise Step2AError("Population rescue manifest does not contain any record ids.")
    rows = load_output_workbook_rows(manifest.output_path)
    target_ids = set(record_ids)
    target_rows = [
        row
        for row in rows
        if normalize_source_index_value(row.source_index).isdigit()
        and int(normalize_source_index_value(row.source_index)) in target_ids
    ]
    if not target_rows:
        raise Step2AError("No workbook rows matched the population rescue manifest record ids.")
    source_by_id = _load_selected_source_records(manifest.source_path, record_ids)
    if not source_by_id:
        raise Step2AError("Population rescue could not load the source records referenced by the manifest.")

    client = Step2APopulationRescueClient(
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        timeout_seconds=timeout_seconds,
        retries=retries,
        should_stop=should_stop,
        user_agent="Step2A-PopulationRescue/1.0 (+desktop)",
    )
    resolved_model = client.resolve_runtime_model()
    runtime = build_runtime_settings(
        provider=provider,
        model=resolved_model,
        timeout_seconds=timeout_seconds,
        retries=retries,
        concurrency=concurrency,
        progress=progress,
    )
    client.timeout_seconds = runtime.timeout_seconds
    client.retries = runtime.retries

    state: Dict[str, Any] = {
        "processed_records": 0,
        "total_records": len(source_by_id),
        "output_rows": len(target_rows),
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "checkpoint_path": manifest_path,
        "failed_batches": 0,
        "failed_records": 0,
        "recovered_failed_batches": 0,
        "recovered_failed_records": 0,
        "failure_manifest_path": "",
        "population_rescue_manifest_path": manifest_path,
        "population_rescue_candidate_records": len(source_by_id),
        "population_rescue_population_candidates": int(manifest.metadata.get("population_candidate_count", 0) or 0),
        "population_rescue_severity_candidates": int(manifest.metadata.get("severity_candidate_count", 0) or 0),
        "population_rescue_no_cue_records": int(manifest.metadata.get("no_cue_record_count", 0) or 0),
    }
    if progress:
        progress(
            f"Starting population rescue for {len(source_by_id)} records using resolved model `{resolved_model}`."
        )
    if on_state:
        on_state(dict(state))

    def dispatch_partial_state(partial: Dict[str, Any]) -> None:
        if not partial:
            return
        state.update(partial)
        if on_state:
            on_state(dict(state))

    rescue_service = PopulationRescueService(
        client=client,
        progress=progress,
        on_state=dispatch_partial_state,
        should_stop=should_stop,
        cancel_exception_cls=UserCancelledError,
        max_batch_records=max(1, int(batch_size)),
        max_workers=max(1, int(runtime.concurrency)),
    )

    rescue_summary = rescue_service.enrich_rows(target_rows, source_by_id)
    state["processed_records"] = len(source_by_id)
    state["prompt_tokens"] = rescue_summary.prompt_tokens
    state["completion_tokens"] = rescue_summary.completion_tokens
    state["total_tokens"] = rescue_summary.total_tokens
    state["output_rows"] = len(rows)
    state["population_rescue_population_candidates"] = rescue_summary.llm_population_rows + rescue_summary.local_population_rows
    state["population_rescue_severity_candidates"] = rescue_summary.llm_severity_rows + rescue_summary.local_severity_rows
    state["population_rescue_no_cue_records"] = rescue_summary.no_cue_records

    write_output_workbook(manifest.output_path, rows, manifest.source_path)
    update_population_rescue_manifest(
        manifest_path,
        status="completed",
        metadata_updates={
            "completed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "provider": provider,
            "model": resolved_model,
            "summary": rescue_summary.to_dict(),
        },
    )
    if progress:
        progress(
            "Population rescue completed: "
            f"broadcast rows={rescue_summary.broadcast_rows}, "
            f"local population fills={rescue_summary.local_population_rows}, "
            f"local severity fills={rescue_summary.local_severity_rows}, "
            f"llm population fills={rescue_summary.llm_population_rows}, "
            f"llm severity fills={rescue_summary.llm_severity_rows}."
        )
    if on_state:
        on_state(dict(state))
    return dict(state)
